from asyncio.log import logger
import sys
import getpass
import os


from PyQt5 import QtWidgets
from PyQt5 import QtCore
from PyQt5 import QtGui
from PyQt5.QtWidgets import *
from PyQt5.QtCore import *
from PyQt5.QtGui import *
from tkinter import *

from openpyxl import Workbook



import project as ui
import pymysql
import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.styles import Border,Side
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from sql import *
import sql
#import auto_login

connection = pymysql.connect(host='192.168.101.64',
                        port=3306,
                        user='root',
                        passwd='jackey8869',
                        database='work_order',
                        charset='utf8',
                        cursorclass=pymysql.cursors.DictCursor)


# cursor = connection.cursor()


class Main(QMainWindow, ui.Ui_Form):
    def __init__(self):
        super().__init__()
        self.setupUi(self)


        #create_table()
        #del_data()
        #update_data()
        #insert_data()

        ###增加下拉式選單客戶
        self.comboBox_company_name.addItems(company_name())


        def index_change():
            try:
                full_name,phone,address,taxID = company_name_change(self.comboBox_company_name.currentText())
                self.lineEdit_phone.setText(phone)
            except:
                pass
        self.comboBox_company_name.currentTextChanged.connect(index_change)

        ###增加下拉式選單工作人員
        for i in range(1,6):
            exec("self.comboBox_employee_"+str(i)+".addItems(employee_name())")
                
        self.comboBox_pack.addItems(pack_name())
        self.comboBox_transport.addItems(transport_name())

        ###增加下拉式選單中間大表
        for i in range(1,16):
            exec("self.comboBox_product_"+str(i)+".addItems(product_name())")
            exec("self.comboBox_material_"+str(i)+".addItems(material_name())")
            exec("self.comboBox_process_"+str(i)+".addItems(process_name())")
            exec("self.comboBox_plate_"+str(i)+".addItems(plate_name())")
            exec("self.comboBox_thicknes_"+str(i)+".addItems(thickness_name())")
            exec("self.comboBox_others_"+str(i)+".addItems(others_name())")
        
    
        ###點擊"小記" 算出價錢   
        def cal_price(self):
            try:
                for i in range(1,16):
                    exec("width = self.lineEdit_width_"+str(i)+".text()")
                    exec("if width == '': width = '0'")
                    exec("height = self.lineEdit_height_"+str(i)+".text()")
                    exec("if height == '': height = '0'")
                    exec("amount = self.lineEdit_amount_"+str(i)+".text()")
                    exec("if amount == '': amount = '1'")
                    exec("CBM1 = math.ceil(int(width) * int(height) / 900)")
                    exec("CBM2 = int(CBM1) * int(amount)")
                    exec("if CBM2 == 0: CBM2 = ''")
                    exec("self.lineEdit_CBM_"+str(i)+".setText(str(CBM2))")
                    exec("CBM= self.lineEdit_CBM_"+str(i)+".text()")
                    exec("if CBM == '': CBM = '0'")
                    exec("CBMprice= self.lineEdit_CBMprice_"+str(i)+".text()")
                    exec("if CBMprice == '': CBMprice = '0'")
                    exec("result = int(CBM)*int(CBMprice)")
                    exec("if result == 0: result = '' ")
                    exec("self.lineEdit_single_price_"+str(i)+".setText(str(result))")
            except Exception as e:
                QMessageBox.critical(self, 'Error Information', repr(e), QMessageBox.Ok | QMessageBox.Close, QMessageBox.Close)     
        self.pushButton_cal.clicked.connect(lambda:cal_price(self))

        #self.pushButton_cal.clicked.connect(self.confirm_window.show)

        ###點擊"計算"，算出總價
        def cal_final_price(self):
            try:
                final_price = []
                for i in range(1,31):
                    exec("single_price = self.lineEdit_single_price_"+str(i)+".text()")
                    exec("if single_price == '':single_price = '0'")
                    exec("final_price.append(single_price)")               
                final_price = list(map(int,final_price))
                result = sum(final_price)
                
                tax = result*0.05
                tax = int(tax+0.5)
                answer = result + tax
                self.lineEdit_tmpprice.setText(str(result))
                self.lineEdit_tax.setText(str(tax))
                self.lineEdit_final_price.setText(str(answer))
            except Exception as e:
                QMessageBox.warning(self, 'Error Information', repr(e), QMessageBox.Ok | QMessageBox.Close, QMessageBox.Close)
        self.pushButton_finalcal.clicked.connect(lambda:cal_final_price(self))
       

       ### 直接存basic_data
        def save_basic_data():
            try:
                dict_data = {
                    'worknum' : self.lineEdit_worknum.text(),
                    'case_name' : self.lineEdit_case_name.text(),
                    'company_name' :str(self.comboBox_company_name.currentText()),
                    'phone' : self.lineEdit_phone.text(),
                    'client_name' : self.lineEdit_client_name.text(),
                    'worktime' : self.lineEdit_worktime.text(),
                    'cleanuptime' : self.lineEdit_cleanup_time.text(),
                    'workaddress' : self.lineEdit_workaddress.text(),
                    'pack' : str(self.comboBox_pack.currentIndex()),
                    'transport' : str(self.comboBox_transport.currentIndex()),
                    'cemployee1' : str(self.comboBox_employee_1.currentIndex()),
                    'cemployee2' : str(self.comboBox_employee_2.currentIndex()),
                    'cemployee3' : str(self.comboBox_employee_3.currentIndex()),
                    'cemployee4' : str(self.comboBox_employee_4.currentIndex()),
                    'cemployee5' : str(self.comboBox_employee_5.currentIndex()),
                    'crossbar_width' : self.lineEdit_crossbar_width.text(),
                    'crossbar_amount': self.lineEdit_crossbar_amount.text(),
                    'crossbar_remark': self.lineEdit_crossbar_remark.text(),
                    '150shelter':self.checkBox_150iron_shelter.isChecked(),
                    '180shelter':self.checkBox_180iron_shelter.isChecked(),
                    'iron_Shelter_amount':self.lineEdit_iron_shelter_amount.text(),
                    'iron_Shelter_remark':self.lineEdit_iron_shelter_remark.text(),
                    'paper_Shelter_height':self.lineEdit_paper_shelter_height.text(),
                    'paper_Shelter_amount':self.lineEdit_paper_shelter_amount.text(),
                    'paper_Shelter_remark':self.lineEdit_paper_shelter_remark.text(),
                    'stand_style' : self.lineEdit_stand_style.text(),
                    'stand_amount' : self.lineEdit_stand_amount.text(),
                    'stand_remark' : self.lineEdit_stand_remark.text(),
                    'rent1' : self.lineEdit_rent_1.text(),
                    'rent2' : self.lineEdit_rent_2.text(),
                    'remark' : self.textEdit_remark.toPlainText()
                }
                Update_basic_data(dict_data)
            except Exception as e:
                QMessageBox.warning(self, 'Error Information', repr(e), QMessageBox.Ok | QMessageBox.Close, QMessageBox.Close)


        ###直接存central_data
        def save_central_data():
            try:
                dict_data = {
                    'worknum':self.lineEdit_worknum.text(),
                    'product_1':str(self.comboBox_product_1.currentText()),
                    'product_2':str(self.comboBox_product_2.currentText()),
                    'product_3':str(self.comboBox_product_3.currentText()),
                    'product_4':str(self.comboBox_product_4.currentText()),
                    'product_5':str(self.comboBox_product_5.currentText()),
                    'product_6':str(self.comboBox_product_6.currentText()),
                    'product_7':str(self.comboBox_product_7.currentText()),
                    'product_8':str(self.comboBox_product_8.currentText()),
                    'product_9':str(self.comboBox_product_9.currentText()),
                    'product_10':str(self.comboBox_product_10.currentText()),
                    'product_11':str(self.comboBox_product_11.currentText()),
                    'product_12':str(self.comboBox_product_12.currentText()),
                    'product_13':str(self.comboBox_product_13.currentText()),
                    'product_14':str(self.comboBox_product_14.currentText()),
                    'product_15':str(self.comboBox_product_15.currentText()),
                    'width_1':self.lineEdit_width_1.text(),
                    'width_2':self.lineEdit_width_2.text(),
                    'width_3':self.lineEdit_width_3.text(),
                    'width_4':self.lineEdit_width_4.text(),
                    'width_5':self.lineEdit_width_5.text(),
                    'width_6':self.lineEdit_width_6.text(),
                    'width_7':self.lineEdit_width_7.text(),
                    'width_8':self.lineEdit_width_8.text(),
                    'width_9':self.lineEdit_width_9.text(),
                    'width_10':self.lineEdit_width_10.text(),
                    'width_11':self.lineEdit_width_11.text(),
                    'width_12':self.lineEdit_width_12.text(),
                    'width_13':self.lineEdit_width_13.text(),
                    'width_14':self.lineEdit_width_14.text(),
                    'width_15':self.lineEdit_width_15.text(),
                    'height_1':self.lineEdit_height_1.text(),
                    'height_2':self.lineEdit_height_2.text(),
                    'height_3':self.lineEdit_height_3.text(),
                    'height_4':self.lineEdit_height_4.text(),
                    'height_5':self.lineEdit_height_5.text(),
                    'height_6':self.lineEdit_height_6.text(),
                    'height_7':self.lineEdit_height_7.text(),
                    'height_8':self.lineEdit_height_8.text(),
                    'height_9':self.lineEdit_height_9.text(),
                    'height_10':self.lineEdit_height_10.text(),
                    'height_11':self.lineEdit_height_11.text(),
                    'height_12':self.lineEdit_height_12.text(),
                    'height_13':self.lineEdit_height_13.text(),
                    'height_14':self.lineEdit_height_14.text(),
                    'height_15':self.lineEdit_height_15.text(),
                    'amount_1':self.lineEdit_amount_1.text(),
                    'amount_2':self.lineEdit_amount_2.text(),
                    'amount_3':self.lineEdit_amount_3.text(),
                    'amount_4':self.lineEdit_amount_4.text(),
                    'amount_5':self.lineEdit_amount_5.text(),
                    'amount_6':self.lineEdit_amount_6.text(),
                    'amount_7':self.lineEdit_amount_7.text(),
                    'amount_8':self.lineEdit_amount_8.text(),
                    'amount_9':self.lineEdit_amount_9.text(),
                    'amount_10':self.lineEdit_amount_10.text(),
                    'amount_11':self.lineEdit_amount_11.text(),
                    'amount_12':self.lineEdit_amount_12.text(),
                    'amount_13':self.lineEdit_amount_13.text(),
                    'amount_14':self.lineEdit_amount_14.text(),
                    'amount_15':self.lineEdit_amount_15.text(),
                    'material_1':str(self.comboBox_material_1.currentText()),
                    'material_2':str(self.comboBox_material_2.currentText()),
                    'material_3':str(self.comboBox_material_3.currentText()),
                    'material_4':str(self.comboBox_material_4.currentText()),
                    'material_5':str(self.comboBox_material_5.currentText()),
                    'material_6':str(self.comboBox_material_6.currentText()),
                    'material_7':str(self.comboBox_material_7.currentText()),
                    'material_8':str(self.comboBox_material_8.currentText()),
                    'material_9':str(self.comboBox_material_9.currentText()),
                    'material_10':str(self.comboBox_material_10.currentText()),
                    'material_11':str(self.comboBox_material_11.currentText()),
                    'material_12':str(self.comboBox_material_12.currentText()),
                    'material_13':str(self.comboBox_material_13.currentText()),
                    'material_14':str(self.comboBox_material_14.currentText()),
                    'material_15':str(self.comboBox_material_15.currentText()),
                    'process_1':str(self.comboBox_process_1.currentText()),
                    'process_2':str(self.comboBox_process_2.currentText()),
                    'process_3':str(self.comboBox_process_3.currentText()),
                    'process_4':str(self.comboBox_process_4.currentText()),
                    'process_5':str(self.comboBox_process_5.currentText()),
                    'process_6':str(self.comboBox_process_6.currentText()),
                    'process_7':str(self.comboBox_process_7.currentText()),
                    'process_8':str(self.comboBox_process_8.currentText()),
                    'process_9':str(self.comboBox_process_9.currentText()),
                    'process_10':str(self.comboBox_process_10.currentText()),
                    'process_11':str(self.comboBox_process_11.currentText()),
                    'process_12':str(self.comboBox_process_12.currentText()),
                    'process_13':str(self.comboBox_process_13.currentText()),
                    'process_14':str(self.comboBox_process_14.currentText()),
                    'process_15':str(self.comboBox_process_15.currentText()),
                    'plate_1':str(self.comboBox_plate_1.currentText()),
                    'plate_2':str(self.comboBox_plate_2.currentText()),
                    'plate_3':str(self.comboBox_plate_3.currentText()),
                    'plate_4':str(self.comboBox_plate_4.currentText()),
                    'plate_5':str(self.comboBox_plate_5.currentText()),
                    'plate_6':str(self.comboBox_plate_6.currentText()),
                    'plate_7':str(self.comboBox_plate_7.currentText()),
                    'plate_8':str(self.comboBox_plate_8.currentText()),
                    'plate_9':str(self.comboBox_plate_9.currentText()),
                    'plate_10':str(self.comboBox_plate_10.currentText()),
                    'plate_11':str(self.comboBox_plate_11.currentText()),
                    'plate_12':str(self.comboBox_plate_12.currentText()),
                    'plate_13':str(self.comboBox_plate_13.currentText()),
                    'plate_14':str(self.comboBox_plate_14.currentText()),
                    'plate_15':str(self.comboBox_plate_15.currentText()),
                    'thicknes_1':str(self.comboBox_thicknes_1.currentText()),
                    'thicknes_2':str(self.comboBox_thicknes_2.currentText()),
                    'thicknes_3':str(self.comboBox_thicknes_3.currentText()),
                    'thicknes_4':str(self.comboBox_thicknes_4.currentText()),
                    'thicknes_5':str(self.comboBox_thicknes_5.currentText()),
                    'thicknes_6':str(self.comboBox_thicknes_6.currentText()),
                    'thicknes_7':str(self.comboBox_thicknes_7.currentText()),
                    'thicknes_8':str(self.comboBox_thicknes_8.currentText()),
                    'thicknes_9':str(self.comboBox_thicknes_9.currentText()),
                    'thicknes_10':str(self.comboBox_thicknes_10.currentText()),
                    'thicknes_11':str(self.comboBox_thicknes_11.currentText()),
                    'thicknes_12':str(self.comboBox_thicknes_12.currentText()),
                    'thicknes_13':str(self.comboBox_thicknes_13.currentText()),
                    'thicknes_14':str(self.comboBox_thicknes_14.currentText()),
                    'thicknes_15':str(self.comboBox_thicknes_15.currentText()),
                    'others_1':str(self.comboBox_others_1.currentText()),
                    'others_2':str(self.comboBox_others_2.currentText()),
                    'others_3':str(self.comboBox_others_3.currentText()),
                    'others_4':str(self.comboBox_others_4.currentText()),
                    'others_5':str(self.comboBox_others_5.currentText()),
                    'others_6':str(self.comboBox_others_6.currentText()),
                    'others_7':str(self.comboBox_others_7.currentText()),
                    'others_8':str(self.comboBox_others_8.currentText()),
                    'others_9':str(self.comboBox_others_9.currentText()),
                    'others_10':str(self.comboBox_others_10.currentText()),
                    'others_11':str(self.comboBox_others_11.currentText()),
                    'others_12':str(self.comboBox_others_12.currentText()),
                    'others_13':str(self.comboBox_others_13.currentText()),
                    'others_14':str(self.comboBox_others_14.currentText()),
                    'others_15':str(self.comboBox_others_15.currentText()),
                    'others_amount_1':str(self.lineEdit_others_amount_1.text()),
                    'others_amount_2':str(self.lineEdit_others_amount_2.text()),
                    'others_amount_3':str(self.lineEdit_others_amount_3.text()),
                    'others_amount_4':str(self.lineEdit_others_amount_4.text()),
                    'others_amount_5':str(self.lineEdit_others_amount_5.text()),
                    'others_amount_6':str(self.lineEdit_others_amount_6.text()),
                    'others_amount_7':str(self.lineEdit_others_amount_7.text()),
                    'others_amount_8':str(self.lineEdit_others_amount_8.text()),
                    'others_amount_9':str(self.lineEdit_others_amount_9.text()),
                    'others_amount_10':str(self.lineEdit_others_amount_10.text()),
                    'others_amount_11':str(self.lineEdit_others_amount_11.text()),
                    'others_amount_12':str(self.lineEdit_others_amount_12.text()),
                    'others_amount_13':str(self.lineEdit_others_amount_13.text()),
                    'others_amount_14':str(self.lineEdit_others_amount_14.text()),
                    'others_amount_15':str(self.lineEdit_others_amount_15.text()),

                }
                Update_central_data(dict_data)
            except Exception as e:
                QMessageBox.warning(self, 'Error Information', repr(e), QMessageBox.Ok | QMessageBox.Close, QMessageBox.Close)

        ###直接存price_data        
        def save_price_data():
            try:
                dict_data={
                    'worknum':self.lineEdit_worknum.text(),
                    'lineEdit_CBM_1':self.lineEdit_CBM_1.text(),
                    'lineEdit_CBM_2':self.lineEdit_CBM_2.text(),
                    'lineEdit_CBM_3':self.lineEdit_CBM_3.text(),
                    'lineEdit_CBM_4':self.lineEdit_CBM_4.text(),
                    'lineEdit_CBM_5':self.lineEdit_CBM_5.text(),
                    'lineEdit_CBM_6':self.lineEdit_CBM_6.text(),
                    'lineEdit_CBM_7':self.lineEdit_CBM_7.text(),
                    'lineEdit_CBM_8':self.lineEdit_CBM_8.text(),
                    'lineEdit_CBM_9':self.lineEdit_CBM_9.text(),
                    'lineEdit_CBM_10':self.lineEdit_CBM_10.text(),
                    'lineEdit_CBM_11':self.lineEdit_CBM_11.text(),
                    'lineEdit_CBM_12':self.lineEdit_CBM_12.text(),
                    'lineEdit_CBM_13':self.lineEdit_CBM_13.text(),
                    'lineEdit_CBM_14':self.lineEdit_CBM_14.text(),
                    'lineEdit_CBM_15':self.lineEdit_CBM_15.text(),
                    'lineEdit_CBMprice_1':self.lineEdit_CBMprice_1.text(),
                    'lineEdit_CBMprice_2':self.lineEdit_CBMprice_2.text(),
                    'lineEdit_CBMprice_3':self.lineEdit_CBMprice_3.text(),
                    'lineEdit_CBMprice_4':self.lineEdit_CBMprice_4.text(),
                    'lineEdit_CBMprice_5':self.lineEdit_CBMprice_5.text(),
                    'lineEdit_CBMprice_6':self.lineEdit_CBMprice_6.text(),
                    'lineEdit_CBMprice_7':self.lineEdit_CBMprice_7.text(),
                    'lineEdit_CBMprice_8':self.lineEdit_CBMprice_8.text(),
                    'lineEdit_CBMprice_9':self.lineEdit_CBMprice_9.text(),
                    'lineEdit_CBMprice_10':self.lineEdit_CBMprice_10.text(),
                    'lineEdit_CBMprice_11':self.lineEdit_CBMprice_11.text(),
                    'lineEdit_CBMprice_12':self.lineEdit_CBMprice_12.text(),
                    'lineEdit_CBMprice_13':self.lineEdit_CBMprice_13.text(),
                    'lineEdit_CBMprice_14':self.lineEdit_CBMprice_14.text(),
                    'lineEdit_CBMprice_15':self.lineEdit_CBMprice_15.text(),
                    'lineEdit_single_price_1':self.lineEdit_single_price_1.text(),
                    'lineEdit_single_price_2':self.lineEdit_single_price_2.text(),
                    'lineEdit_single_price_3':self.lineEdit_single_price_3.text(),
                    'lineEdit_single_price_4':self.lineEdit_single_price_4.text(),
                    'lineEdit_single_price_5':self.lineEdit_single_price_5.text(),
                    'lineEdit_single_price_6':self.lineEdit_single_price_6.text(),
                    'lineEdit_single_price_7':self.lineEdit_single_price_7.text(),
                    'lineEdit_single_price_8':self.lineEdit_single_price_8.text(),
                    'lineEdit_single_price_9':self.lineEdit_single_price_9.text(),
                    'lineEdit_single_price_10':self.lineEdit_single_price_10.text(),
                    'lineEdit_single_price_11':self.lineEdit_single_price_11.text(),
                    'lineEdit_single_price_12':self.lineEdit_single_price_12.text(),
                    'lineEdit_single_price_13':self.lineEdit_single_price_13.text(),
                    'lineEdit_single_price_14':self.lineEdit_single_price_14.text(),
                    'lineEdit_single_price_15':self.lineEdit_single_price_15.text(),
                    'lineEdit_single_price_16':self.lineEdit_single_price_16.text(),
                    'lineEdit_single_price_17':self.lineEdit_single_price_17.text(),
                    'lineEdit_single_price_18':self.lineEdit_single_price_18.text(),
                    'lineEdit_single_price_19':self.lineEdit_single_price_19.text(),
                    'lineEdit_single_price_20':self.lineEdit_single_price_20.text(),
                    'lineEdit_single_price_21':self.lineEdit_single_price_21.text(),
                    'lineEdit_single_price_22':self.lineEdit_single_price_22.text(),
                    'lineEdit_single_price_23':self.lineEdit_single_price_23.text(),
                    'lineEdit_single_price_24':self.lineEdit_single_price_24.text(),
                    'lineEdit_single_price_25':self.lineEdit_single_price_25.text(),
                    'lineEdit_single_price_26':self.lineEdit_single_price_26.text(),
                    'lineEdit_single_price_27':self.lineEdit_single_price_27.text(),
                    'lineEdit_single_price_28':self.lineEdit_single_price_28.text(),
                    'lineEdit_single_price_29':self.lineEdit_single_price_29.text(),
                    'lineEdit_single_price_30':self.lineEdit_single_price_30.text(),
                    'lineEdit_tmpprice':self.lineEdit_tmpprice.text(),
                    'lineEdit_tax':self.lineEdit_tax.text(),
                    'lineEdit_final_price':self.lineEdit_final_price.text()
                }    
                Update_price_data(dict_data)
            except Exception as e:
                QMessageBox.warning(self, 'Error Information', repr(e), QMessageBox.Ok | QMessageBox.Close, QMessageBox.Close)


        ###點擊儲存，直接生一個工單編號在三個表裡面
        def save_data():
            try:
                reply = QMessageBox.information(self,'Check Again!','請確認你輸入的資料是否正確', QMessageBox.Ok | QMessageBox.Abort , QMessageBox.Abort)
                if reply == QMessageBox.Ok:
                    create_work_order(self.lineEdit_worknum.text())
                    save_basic_data()
                    save_central_data()
                    save_price_data()
                    excel_all()
                    QMessageBox.information(self,'Save Success!','儲存成功',QMessageBox.Ok)
                else:
                    pass
            except Exception as e:
                QMessageBox.warning(self, 'Error Information', repr(e), QMessageBox.Ok | QMessageBox.Close, QMessageBox.Close)
        self.pushButton_save.clicked.connect(save_data)

        ### 初始化視窗內所有表格
        def init_data():
            self.lineEdit_worknum.setText("")
            self.lineEdit_case_name.setText("")
            self.comboBox_company_name.setCurrentIndex(0)
            self.lineEdit_phone.setText("")
            self.lineEdit_client_name.setText("")
            self.lineEdit_worktime.setText("")
            self.lineEdit_cleanup_time.setText("")
            self.lineEdit_workaddress.setText("")
            self.comboBox_pack.setCurrentIndex(0)
            self.comboBox_transport.setCurrentIndex(0)
            self.comboBox_employee_1.setCurrentIndex(0)
            self.comboBox_employee_2.setCurrentIndex(0)
            self.comboBox_employee_3.setCurrentIndex(0)
            self.comboBox_employee_4.setCurrentIndex(0)
            self.comboBox_employee_5.setCurrentIndex(0)
            self.lineEdit_crossbar_width.setText("")
            self.lineEdit_crossbar_amount.setText("")
            self.lineEdit_crossbar_remark.setText("")
            self.checkBox_150iron_shelter.setChecked(False)
            self.checkBox_180iron_shelter.setChecked(False)
            self.lineEdit_iron_shelter_amount.setText("")
            self.lineEdit_iron_shelter_remark.setText("")
            self.lineEdit_paper_shelter_height.setText("")
            self.lineEdit_paper_shelter_amount.setText("")
            self.lineEdit_paper_shelter_remark.setText("")
            self.lineEdit_stand_style.setText("")
            self.lineEdit_stand_amount.setText("")
            self.lineEdit_stand_remark.setText("")
            self.lineEdit_rent_1.setText("")
            self.lineEdit_rent_2.setText("")
            self.textEdit_remark.setText("")

            for i in range(1,16):
                exec("self.comboBox_product_"+str(i)+".setCurrentIndex(0)")
                exec("self.lineEdit_width_"+str(i)+".setText('')")
                exec("self.lineEdit_height_"+str(i)+".setText('')")
                exec("self.lineEdit_amount_"+str(i)+".setText('')")
                exec("self.comboBox_material_"+str(i)+".setCurrentIndex(0)")
                exec("self.comboBox_process_"+str(i)+".setCurrentIndex(0)")
                exec("self.comboBox_plate_"+str(i)+".setCurrentIndex(0)")
                exec("self.comboBox_thicknes_"+str(i)+".setCurrentIndex(0)")
                exec("self.comboBox_others_"+str(i)+".setCurrentIndex(0)")
                exec("self.lineEdit_others_amount_"+str(i)+".setText('')")
                exec("self.lineEdit_CBM_"+str(i)+".setText('')")
                exec("self.lineEdit_CBMprice_"+str(i)+".setText('')")
                exec("self.lineEdit_single_price_"+str(i)+".setText('')")
            for i in range(16,31):
                exec("self.lineEdit_single_price_"+str(i)+".setText('')")
            self.lineEdit_tmpprice.setText("")
            self.lineEdit_tax.setText("")
            self.lineEdit_final_price.setText("")
        
        ### 顯示basic_data
        def show_basic_data(data):
            try:
                self.lineEdit_worknum.setText(data['worknum'])
                self.lineEdit_case_name.setText(data['case_name'])
                self.comboBox_company_name.setCurrentText(data['company_name'])
                self.lineEdit_phone.setText(data['phone'])
                self.lineEdit_client_name.setText(data['client_name'])
                self.lineEdit_worktime.setText(data['worktime'])
                self.lineEdit_cleanup_time.setText(data['cleanuptime'])
                self.lineEdit_workaddress.setText(data['workaddress'])
                self.comboBox_pack.setCurrentIndex(int(data['pack']))
                self.comboBox_transport.setCurrentIndex(int(data['transport']))
                self.comboBox_employee_1.setCurrentIndex(int(data['cemployee1']))
                self.comboBox_employee_2.setCurrentIndex(int(data['cemployee2']))
                self.comboBox_employee_3.setCurrentIndex(int(data['cemployee3']))
                self.comboBox_employee_4.setCurrentIndex(int(data['cemployee4']))
                self.comboBox_employee_5.setCurrentIndex(int(data['cemployee5']))
                self.lineEdit_crossbar_width.setText(data['crossbar_width'])
                self.lineEdit_crossbar_amount.setText(data['crossbar_amount'])
                self.lineEdit_crossbar_remark.setText(data['crossbar_remark'])
                if data['150shelter'] == "True": 
                    self.checkBox_150iron_shelter.toggle()
                if data['180shelter'] == "True": 
                    self.checkBox_180iron_shelter.toggle()
                self.lineEdit_iron_shelter_amount.setText(data['iron_Shelter_amount'])
                self.lineEdit_iron_shelter_remark.setText(data['iron_Shelter_remark'])
                self.lineEdit_paper_shelter_height.setText(data['paper_Shelter_height'])
                self.lineEdit_paper_shelter_amount.setText(data['paper_Shelter_amount'])
                self.lineEdit_paper_shelter_remark.setText(data['paper_Shelter_remark'])
                self.lineEdit_stand_style.setText(data['stand_style'])
                self.lineEdit_stand_amount.setText(data['stand_amount'])
                self.lineEdit_stand_remark.setText(data['stand_remark'])
                self.lineEdit_rent_1.setText(data['rent1'])
                self.lineEdit_rent_2.setText(data['rent2'])
                self.textEdit_remark.setText(data['remark'])
            except Exception as e:
                QMessageBox.warning(self, 'Error Information', repr(e), QMessageBox.Ok | QMessageBox.Close, QMessageBox.Close)
        
        ###顯示Central_data
        def show_central_data(data):
            try:
                self.comboBox_product_1.setCurrentText(data['comboBox_product_1'])
                self.comboBox_product_2.setCurrentText(data['comboBox_product_2'])
                self.comboBox_product_3.setCurrentText(data['comboBox_product_3'])
                self.comboBox_product_4.setCurrentText(data['comboBox_product_4'])
                self.comboBox_product_5.setCurrentText(data['comboBox_product_5'])
                self.comboBox_product_6.setCurrentText(data['comboBox_product_6'])
                self.comboBox_product_7.setCurrentText(data['comboBox_product_7'])
                self.comboBox_product_8.setCurrentText(data['comboBox_product_8'])
                self.comboBox_product_9.setCurrentText(data['comboBox_product_9'])
                self.comboBox_product_10.setCurrentText(data['comboBox_product_10'])
                self.comboBox_product_11.setCurrentText(data['comboBox_product_11'])
                self.comboBox_product_12.setCurrentText(data['comboBox_product_12'])
                self.comboBox_product_13.setCurrentText(data['comboBox_product_13'])
                self.comboBox_product_14.setCurrentText(data['comboBox_product_14'])
                self.comboBox_product_15.setCurrentText(data['comboBox_product_15'])
                self.lineEdit_width_1.setText(data['lineEdit_width_1'])
                self.lineEdit_width_2.setText(data['lineEdit_width_2'])
                self.lineEdit_width_3.setText(data['lineEdit_width_3'])
                self.lineEdit_width_4.setText(data['lineEdit_width_4'])
                self.lineEdit_width_5.setText(data['lineEdit_width_5'])
                self.lineEdit_width_6.setText(data['lineEdit_width_6'])
                self.lineEdit_width_7.setText(data['lineEdit_width_7'])
                self.lineEdit_width_8.setText(data['lineEdit_width_8'])
                self.lineEdit_width_9.setText(data['lineEdit_width_9'])
                self.lineEdit_width_10.setText(data['lineEdit_width_10'])
                self.lineEdit_width_11.setText(data['lineEdit_width_11'])
                self.lineEdit_width_12.setText(data['lineEdit_width_12'])
                self.lineEdit_width_13.setText(data['lineEdit_width_13'])
                self.lineEdit_width_14.setText(data['lineEdit_width_14'])
                self.lineEdit_width_15.setText(data['lineEdit_width_15'])
                self.lineEdit_height_1.setText(data['lineEdit_height_1'])
                self.lineEdit_height_2.setText(data['lineEdit_height_2'])
                self.lineEdit_height_3.setText(data['lineEdit_height_3'])
                self.lineEdit_height_4.setText(data['lineEdit_height_4'])
                self.lineEdit_height_5.setText(data['lineEdit_height_5'])
                self.lineEdit_height_6.setText(data['lineEdit_height_6'])
                self.lineEdit_height_7.setText(data['lineEdit_height_7'])
                self.lineEdit_height_8.setText(data['lineEdit_height_8'])
                self.lineEdit_height_9.setText(data['lineEdit_height_9'])
                self.lineEdit_height_10.setText(data['lineEdit_height_10'])
                self.lineEdit_height_11.setText(data['lineEdit_height_11'])
                self.lineEdit_height_12.setText(data['lineEdit_height_12'])
                self.lineEdit_height_13.setText(data['lineEdit_height_13'])
                self.lineEdit_height_14.setText(data['lineEdit_height_14'])
                self.lineEdit_height_15.setText(data['lineEdit_height_15'])
                self.lineEdit_amount_1.setText(data['lineEdit_amount_1'])
                self.lineEdit_amount_2.setText(data['lineEdit_amount_2'])
                self.lineEdit_amount_3.setText(data['lineEdit_amount_3'])
                self.lineEdit_amount_4.setText(data['lineEdit_amount_4'])
                self.lineEdit_amount_5.setText(data['lineEdit_amount_5'])
                self.lineEdit_amount_6.setText(data['lineEdit_amount_6'])
                self.lineEdit_amount_7.setText(data['lineEdit_amount_7'])
                self.lineEdit_amount_8.setText(data['lineEdit_amount_8'])
                self.lineEdit_amount_9.setText(data['lineEdit_amount_9'])
                self.lineEdit_amount_10.setText(data['lineEdit_amount_10'])
                self.lineEdit_amount_11.setText(data['lineEdit_amount_11'])
                self.lineEdit_amount_12.setText(data['lineEdit_amount_12'])
                self.lineEdit_amount_13.setText(data['lineEdit_amount_13'])
                self.lineEdit_amount_14.setText(data['lineEdit_amount_14'])
                self.lineEdit_amount_15.setText(data['lineEdit_amount_15'])
                self.comboBox_material_1.setCurrentText(data['comboBox_material_1'])
                self.comboBox_material_2.setCurrentText(data['comboBox_material_2'])
                self.comboBox_material_3.setCurrentText(data['comboBox_material_3'])
                self.comboBox_material_4.setCurrentText(data['comboBox_material_4'])
                self.comboBox_material_5.setCurrentText(data['comboBox_material_5'])
                self.comboBox_material_6.setCurrentText(data['comboBox_material_6'])
                self.comboBox_material_7.setCurrentText(data['comboBox_material_7'])
                self.comboBox_material_8.setCurrentText(data['comboBox_material_8'])
                self.comboBox_material_9.setCurrentText(data['comboBox_material_9'])
                self.comboBox_material_10.setCurrentText(data['comboBox_material_10'])
                self.comboBox_material_11.setCurrentText(data['comboBox_material_11'])
                self.comboBox_material_12.setCurrentText(data['comboBox_material_12'])
                self.comboBox_material_13.setCurrentText(data['comboBox_material_13'])
                self.comboBox_material_14.setCurrentText(data['comboBox_material_14'])
                self.comboBox_material_15.setCurrentText(data['comboBox_material_15'])
                self.comboBox_process_1.setCurrentText(data['comboBox_process_1'])
                self.comboBox_process_2.setCurrentText(data['comboBox_process_2'])
                self.comboBox_process_3.setCurrentText(data['comboBox_process_3'])
                self.comboBox_process_4.setCurrentText(data['comboBox_process_4'])
                self.comboBox_process_5.setCurrentText(data['comboBox_process_5'])
                self.comboBox_process_6.setCurrentText(data['comboBox_process_6'])
                self.comboBox_process_7.setCurrentText(data['comboBox_process_7'])
                self.comboBox_process_8.setCurrentText(data['comboBox_process_8'])
                self.comboBox_process_9.setCurrentText(data['comboBox_process_9'])
                self.comboBox_process_10.setCurrentText(data['comboBox_process_10'])
                self.comboBox_process_11.setCurrentText(data['comboBox_process_11'])
                self.comboBox_process_12.setCurrentText(data['comboBox_process_12'])
                self.comboBox_process_13.setCurrentText(data['comboBox_process_13'])
                self.comboBox_process_14.setCurrentText(data['comboBox_process_14'])
                self.comboBox_process_15.setCurrentText(data['comboBox_process_15'])
                self.comboBox_plate_1.setCurrentText(data['comboBox_plate_1'])
                self.comboBox_plate_2.setCurrentText(data['comboBox_plate_2'])
                self.comboBox_plate_3.setCurrentText(data['comboBox_plate_3'])
                self.comboBox_plate_4.setCurrentText(data['comboBox_plate_4'])
                self.comboBox_plate_5.setCurrentText(data['comboBox_plate_5'])
                self.comboBox_plate_6.setCurrentText(data['comboBox_plate_6'])
                self.comboBox_plate_7.setCurrentText(data['comboBox_plate_7'])
                self.comboBox_plate_8.setCurrentText(data['comboBox_plate_8'])
                self.comboBox_plate_9.setCurrentText(data['comboBox_plate_9'])
                self.comboBox_plate_10.setCurrentText(data['comboBox_plate_10'])
                self.comboBox_plate_11.setCurrentText(data['comboBox_plate_11'])
                self.comboBox_plate_12.setCurrentText(data['comboBox_plate_12'])
                self.comboBox_plate_13.setCurrentText(data['comboBox_plate_13'])
                self.comboBox_plate_14.setCurrentText(data['comboBox_plate_14'])
                self.comboBox_plate_15.setCurrentText(data['comboBox_plate_15'])
                self.comboBox_thicknes_1.setCurrentText(data['comboBox_thicknes_1'])
                self.comboBox_thicknes_2.setCurrentText(data['comboBox_thicknes_2'])
                self.comboBox_thicknes_3.setCurrentText(data['comboBox_thicknes_3'])
                self.comboBox_thicknes_4.setCurrentText(data['comboBox_thicknes_4'])
                self.comboBox_thicknes_5.setCurrentText(data['comboBox_thicknes_5'])
                self.comboBox_thicknes_6.setCurrentText(data['comboBox_thicknes_6'])
                self.comboBox_thicknes_7.setCurrentText(data['comboBox_thicknes_7'])
                self.comboBox_thicknes_8.setCurrentText(data['comboBox_thicknes_8'])
                self.comboBox_thicknes_9.setCurrentText(data['comboBox_thicknes_9'])
                self.comboBox_thicknes_10.setCurrentText(data['comboBox_thicknes_10'])
                self.comboBox_thicknes_11.setCurrentText(data['comboBox_thicknes_11'])
                self.comboBox_thicknes_12.setCurrentText(data['comboBox_thicknes_12'])
                self.comboBox_thicknes_13.setCurrentText(data['comboBox_thicknes_13'])
                self.comboBox_thicknes_14.setCurrentText(data['comboBox_thicknes_14'])
                self.comboBox_thicknes_15.setCurrentText(data['comboBox_thicknes_15'])
                self.comboBox_others_1.setCurrentText(data['comboBox_others_1'])
                self.comboBox_others_2.setCurrentText(data['comboBox_others_2'])
                self.comboBox_others_3.setCurrentText(data['comboBox_others_3'])
                self.comboBox_others_4.setCurrentText(data['comboBox_others_4'])
                self.comboBox_others_5.setCurrentText(data['comboBox_others_5'])
                self.comboBox_others_6.setCurrentText(data['comboBox_others_6'])
                self.comboBox_others_7.setCurrentText(data['comboBox_others_7'])
                self.comboBox_others_8.setCurrentText(data['comboBox_others_8'])
                self.comboBox_others_9.setCurrentText(data['comboBox_others_9'])
                self.comboBox_others_10.setCurrentText(data['comboBox_others_10'])
                self.comboBox_others_11.setCurrentText(data['comboBox_others_11'])
                self.comboBox_others_12.setCurrentText(data['comboBox_others_12'])
                self.comboBox_others_13.setCurrentText(data['comboBox_others_13'])
                self.comboBox_others_14.setCurrentText(data['comboBox_others_14'])
                self.comboBox_others_15.setCurrentText(data['comboBox_others_15'])
                self.lineEdit_others_amount_1.setText(data['lineEdit_others_amount_1'])
                self.lineEdit_others_amount_2.setText(data['lineEdit_others_amount_2'])
                self.lineEdit_others_amount_3.setText(data['lineEdit_others_amount_3'])
                self.lineEdit_others_amount_4.setText(data['lineEdit_others_amount_4'])
                self.lineEdit_others_amount_5.setText(data['lineEdit_others_amount_5'])
                self.lineEdit_others_amount_6.setText(data['lineEdit_others_amount_6'])
                self.lineEdit_others_amount_7.setText(data['lineEdit_others_amount_7'])
                self.lineEdit_others_amount_8.setText(data['lineEdit_others_amount_8'])
                self.lineEdit_others_amount_9.setText(data['lineEdit_others_amount_9'])
                self.lineEdit_others_amount_10.setText(data['lineEdit_others_amount_10'])
                self.lineEdit_others_amount_11.setText(data['lineEdit_others_amount_11'])
                self.lineEdit_others_amount_12.setText(data['lineEdit_others_amount_12'])
                self.lineEdit_others_amount_13.setText(data['lineEdit_others_amount_13'])
                self.lineEdit_others_amount_14.setText(data['lineEdit_others_amount_14'])
                self.lineEdit_others_amount_15.setText(data['lineEdit_others_amount_15'])
            except Exception as e:
                QMessageBox.warning(self, 'Error Information', repr(e), QMessageBox.Ok | QMessageBox.Close, QMessageBox.Close)
        
        ###顯示price data
        def show_price_data(data):
            try:
                self.lineEdit_CBM_1.setText(data['lineEdit_CBM_1'])
                self.lineEdit_CBM_2.setText(data['lineEdit_CBM_2'])
                self.lineEdit_CBM_3.setText(data['lineEdit_CBM_3'])
                self.lineEdit_CBM_4.setText(data['lineEdit_CBM_4'])
                self.lineEdit_CBM_5.setText(data['lineEdit_CBM_5'])
                self.lineEdit_CBM_6.setText(data['lineEdit_CBM_6'])
                self.lineEdit_CBM_7.setText(data['lineEdit_CBM_7'])
                self.lineEdit_CBM_8.setText(data['lineEdit_CBM_8'])
                self.lineEdit_CBM_9.setText(data['lineEdit_CBM_9'])
                self.lineEdit_CBM_10.setText(data['lineEdit_CBM_10'])
                self.lineEdit_CBM_11.setText(data['lineEdit_CBM_11'])
                self.lineEdit_CBM_12.setText(data['lineEdit_CBM_12'])
                self.lineEdit_CBM_13.setText(data['lineEdit_CBM_13'])
                self.lineEdit_CBM_14.setText(data['lineEdit_CBM_14'])
                self.lineEdit_CBM_15.setText(data['lineEdit_CBM_15'])
                self.lineEdit_CBMprice_1.setText(data['lineEdit_CBMprice_1'])
                self.lineEdit_CBMprice_2.setText(data['lineEdit_CBMprice_2'])
                self.lineEdit_CBMprice_3.setText(data['lineEdit_CBMprice_3'])
                self.lineEdit_CBMprice_4.setText(data['lineEdit_CBMprice_4'])
                self.lineEdit_CBMprice_5.setText(data['lineEdit_CBMprice_5'])
                self.lineEdit_CBMprice_6.setText(data['lineEdit_CBMprice_6'])
                self.lineEdit_CBMprice_7.setText(data['lineEdit_CBMprice_7'])
                self.lineEdit_CBMprice_8.setText(data['lineEdit_CBMprice_8'])
                self.lineEdit_CBMprice_9.setText(data['lineEdit_CBMprice_9'])
                self.lineEdit_CBMprice_10.setText(data['lineEdit_CBMprice_10'])
                self.lineEdit_CBMprice_11.setText(data['lineEdit_CBMprice_11'])
                self.lineEdit_CBMprice_12.setText(data['lineEdit_CBMprice_12'])
                self.lineEdit_CBMprice_13.setText(data['lineEdit_CBMprice_13'])
                self.lineEdit_CBMprice_14.setText(data['lineEdit_CBMprice_14'])
                self.lineEdit_CBMprice_15.setText(data['lineEdit_CBMprice_15'])
                self.lineEdit_single_price_1.setText(data['lineEdit_single_price_1'])
                self.lineEdit_single_price_2.setText(data['lineEdit_single_price_2'])
                self.lineEdit_single_price_3.setText(data['lineEdit_single_price_3'])
                self.lineEdit_single_price_4.setText(data['lineEdit_single_price_4'])
                self.lineEdit_single_price_5.setText(data['lineEdit_single_price_5'])
                self.lineEdit_single_price_6.setText(data['lineEdit_single_price_6'])
                self.lineEdit_single_price_7.setText(data['lineEdit_single_price_7'])
                self.lineEdit_single_price_8.setText(data['lineEdit_single_price_8'])
                self.lineEdit_single_price_9.setText(data['lineEdit_single_price_9'])
                self.lineEdit_single_price_10.setText(data['lineEdit_single_price_10'])
                self.lineEdit_single_price_11.setText(data['lineEdit_single_price_11'])
                self.lineEdit_single_price_12.setText(data['lineEdit_single_price_12'])
                self.lineEdit_single_price_13.setText(data['lineEdit_single_price_13'])
                self.lineEdit_single_price_14.setText(data['lineEdit_single_price_14'])
                self.lineEdit_single_price_15.setText(data['lineEdit_single_price_15'])
                self.lineEdit_single_price_16.setText(data['lineEdit_single_price_16'])
                self.lineEdit_single_price_17.setText(data['lineEdit_single_price_17'])
                self.lineEdit_single_price_18.setText(data['lineEdit_single_price_18'])
                self.lineEdit_single_price_19.setText(data['lineEdit_single_price_19'])
                self.lineEdit_single_price_20.setText(data['lineEdit_single_price_20'])
                self.lineEdit_single_price_21.setText(data['lineEdit_single_price_21'])
                self.lineEdit_single_price_22.setText(data['lineEdit_single_price_22'])
                self.lineEdit_single_price_23.setText(data['lineEdit_single_price_23'])
                self.lineEdit_single_price_24.setText(data['lineEdit_single_price_24'])
                self.lineEdit_single_price_25.setText(data['lineEdit_single_price_25'])
                self.lineEdit_single_price_26.setText(data['lineEdit_single_price_26'])
                self.lineEdit_single_price_27.setText(data['lineEdit_single_price_27'])
                self.lineEdit_single_price_28.setText(data['lineEdit_single_price_28'])
                self.lineEdit_single_price_29.setText(data['lineEdit_single_price_29'])
                self.lineEdit_single_price_30.setText(data['lineEdit_single_price_30'])
                self.lineEdit_tmpprice.setText(data['lineEdit_tmpprice'])
                self.lineEdit_tax.setText(data['lineEdit_tax'])
                self.lineEdit_final_price.setText(data['lineEdit_final_price'])
            except Exception as e:
                QMessageBox.warning(self, 'Error Information', repr(e), QMessageBox.Ok | QMessageBox.Close, QMessageBox.Close)
        ###點擊開啟
        def call_data():
            try:
                data_basic = call__basic_data(self.lineEdit_worknum.text())
                data_central = call_central_data(self.lineEdit_worknum.text())
                data_price = call_price_data(self.lineEdit_worknum.text())
                init_data()
                show_basic_data(data_basic)
                show_central_data(data_central)
                show_price_data(data_price)
            except Exception as e:
                QMessageBox.warning(self, 'Error Information', repr(e), QMessageBox.Ok | QMessageBox.Close, QMessageBox.Close)
        self.pushButton_open.clicked.connect(call_data)


        def excel_progess():
            try:
                workbook = openpyxl.load_workbook('excel_progess.xlsx')
                worksheet = workbook.worksheets[0]

                new_workbook = openpyxl.Workbook()
                new_worksheet = new_workbook.active

                for value in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column, values_only=True):
                    value = list(value)
                    new_worksheet.append(value)
                
                new_worksheet['B1'] = self.lineEdit_case_name.text()
                new_worksheet['D1'] = self.comboBox_company_name.currentText()
                new_worksheet['F1'] = self.lineEdit_phone.text()
                new_worksheet['H1'] = self.lineEdit_worknum.text()
                new_worksheet['B2'] = self.lineEdit_client_name.text()
                new_worksheet['D2'] = self.lineEdit_worktime.text()
                new_worksheet['F2'] = self.lineEdit_cleanup_time.text()
                new_worksheet['H2'] = self.lineEdit_workaddress.text()
                new_worksheet['B20'] = self.comboBox_pack.currentText() 
                new_worksheet['B21'] = self.comboBox_transport.currentText()
                new_worksheet['J20'] = self.comboBox_employee_1.currentText()
                new_worksheet['J21'] = self.comboBox_employee_2.currentText()
                new_worksheet['J22'] = self.comboBox_employee_3.currentText()
                new_worksheet['J23'] = self.comboBox_employee_4.currentText()
                new_worksheet['J24'] = self.comboBox_employee_5.currentText()
                new_worksheet['E20'] = self.lineEdit_crossbar_width.text()
                new_worksheet['G20'] = self.lineEdit_crossbar_amount.text()
                if self.checkBox_150iron_shelter.isChecked() == True:
                    new_worksheet['E21'] = '150'
                elif self.checkBox_180iron_shelter.isChecked() == True:
                    new_worksheet['E21'] = '180'
                elif self.checkBox_150iron_shelter.isChecked() == True and self.checkBox_180iron_shelter.isChecked() == True:
                    new_worksheet['E21'] = '150 & 180'
                else:
                    pass
                new_worksheet['G21'] = self.lineEdit_iron_shelter_amount.text()
                new_worksheet['E22'] = self.lineEdit_paper_shelter_height.text()
                new_worksheet['G22'] = self.lineEdit_paper_shelter_amount.text()
                new_worksheet['E23'] = self.lineEdit_stand_style.text()
                new_worksheet['G23'] = self.lineEdit_stand_amount.text()
                new_worksheet['D24'] = self.lineEdit_rent_1.text()
                new_worksheet['G24'] = self.lineEdit_rent_2.text()
                new_worksheet['A4'] = self.comboBox_product_1.currentText()
                new_worksheet['A5'] = self.comboBox_product_2.currentText()
                new_worksheet['A6'] = self.comboBox_product_3.currentText()
                new_worksheet['A7'] = self.comboBox_product_4.currentText()
                new_worksheet['A8'] = self.comboBox_product_5.currentText()
                new_worksheet['A9'] = self.comboBox_product_6.currentText()
                new_worksheet['A10'] = self.comboBox_product_7.currentText()
                new_worksheet['A11'] = self.comboBox_product_8.currentText()
                new_worksheet['A12'] = self.comboBox_product_9.currentText()
                new_worksheet['A13'] = self.comboBox_product_10.currentText()
                new_worksheet['A14'] = self.comboBox_product_11.currentText()
                new_worksheet['A15'] = self.comboBox_product_12.currentText()
                new_worksheet['A16'] = self.comboBox_product_13.currentText()
                new_worksheet['A17'] = self.comboBox_product_14.currentText()
                new_worksheet['A18'] = self.comboBox_product_15.currentText()
                new_worksheet['B4'] = self.lineEdit_width_1.text()
                new_worksheet['B5'] = self.lineEdit_width_2.text()
                new_worksheet['B6'] = self.lineEdit_width_3.text()
                new_worksheet['B7'] = self.lineEdit_width_4.text()
                new_worksheet['B8'] = self.lineEdit_width_5.text()
                new_worksheet['B9'] = self.lineEdit_width_6.text()
                new_worksheet['B10'] = self.lineEdit_width_7.text()
                new_worksheet['B11'] = self.lineEdit_width_8.text()
                new_worksheet['B12'] = self.lineEdit_width_9.text()
                new_worksheet['B13'] = self.lineEdit_width_10.text()
                new_worksheet['B14'] = self.lineEdit_width_11.text()
                new_worksheet['B15'] = self.lineEdit_width_12.text()
                new_worksheet['B16'] = self.lineEdit_width_13.text()
                new_worksheet['B17'] = self.lineEdit_width_14.text()
                new_worksheet['B18'] = self.lineEdit_width_15.text()
                new_worksheet['D4'] = self.lineEdit_height_1.text()
                new_worksheet['D5'] = self.lineEdit_height_2.text()
                new_worksheet['D6'] = self.lineEdit_height_3.text()
                new_worksheet['D7'] = self.lineEdit_height_4.text()
                new_worksheet['D8'] = self.lineEdit_height_5.text()
                new_worksheet['D9'] = self.lineEdit_height_6.text()
                new_worksheet['D10'] = self.lineEdit_height_7.text()
                new_worksheet['D11'] = self.lineEdit_height_8.text()
                new_worksheet['D12'] = self.lineEdit_height_9.text()
                new_worksheet['D13'] = self.lineEdit_height_10.text()
                new_worksheet['D14'] = self.lineEdit_height_11.text()
                new_worksheet['D15'] = self.lineEdit_height_12.text()
                new_worksheet['D16'] = self.lineEdit_height_13.text()
                new_worksheet['D17'] = self.lineEdit_height_14.text()
                new_worksheet['D18'] = self.lineEdit_height_15.text()
                new_worksheet['E4'] = self.lineEdit_amount_1.text()
                new_worksheet['E5'] = self.lineEdit_amount_2.text()
                new_worksheet['E6'] = self.lineEdit_amount_3.text()
                new_worksheet['E7'] = self.lineEdit_amount_4.text()
                new_worksheet['E8'] = self.lineEdit_amount_5.text()
                new_worksheet['E9'] = self.lineEdit_amount_6.text()
                new_worksheet['E10'] = self.lineEdit_amount_7.text()
                new_worksheet['E11'] = self.lineEdit_amount_8.text()
                new_worksheet['E12'] = self.lineEdit_amount_9.text()
                new_worksheet['E13'] = self.lineEdit_amount_10.text()
                new_worksheet['E14'] = self.lineEdit_amount_11.text()
                new_worksheet['E15'] = self.lineEdit_amount_12.text()
                new_worksheet['E16'] = self.lineEdit_amount_13.text()
                new_worksheet['E17'] = self.lineEdit_amount_14.text()
                new_worksheet['E18'] = self.lineEdit_amount_15.text()
                new_worksheet['F4'] = self.comboBox_material_1.currentText()
                new_worksheet['F5'] = self.comboBox_material_2.currentText()
                new_worksheet['F6'] = self.comboBox_material_3.currentText()
                new_worksheet['F7'] = self.comboBox_material_4.currentText()
                new_worksheet['F8'] = self.comboBox_material_5.currentText()
                new_worksheet['F9'] = self.comboBox_material_6.currentText()
                new_worksheet['F10'] = self.comboBox_material_7.currentText()
                new_worksheet['F11'] = self.comboBox_material_8.currentText()
                new_worksheet['F12'] = self.comboBox_material_9.currentText()
                new_worksheet['F13'] = self.comboBox_material_10.currentText()
                new_worksheet['F14'] = self.comboBox_material_11.currentText()
                new_worksheet['F15'] = self.comboBox_material_12.currentText()
                new_worksheet['F16'] = self.comboBox_material_13.currentText()
                new_worksheet['F17'] = self.comboBox_material_14.currentText()
                new_worksheet['F18'] = self.comboBox_material_15.currentText()
                new_worksheet['G4'] = self.comboBox_process_1.currentText()
                new_worksheet['G5'] = self.comboBox_process_2.currentText()
                new_worksheet['G6'] = self.comboBox_process_3.currentText()
                new_worksheet['G7'] = self.comboBox_process_4.currentText()
                new_worksheet['G8'] = self.comboBox_process_5.currentText()
                new_worksheet['G9'] = self.comboBox_process_6.currentText()
                new_worksheet['G10'] = self.comboBox_process_7.currentText()
                new_worksheet['G11'] = self.comboBox_process_8.currentText()
                new_worksheet['G12'] = self.comboBox_process_9.currentText()
                new_worksheet['G13'] = self.comboBox_process_10.currentText()
                new_worksheet['G14'] = self.comboBox_process_11.currentText()
                new_worksheet['G15'] = self.comboBox_process_12.currentText()
                new_worksheet['G16'] = self.comboBox_process_13.currentText()
                new_worksheet['G17'] = self.comboBox_process_14.currentText()
                new_worksheet['G18'] = self.comboBox_process_15.currentText()
                new_worksheet['H4'] = self.comboBox_plate_1.currentText()
                new_worksheet['H5'] = self.comboBox_plate_2.currentText()
                new_worksheet['H6'] = self.comboBox_plate_3.currentText()
                new_worksheet['H7'] = self.comboBox_plate_4.currentText()
                new_worksheet['H8'] = self.comboBox_plate_5.currentText()
                new_worksheet['H9'] = self.comboBox_plate_6.currentText()
                new_worksheet['H10'] = self.comboBox_plate_7.currentText()
                new_worksheet['H11'] = self.comboBox_plate_8.currentText()
                new_worksheet['H12'] = self.comboBox_plate_9.currentText()
                new_worksheet['H13'] = self.comboBox_plate_10.currentText()
                new_worksheet['H14'] = self.comboBox_plate_11.currentText()
                new_worksheet['H15'] = self.comboBox_plate_12.currentText()
                new_worksheet['H16'] = self.comboBox_plate_13.currentText()
                new_worksheet['H17'] = self.comboBox_plate_14.currentText()
                new_worksheet['H18'] = self.comboBox_plate_15.currentText()
                new_worksheet['I4'] = self.comboBox_thicknes_1.currentText()
                new_worksheet['I5'] = self.comboBox_thicknes_2.currentText()
                new_worksheet['I6'] = self.comboBox_thicknes_3.currentText()
                new_worksheet['I7'] = self.comboBox_thicknes_4.currentText()
                new_worksheet['I8'] = self.comboBox_thicknes_5.currentText()
                new_worksheet['I9'] = self.comboBox_thicknes_6.currentText()
                new_worksheet['I10'] = self.comboBox_thicknes_7.currentText()
                new_worksheet['I11'] = self.comboBox_thicknes_8.currentText()
                new_worksheet['I12'] = self.comboBox_thicknes_9.currentText()
                new_worksheet['I13'] = self.comboBox_thicknes_10.currentText()
                new_worksheet['I14'] = self.comboBox_thicknes_11.currentText()
                new_worksheet['I15'] = self.comboBox_thicknes_12.currentText()
                new_worksheet['I16'] = self.comboBox_thicknes_13.currentText()
                new_worksheet['I17'] = self.comboBox_thicknes_14.currentText()
                new_worksheet['I18'] = self.comboBox_thicknes_15.currentText()
                new_worksheet['J4'] = self.comboBox_others_1.currentText()
                new_worksheet['J5'] = self.comboBox_others_2.currentText()
                new_worksheet['J6'] = self.comboBox_others_3.currentText()
                new_worksheet['J7'] = self.comboBox_others_4.currentText()
                new_worksheet['J8'] = self.comboBox_others_5.currentText()
                new_worksheet['J9'] = self.comboBox_others_6.currentText()
                new_worksheet['J10'] = self.comboBox_others_7.currentText()
                new_worksheet['J11'] = self.comboBox_others_8.currentText()
                new_worksheet['J12'] = self.comboBox_others_9.currentText()
                new_worksheet['J13'] = self.comboBox_others_10.currentText()
                new_worksheet['J14'] = self.comboBox_others_11.currentText()
                new_worksheet['J15'] = self.comboBox_others_12.currentText()
                new_worksheet['J16'] = self.comboBox_others_13.currentText()
                new_worksheet['J17'] = self.comboBox_others_14.currentText()
                new_worksheet['J18'] = self.comboBox_others_15.currentText()
                new_worksheet['K4'] = self.lineEdit_others_amount_1.text()
                new_worksheet['K5'] = self.lineEdit_others_amount_2.text()
                new_worksheet['K6'] = self.lineEdit_others_amount_3.text()
                new_worksheet['K7'] = self.lineEdit_others_amount_4.text()
                new_worksheet['K8'] = self.lineEdit_others_amount_5.text()
                new_worksheet['K9'] = self.lineEdit_others_amount_6.text()
                new_worksheet['K10'] = self.lineEdit_others_amount_7.text()
                new_worksheet['K11'] = self.lineEdit_others_amount_8.text()
                new_worksheet['K12'] = self.lineEdit_others_amount_9.text()
                new_worksheet['K13'] = self.lineEdit_others_amount_10.text()
                new_worksheet['K14'] = self.lineEdit_others_amount_11.text()
                new_worksheet['K15'] = self.lineEdit_others_amount_12.text()
                new_worksheet['K16'] = self.lineEdit_others_amount_13.text()
                new_worksheet['K17'] = self.lineEdit_others_amount_14.text()
                new_worksheet['K18'] = self.lineEdit_others_amount_15.text()



                font = Font('新細明體',size = 22,)
                align = Alignment(horizontal="center",vertical="center",wrapText=True)

                for row in range(new_worksheet.max_row):
                    for column in range(new_worksheet.max_column):
                        new_worksheet.cell(row=row+1,column=column+1).font = font
                        new_worksheet.cell(row=row+1,column=column+1).alignment = align

                lks=[]
                for i in range(1,new_worksheet.max_column+1):
                    lk = 1
                    for j in range(1,new_worksheet.max_row+1):
                        sz = new_worksheet.cell(row = j, column = i).value
                        if isinstance(sz,str):
                            lk1 = len(sz.encode('utf-8'))
                        else:
                            lk1 = len(str(sz))
                        if lk < lk1:
                            lk = lk1
                    lks.append(lk)

                for i in range(1,new_worksheet.max_column+1):
                    k = get_column_letter(i)
                    new_worksheet.column_dimensions[k].width = lks[i-1]+10
                
                fille = PatternFill('solid',fgColor='E0FFFF')
                for row in new_worksheet['A3:K3']:
                    for cell in row:
                        cell.fill = fille 

                line_t = Side(style='thin',color = '000000')
                line_T = Side(style='thick',color = '000000')
                border = Border(left=line_t,right=line_t,top=line_t,bottom=line_t)
                border_leftup = Border(left=line_T,right=line_t,top=line_T,bottom=line_t)
                border_rightup = Border(left=line_t,right=line_T,top=line_T,bottom=line_t)
                border_leftdown = Border(left=line_T,right=line_t,top=line_t,bottom=line_T)
                border_rightdown = Border(left=line_t,right=line_T,top=line_t,bottom=line_T)
                border_left = Border(left=line_T,right=line_t,top=line_t,bottom=line_t)
                border_up = Border(left=line_t,right=line_t,top=line_T,bottom=line_t)
                border_down = Border(left=line_t,right=line_t,top=line_t,bottom=line_T)
                border_right = Border(left=line_t,right=line_T,top=line_t,bottom=line_t)
                
                new_worksheet['A1'].border = border_leftup
                new_worksheet['A2'].border = border_leftdown
                new_worksheet['H1'].border = border_rightup
                new_worksheet['H2'].border = border_rightdown
                new_worksheet['A3'].border = border_leftup
                new_worksheet['K3'].border = border_rightup
                new_worksheet['A18'].border = border_leftdown
                new_worksheet['K18'].border = border_rightdown
                for row in new_worksheet['B1:G1']:
                    for cell in row:
                        cell.border = border_up
                for row in new_worksheet['B2:G2']:
                    for cell in row:
                        cell.border = border_down
                
                for row in new_worksheet['A4:A17']:
                    for cell in row:
                        cell.border = border_left
                for row in new_worksheet['B3:J3']:
                    for cell in row:
                        cell.border = border_up
                for row in new_worksheet['K4:K17']:
                    for cell in row:
                        cell.border = border_right
                for row in new_worksheet['B18:J18']:
                    for cell in row:
                        cell.border = border_down
                for row in new_worksheet['B4:J17']:
                    for cell in row:
                        cell.border = border

                new_worksheet['A19'].border = border_leftup
                new_worksheet['B19'].border = border_rightup
                new_worksheet['A21'].border = border_leftdown
                new_worksheet['B21'].border = border_rightdown
                new_worksheet['A20'].border = border_left
                new_worksheet['B20'].border = border_right

                new_worksheet['C19'].border = border_leftup
                new_worksheet['I19'].border = border_rightup
                new_worksheet['C24'].border = border_leftdown
                new_worksheet['I24'].border = border_rightdown

                for row in new_worksheet['D19:I19']:
                    for cell in row:
                        cell.border = border_up
                for row in new_worksheet['C20:C23']:
                    for cell in row:
                        cell.border = border_left
                for row in new_worksheet['I20:I23']:
                    for cell in row:
                        cell.border = border_right
                for row in new_worksheet['D24:H24']:
                    for cell in row:
                        cell.border = border_down
                for row in new_worksheet['D20:H23']:
                    for cell in row:
                        cell.border = border

                new_worksheet['J19'].border = border_leftup
                new_worksheet['K19'].border = border_rightup
                new_worksheet['J24'].border = border_leftdown
                new_worksheet['K24'].border = border_rightdown

                for row in new_worksheet['J20:J23']:
                    for cell in row:
                        cell.border = border_left
                
                for row in new_worksheet['K20:K23']:
                    for cell in row:
                        cell.border = border_right
        
                user_name = getpass.getuser()
                save_path = "C:/Users/" + user_name + "/Desktop/" + self.lineEdit_worknum.text() + "-" + self.comboBox_company_name.currentText() + "-" + self.lineEdit_case_name.text() + ".xlsx"
                print(save_path)
                new_workbook.save(save_path)
                QMessageBox.information(self,'Export Success','匯出成功',QMessageBox.Ok)
            except Exception as e:
                QMessageBox.warning(self, 'Error Information', repr(e), QMessageBox.Ok | QMessageBox.Close, QMessageBox.Close)
        self.pushButton_excelProgess.clicked.connect(excel_progess)

        def excel_all():
            try:
                workbook = openpyxl.load_workbook('excel_all.xlsx')
                ws = workbook.worksheets[0]

                font = Font('新細明體',size = 22)
                align = Alignment(horizontal="center",vertical="center",wrapText=True)

                ws['B1'] = self.lineEdit_case_name.text()
                ws['B1'].font = font
                ws['B1'].alignment = align
                ws['E1'] = self.comboBox_company_name.currentText()
                ws['E1'].font = font
                ws['E1'].alignment = align
                ws['G1'] = self.lineEdit_phone.text()
                ws['G1'].font = font
                ws['G1'].alignment = align
                ws['I1'] = self.lineEdit_worknum.text()
                ws['I1'].font = font
                ws['I1'].alignment = align
                ws['B2'] = self.lineEdit_client_name.text()
                ws['B2'].font = font
                ws['B2'].alignment = align
                ws['E2'] = self.lineEdit_worktime.text()
                ws['E2'].font = font
                ws['E2'].alignment = align
                ws['G2'] = self.lineEdit_cleanup_time.text()
                ws['G2'].font = font
                ws['G2'].alignment = align
                ws['I2'] = self.lineEdit_workaddress.text()
                ws['I2'].font = font
                ws['I2'].alignment = align

                ws['B20'] = self.comboBox_pack.currentText()
                ws['B21'] = self.comboBox_transport.currentText()
                ws['E20'] = self.lineEdit_crossbar_width.text()
                ws['G20'] = self.lineEdit_crossbar_amount.text()
                if self.checkBox_150iron_shelter.isChecked() == True:
                    ws['E21'] = '150'
                elif self.checkBox_180iron_shelter.isChecked() == True:
                    ws['E21'] = '180'
                elif self.checkBox_150iron_shelter.isChecked() == True and self.checkBox_180iron_shelter.isChecked() == True:
                    ws['E21'] = '150 and 180'
                else:
                    pass
                ws['G21'] = self.lineEdit_iron_shelter_amount.text()
                ws['E22'] = self.lineEdit_paper_shelter_height.text()
                ws['G22'] = self.lineEdit_paper_shelter_amount.text()
                ws['E23'] = self.lineEdit_stand_style.text()
                ws['G23'] = self.lineEdit_stand_amount.text()
                ws['D24'] = self.lineEdit_rent_1.text()
                ws['G24'] = self.lineEdit_rent_2.text()

                ws['A23'] = self.comboBox_employee_1.currentText()
                ws['A24'] = self.comboBox_employee_2.currentText()
                ws['A25'] = self.comboBox_employee_3.currentText()
                ws['A26'] = self.comboBox_employee_4.currentText()
                ws['A27'] = self.comboBox_employee_5.currentText()
    ###############################################################################
                ws['A4'] = self.comboBox_product_1.currentText()
                ws['B4'] = self.lineEdit_width_1.text()
                ws['C4'] = self.lineEdit_height_1.text()
                ws['D4'] = self.lineEdit_amount_1.text()
                ws['E4'] = self.comboBox_material_1.currentText()
                ws['F4'] = self.comboBox_process_1.currentText()
                ws['G4'] = self.comboBox_plate_1.currentText()
                ws['H4'] = self.comboBox_thicknes_1.currentText()
                ws['I4'] = self.comboBox_others_1.currentText()
                ws['J4'] = self.lineEdit_others_amount_1.text()
                ws['K4'] = self.lineEdit_CBM_1.text()
                ws['L4'] = self.lineEdit_CBMprice_1.text()
                ws['M4'] = self.lineEdit_single_price_1.text()
                ws['N4'] = self.lineEdit_single_price_16.text()

                ws['A5'] = self.comboBox_product_2.currentText()
                ws['B5'] = self.lineEdit_width_2.text()
                ws['C5'] = self.lineEdit_height_2.text()
                ws['D5'] = self.lineEdit_amount_2.text()
                ws['E5'] = self.comboBox_material_2.currentText()
                ws['F5'] = self.comboBox_process_2.currentText()
                ws['G5'] = self.comboBox_plate_2.currentText()
                ws['H5'] = self.comboBox_thicknes_2.currentText()
                ws['I5'] = self.comboBox_others_2.currentText()
                ws['J5'] = self.lineEdit_others_amount_2.text()
                ws['K5'] = self.lineEdit_CBM_2.text()
                ws['L5'] = self.lineEdit_CBMprice_2.text()
                ws['M5'] = self.lineEdit_single_price_2.text()
                ws['N5'] = self.lineEdit_single_price_17.text()

                ws['A6'] = self.comboBox_product_3.currentText()
                ws['B6'] = self.lineEdit_width_3.text()
                ws['C6'] = self.lineEdit_height_3.text()
                ws['D6'] = self.lineEdit_amount_3.text()
                ws['E6'] = self.comboBox_material_3.currentText()
                ws['F6'] = self.comboBox_process_3.currentText()
                ws['G6'] = self.comboBox_plate_3.currentText()
                ws['H6'] = self.comboBox_thicknes_3.currentText()
                ws['I6'] = self.comboBox_others_3.currentText()
                ws['J6'] = self.lineEdit_others_amount_3.text()
                ws['K6'] = self.lineEdit_CBM_3.text()
                ws['L6'] = self.lineEdit_CBMprice_3.text()
                ws['M6'] = self.lineEdit_single_price_3.text()
                ws['N6'] = self.lineEdit_single_price_18.text()

                ws['A7'] = self.comboBox_product_4.currentText()
                ws['B7'] = self.lineEdit_width_4.text()
                ws['C7'] = self.lineEdit_height_4.text()
                ws['D7'] = self.lineEdit_amount_4.text()
                ws['E7'] = self.comboBox_material_4.currentText()
                ws['F7'] = self.comboBox_process_4.currentText()
                ws['G7'] = self.comboBox_plate_4.currentText()
                ws['H7'] = self.comboBox_thicknes_4.currentText()
                ws['I7'] = self.comboBox_others_4.currentText()
                ws['J7'] = self.lineEdit_others_amount_4.text()
                ws['K7'] = self.lineEdit_CBM_4.text()
                ws['L7'] = self.lineEdit_CBMprice_4.text()
                ws['M7'] = self.lineEdit_single_price_4.text()
                ws['N7'] = self.lineEdit_single_price_19.text()

                ws['A8'] = self.comboBox_product_5.currentText()
                ws['B8'] = self.lineEdit_width_5.text()
                ws['C8'] = self.lineEdit_height_5.text()
                ws['D8'] = self.lineEdit_amount_5.text()
                ws['E8'] = self.comboBox_material_5.currentText()
                ws['F8'] = self.comboBox_process_5.currentText()
                ws['G8'] = self.comboBox_plate_5.currentText()
                ws['H8'] = self.comboBox_thicknes_5.currentText()
                ws['I8'] = self.comboBox_others_5.currentText()
                ws['J8'] = self.lineEdit_others_amount_5.text()
                ws['K8'] = self.lineEdit_CBM_5.text()
                ws['L8'] = self.lineEdit_CBMprice_5.text()
                ws['M8'] = self.lineEdit_single_price_5.text()
                ws['N8'] = self.lineEdit_single_price_20.text()

                ws['A9'] = self.comboBox_product_6.currentText()
                ws['B9'] = self.lineEdit_width_6.text()
                ws['C9'] = self.lineEdit_height_6.text()
                ws['D9'] = self.lineEdit_amount_6.text()
                ws['E9'] = self.comboBox_material_6.currentText()
                ws['F9'] = self.comboBox_process_6.currentText()
                ws['G9'] = self.comboBox_plate_6.currentText()
                ws['H9'] = self.comboBox_thicknes_6.currentText()
                ws['I9'] = self.comboBox_others_6.currentText()
                ws['J9'] = self.lineEdit_others_amount_6.text()
                ws['K9'] = self.lineEdit_CBM_6.text()
                ws['L9'] = self.lineEdit_CBMprice_6.text()
                ws['M9'] = self.lineEdit_single_price_6.text()
                ws['N9'] = self.lineEdit_single_price_21.text()

                ws['A10'] = self.comboBox_product_7.currentText()
                ws['B10'] = self.lineEdit_width_7.text()
                ws['C10'] = self.lineEdit_height_7.text()
                ws['D10'] = self.lineEdit_amount_7.text()
                ws['E10'] = self.comboBox_material_7.currentText()
                ws['F10'] = self.comboBox_process_7.currentText()
                ws['G10'] = self.comboBox_plate_7.currentText()
                ws['H10'] = self.comboBox_thicknes_7.currentText()
                ws['I10'] = self.comboBox_others_7.currentText()
                ws['J10'] = self.lineEdit_others_amount_7.text()
                ws['K10'] = self.lineEdit_CBM_7.text()
                ws['L10'] = self.lineEdit_CBMprice_7.text()
                ws['M10'] = self.lineEdit_single_price_7.text()
                ws['N10'] = self.lineEdit_single_price_22.text()

                ws['A11'] = self.comboBox_product_8.currentText()
                ws['B11'] = self.lineEdit_width_8.text()
                ws['C11'] = self.lineEdit_height_8.text()
                ws['D11'] = self.lineEdit_amount_8.text()
                ws['E11'] = self.comboBox_material_8.currentText()
                ws['F11'] = self.comboBox_process_8.currentText()
                ws['G11'] = self.comboBox_plate_8.currentText()
                ws['H11'] = self.comboBox_thicknes_8.currentText()
                ws['I11'] = self.comboBox_others_8.currentText()
                ws['J11'] = self.lineEdit_others_amount_8.text()
                ws['K11'] = self.lineEdit_CBM_8.text()
                ws['L11'] = self.lineEdit_CBMprice_8.text()
                ws['M11'] = self.lineEdit_single_price_8.text()
                ws['N11'] = self.lineEdit_single_price_23.text()

                ws['A12'] = self.comboBox_product_9.currentText()
                ws['B12'] = self.lineEdit_width_9.text()
                ws['C12'] = self.lineEdit_height_9.text()
                ws['D12'] = self.lineEdit_amount_9.text()
                ws['E12'] = self.comboBox_material_9.currentText()
                ws['F12'] = self.comboBox_process_9.currentText()
                ws['G12'] = self.comboBox_plate_9.currentText()
                ws['H12'] = self.comboBox_thicknes_9.currentText()
                ws['I12'] = self.comboBox_others_9.currentText()
                ws['J12'] = self.lineEdit_others_amount_9.text()
                ws['K12'] = self.lineEdit_CBM_9.text()
                ws['L12'] = self.lineEdit_CBMprice_9.text()
                ws['M12'] = self.lineEdit_single_price_9.text()
                ws['N12'] = self.lineEdit_single_price_24.text()

                ws['A13'] = self.comboBox_product_10.currentText()
                ws['B13'] = self.lineEdit_width_10.text()
                ws['C13'] = self.lineEdit_height_10.text()
                ws['D13'] = self.lineEdit_amount_10.text()
                ws['E13'] = self.comboBox_material_10.currentText()
                ws['F13'] = self.comboBox_process_10.currentText()
                ws['G13'] = self.comboBox_plate_10.currentText()
                ws['H13'] = self.comboBox_thicknes_10.currentText()
                ws['I13'] = self.comboBox_others_10.currentText()
                ws['J13'] = self.lineEdit_others_amount_10.text()
                ws['K13'] = self.lineEdit_CBM_10.text()
                ws['L13'] = self.lineEdit_CBMprice_10.text()
                ws['M13'] = self.lineEdit_single_price_10.text()
                ws['N13'] = self.lineEdit_single_price_25.text()

                ws['A14'] = self.comboBox_product_11.currentText()
                ws['B14'] = self.lineEdit_width_11.text()
                ws['C14'] = self.lineEdit_height_11.text()
                ws['D14'] = self.lineEdit_amount_11.text()
                ws['E14'] = self.comboBox_material_11.currentText()
                ws['F14'] = self.comboBox_process_11.currentText()
                ws['G14'] = self.comboBox_plate_11.currentText()
                ws['H14'] = self.comboBox_thicknes_11.currentText()
                ws['I14'] = self.comboBox_others_11.currentText()
                ws['J14'] = self.lineEdit_others_amount_11.text()
                ws['K14'] = self.lineEdit_CBM_11.text()
                ws['L14'] = self.lineEdit_CBMprice_11.text()
                ws['M14'] = self.lineEdit_single_price_11.text()
                ws['N14'] = self.lineEdit_single_price_26.text()

                ws['A15'] = self.comboBox_product_12.currentText()
                ws['B15'] = self.lineEdit_width_12.text()
                ws['C15'] = self.lineEdit_height_12.text()
                ws['D15'] = self.lineEdit_amount_12.text()
                ws['E15'] = self.comboBox_material_12.currentText()
                ws['F15'] = self.comboBox_process_12.currentText()
                ws['G15'] = self.comboBox_plate_12.currentText()
                ws['H15'] = self.comboBox_thicknes_12.currentText()
                ws['I15'] = self.comboBox_others_12.currentText()
                ws['J15'] = self.lineEdit_others_amount_12.text()
                ws['K15'] = self.lineEdit_CBM_12.text()
                ws['L15'] = self.lineEdit_CBMprice_12.text()
                ws['M15'] = self.lineEdit_single_price_12.text()
                ws['N15'] = self.lineEdit_single_price_27.text()

                ws['A16'] = self.comboBox_product_13.currentText()
                ws['B16'] = self.lineEdit_width_13.text()
                ws['C16'] = self.lineEdit_height_13.text()
                ws['D16'] = self.lineEdit_amount_13.text()
                ws['E16'] = self.comboBox_material_13.currentText()
                ws['F16'] = self.comboBox_process_13.currentText()
                ws['G16'] = self.comboBox_plate_13.currentText()
                ws['H16'] = self.comboBox_thicknes_13.currentText()
                ws['I16'] = self.comboBox_others_13.currentText()
                ws['J16'] = self.lineEdit_others_amount_13.text()
                ws['K16'] = self.lineEdit_CBM_13.text()
                ws['L16'] = self.lineEdit_CBMprice_13.text()
                ws['M16'] = self.lineEdit_single_price_13.text()
                ws['N16'] = self.lineEdit_single_price_28.text()

                ws['A17'] = self.comboBox_product_14.currentText()
                ws['B17'] = self.lineEdit_width_14.text()
                ws['C17'] = self.lineEdit_height_14.text()
                ws['D17'] = self.lineEdit_amount_14.text()
                ws['E17'] = self.comboBox_material_14.currentText()
                ws['F17'] = self.comboBox_process_14.currentText()
                ws['G17'] = self.comboBox_plate_14.currentText()
                ws['H17'] = self.comboBox_thicknes_14.currentText()
                ws['I17'] = self.comboBox_others_14.currentText()
                ws['J17'] = self.lineEdit_others_amount_14.text()
                ws['K17'] = self.lineEdit_CBM_14.text()
                ws['L17'] = self.lineEdit_CBMprice_14.text()
                ws['M17'] = self.lineEdit_single_price_14.text()
                ws['N17'] = self.lineEdit_single_price_29.text()

                ws['A18'] = self.comboBox_product_15.currentText()
                ws['B18'] = self.lineEdit_width_15.text()
                ws['C18'] = self.lineEdit_height_15.text()
                ws['D18'] = self.lineEdit_amount_15.text()
                ws['E18'] = self.comboBox_material_15.currentText()
                ws['F18'] = self.comboBox_process_15.currentText()
                ws['G18'] = self.comboBox_plate_15.currentText()
                ws['H18'] = self.comboBox_thicknes_15.currentText()
                ws['I18'] = self.comboBox_others_15.currentText()
                ws['J18'] = self.lineEdit_others_amount_15.text()
                ws['K18'] = self.lineEdit_CBM_15.text()
                ws['L18'] = self.lineEdit_CBMprice_15.text()
                ws['M18'] = self.lineEdit_single_price_15.text()
                ws['N18'] = self.lineEdit_single_price_30.text()


                ws['L19'] = self.lineEdit_tmpprice.text()
                ws['L21'] = self.lineEdit_tax.text()
                ws['L23'] = self.lineEdit_final_price.text()


                for row in ws['A4:N27']:
                    for cell in row:
                        cell.font = font
                        cell.alignment = align
                txt = (self.lineEdit_worknum.text()).split("-")
                #print(txt)
                path = 'X:\\' + txt[0] +"\\" + txt[1] 
                folder = os.path.exists(path)
                if not folder:
                    os.makedirs(path)
                
                filename = self.lineEdit_worknum.text()
                
                for root,dirs,files in os.walk(path):
                    for name in files:
                        if filename in name:
                            os.remove(path+"\\"+name)

                save_path1 = "X:\\" + txt[0] + "\\" + txt[1] + "\\"  + self.lineEdit_worknum.text() + "-" + self.comboBox_company_name.currentText() + "-" + self.lineEdit_case_name.text() + ".xlsx"               
                workbook.save(save_path1)
                QMessageBox.information(self,'Export Success','匯出成功',QMessageBox.Ok)
            except Exception as e:
                QMessageBox.warning(self, 'Error Information', repr(e), QMessageBox.Ok | QMessageBox.Close, QMessageBox.Close)
        #self.pushButton_excelALL.clicked.connect(excel_all)

        def excel_payment():
            try:
                try:
                    full_name,phone,address,taxID = company_name_change(self.comboBox_company_name.currentText())
                except:
                    full_name = ''
                    phone = ''
                    address = ''
                    taxID = ''
                workbook = openpyxl.load_workbook('excel_payment.xlsx')
                ws = workbook.worksheets[0]

                font = Font('新細明體',size = 22)
                align = Alignment(horizontal="left",vertical="center",wrapText=True)
                align1 = Alignment(horizontal="center",vertical="center",wrapText=True)

                ws['C3'] = full_name
                ws['C3'].font = font
                ws['C3'].alignment = align
                ws['G3'] = self.lineEdit_case_name.text()
                ws['G3'].font = font
                ws['G3'].alignment = align
                ws['I3'] = self.lineEdit_client_name.text()
                ws['I3'].font = font
                ws['I3'].alignment = align
                ws['K3'] = phone
                ws['K3'].font = font
                ws['K3'].alignment = align
                ws['C5'] = taxID
                ws['C5'].font = font
                ws['C5'].alignment = align
                ws['I4'] = self.lineEdit_worktime.text()
                ws['I4'].font = font
                ws['I4'].alignment = align
                ws['K4'] = self.lineEdit_cleanup_time.text()
                ws['K4'].font = font
                ws['K4'].alignment = align
                ws['C4'] = self.lineEdit_workaddress.text()
                ws['C4'].font = font
                ws['C4'].alignment = align
                ws['F5'] = address
                ws['F5'].font = font
                ws['F5'].alignment = align
                
                ws['B7'] = self.comboBox_product_1.currentText()
                if self.lineEdit_width_1.text() != "" : ws['C7'] = self.lineEdit_width_1.text() + " X " + self.lineEdit_height_1.text()
                ws['E7'] = self.lineEdit_amount_1.text()
                ws['F7'] = self.comboBox_material_1.currentText()
                ws['G7'] = self.comboBox_thicknes_1.currentText() + self.comboBox_plate_1.currentText()
                ws['H7'] = self.comboBox_others_1.currentText()
                ws['I7'] = self.lineEdit_others_amount_1.text()
                ws['J7'] = self.lineEdit_CBM_1.text()
                ws['K7'] = self.lineEdit_single_price_1.text()
                ws['L7'] = self.lineEdit_single_price_16.text()

                ws['B8'] = self.comboBox_product_2.currentText()
                if self.lineEdit_width_2.text() != "" : ws['C8'] = self.lineEdit_width_2.text() + " X " + self.lineEdit_height_2.text()
                ws['E8'] = self.lineEdit_amount_2.text()
                ws['F8'] = self.comboBox_material_2.currentText()
                ws['G8'] = self.comboBox_thicknes_2.currentText() + self.comboBox_plate_2.currentText()
                ws['H8'] = self.comboBox_others_2.currentText()
                ws['I8'] = self.lineEdit_others_amount_2.text()
                ws['J8'] = self.lineEdit_CBM_2.text()
                ws['K8'] = self.lineEdit_single_price_2.text()
                ws['L8'] = self.lineEdit_single_price_17.text()

                ws['B9'] = self.comboBox_product_3.currentText()
                if self.lineEdit_width_3.text() != "" : ws['C9'] = self.lineEdit_width_3.text() + " X " + self.lineEdit_height_3.text()
                ws['E9'] = self.lineEdit_amount_3.text()
                ws['F9'] = self.comboBox_material_3.currentText()
                ws['G9'] = self.comboBox_thicknes_3.currentText() + self.comboBox_plate_3.currentText()
                ws['H9'] = self.comboBox_others_3.currentText()
                ws['I9'] = self.lineEdit_others_amount_3.text()
                ws['J9'] = self.lineEdit_CBM_3.text()
                ws['K9'] = self.lineEdit_single_price_3.text()
                ws['L9'] = self.lineEdit_single_price_18.text()

                ws['B10'] = self.comboBox_product_4.currentText()
                if self.lineEdit_width_4.text() != "" : ws['C10'] = self.lineEdit_width_4.text() + " X " + self.lineEdit_height_4.text()
                ws['E10'] = self.lineEdit_amount_4.text()
                ws['F10'] = self.comboBox_material_4.currentText()
                ws['G10'] = self.comboBox_thicknes_4.currentText() + self.comboBox_plate_4.currentText()
                ws['H10'] = self.comboBox_others_4.currentText()
                ws['I10'] = self.lineEdit_others_amount_4.text()
                ws['J10'] = self.lineEdit_CBM_4.text()
                ws['K10'] = self.lineEdit_single_price_4.text()
                ws['L10'] = self.lineEdit_single_price_19.text()

                ws['B11'] = self.comboBox_product_5.currentText()
                if self.lineEdit_width_5.text() != "" : ws['C11'] = self.lineEdit_width_5.text() + " X " + self.lineEdit_height_5.text()
                ws['E11'] = self.lineEdit_amount_5.text()
                ws['F11'] = self.comboBox_material_5.currentText()
                ws['G11'] = self.comboBox_thicknes_5.currentText() + self.comboBox_plate_5.currentText()
                ws['H11'] = self.comboBox_others_5.currentText()
                ws['I11'] = self.lineEdit_others_amount_5.text()
                ws['J11'] = self.lineEdit_CBM_5.text()
                ws['K11'] = self.lineEdit_single_price_5.text()
                ws['L11'] = self.lineEdit_single_price_20.text()

                ws['B12'] = self.comboBox_product_6.currentText()
                if self.lineEdit_width_6.text() != "" : ws['C12'] = self.lineEdit_width_6.text() + " X " + self.lineEdit_height_6.text()
                ws['E12'] = self.lineEdit_amount_6.text()
                ws['F12'] = self.comboBox_material_6.currentText()
                ws['G12'] = self.comboBox_thicknes_6.currentText() + self.comboBox_plate_6.currentText()
                ws['H12'] = self.comboBox_others_6.currentText()
                ws['I12'] = self.lineEdit_others_amount_6.text()
                ws['J12'] = self.lineEdit_CBM_6.text()
                ws['K12'] = self.lineEdit_single_price_6.text()
                ws['L12'] = self.lineEdit_single_price_21.text()

                ws['B13'] = self.comboBox_product_7.currentText()
                if self.lineEdit_width_7.text() != "" : ws['C13'] = self.lineEdit_width_7.text() + " X " + self.lineEdit_height_7.text()
                ws['E13'] = self.lineEdit_amount_7.text()
                ws['F13'] = self.comboBox_material_7.currentText()
                ws['G13'] = self.comboBox_thicknes_7.currentText() + self.comboBox_plate_7.currentText()
                ws['H13'] = self.comboBox_others_7.currentText()
                ws['I13'] = self.lineEdit_others_amount_7.text()
                ws['J13'] = self.lineEdit_CBM_7.text()
                ws['K13'] = self.lineEdit_single_price_7.text()
                ws['L13'] = self.lineEdit_single_price_22.text()

                ws['B14'] = self.comboBox_product_8.currentText()
                if self.lineEdit_width_8.text() != "" : ws['C14'] = self.lineEdit_width_8.text() + " X " + self.lineEdit_height_8.text()
                ws['E14'] = self.lineEdit_amount_8.text()
                ws['F14'] = self.comboBox_material_8.currentText()
                ws['G14'] = self.comboBox_thicknes_8.currentText() + self.comboBox_plate_8.currentText()
                ws['H14'] = self.comboBox_others_8.currentText()
                ws['I14'] = self.lineEdit_others_amount_8.text()
                ws['J14'] = self.lineEdit_CBM_8.text()
                ws['K14'] = self.lineEdit_single_price_8.text()
                ws['L14'] = self.lineEdit_single_price_23.text()

                ws['B15'] = self.comboBox_product_9.currentText()
                if self.lineEdit_width_9.text() != "" : ws['C15'] = self.lineEdit_width_9.text() + " X " + self.lineEdit_height_9.text()
                ws['E15'] = self.lineEdit_amount_9.text()
                ws['F15'] = self.comboBox_material_9.currentText()
                ws['G15'] = self.comboBox_thicknes_9.currentText() + self.comboBox_plate_9.currentText()
                ws['H15'] = self.comboBox_others_9.currentText()
                ws['I15'] = self.lineEdit_others_amount_9.text()
                ws['J15'] = self.lineEdit_CBM_9.text()
                ws['K15'] = self.lineEdit_single_price_9.text()
                ws['L15'] = self.lineEdit_single_price_24.text()

                ws['B16'] = self.comboBox_product_10.currentText()
                if self.lineEdit_width_10.text() != "" : ws['C16'] = self.lineEdit_width_10.text() + " X " + self.lineEdit_height_10.text()
                ws['E16'] = self.lineEdit_amount_10.text()
                ws['F16'] = self.comboBox_material_10.currentText()
                ws['G16'] = self.comboBox_thicknes_10.currentText() + self.comboBox_plate_10.currentText()
                ws['H16'] = self.comboBox_others_10.currentText()
                ws['I16'] = self.lineEdit_others_amount_10.text()
                ws['J16'] = self.lineEdit_CBM_10.text()
                ws['K16'] = self.lineEdit_single_price_10.text()
                ws['L16'] = self.lineEdit_single_price_25.text()

                ws['B17'] = self.comboBox_product_11.currentText()
                if self.lineEdit_width_11.text() != "" : ws['C17'] = self.lineEdit_width_11.text() + " X " + self.lineEdit_height_11.text()
                ws['E17'] = self.lineEdit_amount_11.text()
                ws['F17'] = self.comboBox_material_11.currentText()
                ws['G17'] = self.comboBox_thicknes_11.currentText() + self.comboBox_plate_11.currentText()
                ws['H17'] = self.comboBox_others_11.currentText()
                ws['I17'] = self.lineEdit_others_amount_11.text()
                ws['J17'] = self.lineEdit_CBM_11.text()
                ws['K17'] = self.lineEdit_single_price_11.text()
                ws['L17'] = self.lineEdit_single_price_26.text()

                ws['B18'] = self.comboBox_product_12.currentText()
                if self.lineEdit_width_12.text() != "" : ws['C18'] = self.lineEdit_width_12.text() + " X " + self.lineEdit_height_12.text()
                ws['E18'] = self.lineEdit_amount_12.text()
                ws['F18'] = self.comboBox_material_12.currentText()
                ws['G18'] = self.comboBox_thicknes_12.currentText() + self.comboBox_plate_12.currentText()
                ws['H18'] = self.comboBox_others_12.currentText()
                ws['I18'] = self.lineEdit_others_amount_12.text()
                ws['J18'] = self.lineEdit_CBM_12.text()
                ws['K18'] = self.lineEdit_single_price_12.text()
                ws['L18'] = self.lineEdit_single_price_27.text()

                ws['B19'] = self.comboBox_product_13.currentText()
                if self.lineEdit_width_13.text() != "" : ws['C19'] = self.lineEdit_width_13.text() + " X " + self.lineEdit_height_13.text()
                ws['E19'] = self.lineEdit_amount_13.text()
                ws['F19'] = self.comboBox_material_13.currentText()
                ws['G19'] = self.comboBox_thicknes_13.currentText() + self.comboBox_plate_13.currentText()
                ws['H19'] = self.comboBox_others_13.currentText()
                ws['I19'] = self.lineEdit_others_amount_13.text()
                ws['J19'] = self.lineEdit_CBM_13.text()
                ws['K19'] = self.lineEdit_single_price_13.text()
                ws['L19'] = self.lineEdit_single_price_28.text()

                ws['B20'] = self.comboBox_product_14.currentText()
                if self.lineEdit_width_14.text() != "" : ws['C20'] = self.lineEdit_width_14.text() + " X " + self.lineEdit_height_14.text()
                ws['E20'] = self.lineEdit_amount_14.text()
                ws['F20'] = self.comboBox_material_14.currentText()
                ws['G20'] = self.comboBox_thicknes_14.currentText() + self.comboBox_plate_14.currentText()
                ws['H20'] = self.comboBox_others_14.currentText()
                ws['I20'] = self.lineEdit_others_amount_14.text()
                ws['J20'] = self.lineEdit_CBM_14.text()
                ws['K20'] = self.lineEdit_single_price_14.text()
                ws['L20'] = self.lineEdit_single_price_29.text()

                ws['B21'] = self.comboBox_product_15.currentText()
                if self.lineEdit_width_15.text() != "" : ws['C21'] = self.lineEdit_width_15.text() + " X " + self.lineEdit_height_15.text()
                ws['E21'] = self.lineEdit_amount_15.text()
                ws['F21'] = self.comboBox_material_15.currentText()
                ws['G21'] = self.comboBox_thicknes_15.currentText() + self.comboBox_plate_15.currentText()
                ws['H21'] = self.comboBox_others_15.currentText()
                ws['I21'] = self.lineEdit_others_amount_15.text()
                ws['J21'] = self.lineEdit_CBM_15.text()
                ws['K21'] = self.lineEdit_single_price_15.text()
                ws['L21'] = self.lineEdit_single_price_30.text()

                ws['H22'] = self.lineEdit_tmpprice.text()
                ws['H22'].font = font
                ws['H22'].alignment = align1
                ws['J22'] = self.lineEdit_tax.text()
                ws['J22'].font = font
                ws['J22'].alignment = align1
                ws['L22'] = self.lineEdit_final_price.text()
                ws['L22'].font = font
                ws['L22'].alignment = align1
                for row in ws['B7:L21']:
                    for cell in row:
                        cell.font = font
                        cell.alignment = align1
                user_name = getpass.getuser()
                check = "C:/Users/" + user_name + "/Desktop/報價留底/"
                folder = os.path.exists(check)
                if not folder:
                    os.makedirs(check)
                save_path = "C:/Users/" + user_name + "/Desktop/報價留底/" + self.lineEdit_worknum.text() + "-" + self.comboBox_company_name.currentText() + "-" + self.lineEdit_case_name.text() + ".xlsx"
                workbook.save(save_path)
                QMessageBox.information(self,'Export Success','匯出成功',QMessageBox.Ok)
            except Exception as e:
                QMessageBox.warning(self, 'Error Information', repr(e), QMessageBox.Ok | QMessageBox.Close, QMessageBox.Close)
        self.pushButton_excelpayment.clicked.connect(excel_payment)

        def clean_data():
            init_data()
        self.pushButton_clean.clicked.connect(clean_data)


if __name__ == '__main__':
    import sys
    from PyQt5 import QtWidgets
    from PyQt5 import QtCore
    from PyQt5 import QtGui
    #QtGui.QGuiApplication.setHighDpiScaleFactorRoundingPolicy(QtCore.Qt.HighDpiScaleFactorRoundingPolicy.Floor)
    QtWidgets.QApplication.setAttribute(QtCore.Qt.AA_EnableHighDpiScaling, True)
    app = QtWidgets.QApplication(sys.argv)
    window = Main()
    window.showMaximized()
    window.show()
    sys.exit(app.exec())





