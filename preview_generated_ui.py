from __future__ import annotations

import argparse
import difflib
import json
import os
import re
import subprocess
import sys
import unicodedata
from contextlib import contextmanager
from decimal import Decimal, InvalidOperation, ROUND_CEILING, ROUND_HALF_UP
from pathlib import Path
from urllib import error as urllib_error, request as urllib_request

from PySide6.QtCore import QEvent, QObject, QPoint, QTimer, Qt
from PySide6.QtGui import QAction, QColor, QFont, QKeyEvent, QPalette, QPainter
from PySide6.QtWidgets import (
    QApplication,
    QAbstractItemView,
    QCheckBox,
    QComboBox,
    QCompleter,
    QDialog,
    QDialogButtonBox,
    QFileDialog,
    QFormLayout,
    QHeaderView,
    QLabel,
    QLineEdit,
    QMainWindow,
    QMenu,
    QMessageBox,
    QPushButton,
    QProgressDialog,
    QSizePolicy,
    QStyle,
    QStyleFactory,
    QStyleOptionHeader,
    QStyledItemDelegate,
    QTableWidget,
    QTableWidgetItem,
    QTextEdit,
    QVBoxLayout,
    QWidget,
)

from ui_project_generated import Ui_MainWindow

try:
    import pymysql
except ImportError:  # pragma: no cover - fallback when PyMySQL is unavailable
    pymysql = None

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font
except ImportError:  # pragma: no cover - fallback when openpyxl is unavailable
    load_workbook = None
    Alignment = None
    Font = None


TABLE_HEADERS = [
    "製作項目",
    "寬度",
    "x",
    "長度",
    "數量",
    "材質",
    "冷裱加工",
    "板材種類",
    "板材厚度",
    "其他備料",
    "數量",
    "才數",
    "單價",
    "計價",
    "備料計價",
]

# Excel / 工具表格感：寬欄給描述類欄位，中欄給數值，x 欄極窄固定顯示。
TABLE_COLUMN_WIDTHS = [196, 82, 36, 82, 68, 140, 136, 126, 108, 144, 68, 84, 88, 92, 104]
X_COLUMN_INDEX = 2
LENGTH_COLUMN_INDEX = 3
TABLE_LOCKED_COLUMN_INDEXES = {11, 13}
MIDDLE_TAB_COLUMN_ORDER = [0, 1, 3, 4, 5, 6, 7, 8, 9, 10, 12, 14]
TABLE_SKIP_FOCUS_COLUMN_INDEXES = {X_COLUMN_INDEX, *TABLE_LOCKED_COLUMN_INDEXES}
DEFAULT_LINE_ITEM_ROW_COUNT = 15
BLANK_LINE_ROW_TEMPLATE = ["", "", "x", "", "", "", "", "", "", "", "", "", "", "", ""]
NUMERIC_COLUMN_INDEXES = {1, 3, 4, 10, 11, 12, 13, 14}
SELECT_LIKE_COLUMN_INDEXES = {0, 5, 6, 7, 8, 9}
SUMMARY_LOCKED_FIELD_NAMES = {"le_productionAmount", "le_taxAmount", "le_totalAmount"}
COMBO_COLUMN_OPTIONS = {
    0: ["大圖輸出", "裱板施工", "立牌製作", "桌裙布置", "展場貼圖"],
    5: ["PVC 貼圖", "PP 相紙", "防水帆布", "單透布", "背膠海報"],
    6: ["冷裱", "冷裱 + 修邊", "不上膜", "雙面對裱", "包邊處理"],
    7: ["KT 板", "豪卡板", "塑鋁板", "珍珠板", "發泡板"],
    8: ["1mm", "2mm", "3mm", "5mm", "10mm"],
    9: ["無", "黑膠帶收邊", "腳架孔位", "魔鬼氈", "補強條"],
}
COMBO_COLUMN_GROUPS = {
    0: "production_item",
    5: "material",
    6: "lamination",
    7: "board_type",
    8: "board_thickness",
    9: "extra_material",
}
DB_V2_CONFIG = {
    "host": os.environ.get("WORK_ORDER_V2_DB_HOST", "127.0.0.1"),
    "port": int(os.environ.get("WORK_ORDER_V2_DB_PORT", "3308")),
    "user": os.environ.get("WORK_ORDER_V2_DB_USER", "workorder_v2"),
    "password": os.environ.get("WORK_ORDER_V2_DB_PASSWORD", "123456"),
    "database": os.environ.get("WORK_ORDER_V2_DB_NAME", "work_order_v2"),
    "charset": "utf8mb4",
    "cursorclass": pymysql.cursors.DictCursor if pymysql else None,
}
HEADER_YELLOW_COLUMNS = range(0, 11)
HEADER_PURPLE_COLUMNS = range(11, len(TABLE_HEADERS))
HEADER_YELLOW_BG = QColor("#f6e7a6")
HEADER_YELLOW_FG = QColor("#3f3110")
HEADER_PURPLE_BG = QColor("#c9b1ea")
HEADER_PURPLE_FG = QColor("#31204a")
SUMMARY_LABEL_BG = "#e8dcf8"
SUMMARY_LABEL_FG = "#4b3566"
SUMMARY_LABEL_BORDER = "#c9b1ea"
HEADER_DEFAULT_BG = QColor("#ececec")
HEADER_DEFAULT_FG = QColor("#202124")
HEADER_BORDER_TOP = QColor("#c4c8cc")
HEADER_BORDER_BOTTOM = QColor("#adb3b9")


def resolve_build_commit_hash() -> str:
    repo_root = Path(__file__).resolve().parents[2]
    try:
        return subprocess.check_output(
            ["git", "-C", str(repo_root), "rev-parse", "--short", "HEAD"],
            text=True,
            stderr=subprocess.DEVNULL,
        ).strip() or "unknown"
    except Exception:
        return "unknown"


BUILD_COMMIT_HASH = resolve_build_commit_hash()
BUILD_LABEL_TEXT = f"build: {BUILD_COMMIT_HASH}"


BILLING_TEMPLATE_FILENAME = "excel_payment.xlsx"
BILLING_OUTPUT_FOLDER_NAME = "報價留底"
INPUT_FONT_FAMILY = "新細明體"
TOP_FORM_FONT_POINT_SIZE = 16
TABLE_PRODUCTION_FONT_POINT_SIZE = 20
TABLE_TEXT_FONT_POINT_SIZE = 18
TABLE_NUMBER_FONT_POINT_SIZE = 16
TABLE_HEADER_PRODUCTION_FONT_POINT_SIZE = 20
TABLE_HEADER_TEXT_FONT_POINT_SIZE = 17
TABLE_HEADER_NUMBER_FONT_POINT_SIZE = 15
BILLING_MAX_ROWS_PER_PAGE = 15
BILLING_DETAIL_START_ROW = 7
BILLING_TOTAL_ROW = 22
PAGE_LABELS = [
    "第一頁",
    "第二頁",
    "第三頁",
    "第四頁",
    "第五頁",
    "第六頁",
    "第七頁",
    "第八頁",
    "第九頁",
    "第十頁",
]
IMPORT_ROOT_HINT = r"\\New-super-1682\1682輸出"
IMPORT_IMAGE_SUFFIXES = {".jpg", ".jpeg"}
SIZE_TEXT_PATTERN = re.compile(r"^\d+(?:\.\d+)?(?:mm|cm)?[xX]\d+(?:\.\d+)?(?:mm|cm)?$", re.IGNORECASE)
PAPER_SIZE_PATTERN = re.compile(r"^A\d+$", re.IGNORECASE)
THICKNESS_PATTERN = re.compile(r"^\d+(?:\.\d+)?(?:mm|cm)$", re.IGNORECASE)
INLINE_WH_SIZE_PATTERN = re.compile(
    r"(^|[_\-\s]+)w\s*(\d+(?:\.\d+)?)(mm|cm)?\s*[_\-\s]*h\s*(\d+(?:\.\d+)?)(mm|cm)?(?=$|[_\-\s]+)",
    re.IGNORECASE,
)
CASE_PREFIX_PATTERN = re.compile(r"^\d+-")
QUANTITY_TOKEN_PATTERN = re.compile(r"^[xX](\d+)$")

MATERIAL_ALIASES = {
    "PVC": "PVC",
    "PP": "PP",
}
BOARD_TYPE_ALIASES = {
    "H": "合成板",
    "合成板": "合成板",
    "合成版": "合成板",
    "黑色H": "黑色合成板",
    "黑色合成板": "黑色合成板",
    "黑色合成版": "黑色合成板",
    "F": "發泡板",
    "發泡板": "發泡板",
    "W": "瓦楞板",
    "瓦楞板": "瓦楞板",
    "S": "厚磅紙板",
    "厚磅紙板": "厚磅紙板",
}
BOARD_TYPE_DEFAULT_THICKNESS = {
    "合成板": "5mm",
    "黑色合成板": "5mm",
    "發泡板": "2mm",
    "瓦楞板": "5mm",
    "厚磅紙板": "",
}
IMPORT_AI_MODEL = os.environ.get("WORK_ORDER_IMPORT_AI_MODEL", "deepseek/deepseek-v4-flash")
OPENROUTER_API_URL = os.environ.get("OPENROUTER_API_URL", "https://openrouter.ai/api/v1/chat/completions")
DEEPSEEK_API_URL = os.environ.get("DEEPSEEK_API_URL", "https://api.deepseek.com/chat/completions")
IMPORT_AI_TIMEOUT_SECONDS = int(os.environ.get("WORK_ORDER_IMPORT_AI_TIMEOUT_SECONDS", "45"))


def _connect_kwargs() -> dict:
    return {key: value for key, value in DB_V2_CONFIG.items() if value is not None}


def build_input_font(point_size: int) -> QFont:
    font = QFont(INPUT_FONT_FAMILY, point_size)
    font.setStyleHint(QFont.StyleHint.SansSerif, QFont.StyleStrategy.PreferAntialias)
    return font


def build_table_cell_font(column: int) -> QFont:
    if column == 0:
        return build_input_font(TABLE_PRODUCTION_FONT_POINT_SIZE)
    if column in SELECT_LIKE_COLUMN_INDEXES:
        return build_input_font(TABLE_TEXT_FONT_POINT_SIZE)
    return build_input_font(TABLE_NUMBER_FONT_POINT_SIZE)


def build_table_header_font(column: int) -> QFont:
    if column == 0:
        return build_input_font(TABLE_HEADER_PRODUCTION_FONT_POINT_SIZE)
    if column in SELECT_LIKE_COLUMN_INDEXES:
        return build_input_font(TABLE_HEADER_TEXT_FONT_POINT_SIZE)
    return build_input_font(TABLE_HEADER_NUMBER_FONT_POINT_SIZE)


def resolve_desktop_dir() -> Path:
    candidates = []
    one_drive = os.environ.get("OneDrive")
    user_profile = os.environ.get("USERPROFILE")
    if one_drive:
        candidates.append(Path(one_drive) / "Desktop")
    if user_profile:
        candidates.append(Path(user_profile) / "Desktop")
    candidates.append(Path.home() / "Desktop")

    for candidate in candidates:
        if candidate.exists() and candidate.is_dir():
            return candidate
    return candidates[0]


def resolve_billing_output_dir() -> Path:
    return resolve_desktop_dir() / BILLING_OUTPUT_FOLDER_NAME


def resolve_billing_template_path() -> Path:
    base_dir = Path(getattr(sys, "_MEIPASS", Path(__file__).resolve().parent))
    candidates = [
        Path(sys.executable).resolve().parent / BILLING_TEMPLATE_FILENAME if getattr(sys, "frozen", False) else None,
        base_dir / BILLING_TEMPLATE_FILENAME,
        Path(__file__).resolve().parent / BILLING_TEMPLATE_FILENAME,
        Path.cwd() / BILLING_TEMPLATE_FILENAME,
    ]
    for candidate in candidates:
        if candidate is not None and candidate.exists() and candidate.is_file():
            return candidate
    return (Path(sys.executable).resolve().parent if getattr(sys, "frozen", False) else Path(__file__).resolve().parent) / BILLING_TEMPLATE_FILENAME


def load_clients_from_v2() -> list[dict[str, str | int | None]]:
    if pymysql is None:
        return []

    query = """
        SELECT id, short_name, full_name, phone, address, tax_id
        FROM clients
        WHERE is_active = 1
        ORDER BY short_name, id
    """
    try:
        with pymysql.connect(**_connect_kwargs()) as conn:
            with conn.cursor() as cur:
                cur.execute(query)
                rows = cur.fetchall()
    except Exception:
        return []

    normalized_rows: list[dict[str, str | int | None]] = []
    for row in rows:
        short_name = (row.get("short_name") or "").strip()
        if not short_name:
            continue
        normalized_rows.append(
            {
                "id": row.get("id"),
                "short_name": short_name,
                "full_name": (row.get("full_name") or "").strip() or None,
                "phone": (row.get("phone") or "").strip() or None,
                "address": (row.get("address") or "").strip() or None,
                "tax_id": (row.get("tax_id") or "").strip() or None,
            }
        )
    return normalized_rows


def load_combo_options_from_v2() -> dict[int, list[str]]:
    if pymysql is None:
        return {column: list(options) for column, options in COMBO_COLUMN_OPTIONS.items()}

    query = """
        SELECT option_group, item_name
        FROM option_items
        WHERE is_active = 1
          AND option_group IN (%s, %s, %s, %s, %s, %s)
        ORDER BY option_group, sort_order, id
    """
    options_by_group = {group: [] for group in COMBO_COLUMN_GROUPS.values()}

    try:
        with pymysql.connect(**_connect_kwargs()) as conn:
            with conn.cursor() as cur:
                cur.execute(query, tuple(COMBO_COLUMN_GROUPS.values()))
                for row in cur.fetchall():
                    group = row["option_group"]
                    item_name = (row["item_name"] or "").strip()
                    if item_name:
                        options_by_group.setdefault(group, []).append(item_name)
    except Exception:
        return {column: list(options) for column, options in COMBO_COLUMN_OPTIONS.items()}

    loaded_options: dict[int, list[str]] = {}
    for column, group in COMBO_COLUMN_GROUPS.items():
        loaded_options[column] = options_by_group.get(group) or list(COMBO_COLUMN_OPTIONS[column])
    return loaded_options


class BandHeaderView(QHeaderView):
    def __init__(self, orientation, parent=None) -> None:
        super().__init__(orientation, parent)
        self.setDefaultAlignment(Qt.AlignmentFlag.AlignCenter)
        self.setHighlightSections(False)

    def paintSection(self, painter: QPainter, rect, logical_index: int) -> None:
        if not rect.isValid():
            return

        painter.save()

        if logical_index in HEADER_YELLOW_COLUMNS:
            bg = HEADER_YELLOW_BG
            fg = HEADER_YELLOW_FG
        elif logical_index in HEADER_PURPLE_COLUMNS:
            bg = HEADER_PURPLE_BG
            fg = HEADER_PURPLE_FG
        else:
            bg = HEADER_DEFAULT_BG
            fg = HEADER_DEFAULT_FG

        painter.fillRect(rect, bg)
        painter.setPen(HEADER_BORDER_TOP)
        painter.drawLine(rect.topLeft(), rect.topRight())
        painter.drawLine(rect.topLeft(), rect.bottomLeft())
        painter.drawLine(rect.topRight(), rect.bottomRight())
        painter.setPen(HEADER_BORDER_BOTTOM)
        painter.drawLine(rect.bottomLeft(), rect.bottomRight())

        option = QStyleOptionHeader()
        self.initStyleOption(option)
        option.rect = rect.adjusted(4, 0, -4, 0)
        option.section = logical_index
        option.text = self.model().headerData(logical_index, self.orientation(), Qt.ItemDataRole.DisplayRole) or ""
        option.state &= ~QStyle.StateFlag.State_Sunken
        painter.setPen(fg)
        painter.drawText(option.rect, int(Qt.AlignmentFlag.AlignCenter), option.text)

        painter.restore()


def apply_light_preview_theme(app: QApplication) -> None:
    app.setStyle(QStyleFactory.create("Fusion"))

    palette = QPalette()
    palette.setColor(QPalette.ColorRole.Window, QColor("#f5f5f5"))
    palette.setColor(QPalette.ColorRole.WindowText, QColor("#202124"))
    palette.setColor(QPalette.ColorRole.Base, QColor("#ffffff"))
    palette.setColor(QPalette.ColorRole.AlternateBase, QColor("#f7f7f7"))
    palette.setColor(QPalette.ColorRole.ToolTipBase, QColor("#ffffff"))
    palette.setColor(QPalette.ColorRole.ToolTipText, QColor("#202124"))
    palette.setColor(QPalette.ColorRole.Text, QColor("#202124"))
    palette.setColor(QPalette.ColorRole.Button, QColor("#ffffff"))
    palette.setColor(QPalette.ColorRole.ButtonText, QColor("#202124"))
    palette.setColor(QPalette.ColorRole.BrightText, QColor("#ffffff"))
    palette.setColor(QPalette.ColorRole.Highlight, QColor("#4c8bf5"))
    palette.setColor(QPalette.ColorRole.HighlightedText, QColor("#ffffff"))
    palette.setColor(QPalette.ColorRole.Light, QColor("#ffffff"))
    palette.setColor(QPalette.ColorRole.Midlight, QColor("#e6e6e6"))
    palette.setColor(QPalette.ColorRole.Mid, QColor("#cfcfcf"))
    palette.setColor(QPalette.ColorRole.Dark, QColor("#9e9e9e"))
    palette.setColor(QPalette.ColorRole.Shadow, QColor("#808080"))
    app.setPalette(palette)

    app.setStyleSheet(
        """
        QWidget {
            background-color: #f5f5f5;
            color: #202124;
        }
        QLineEdit, QTextEdit, QComboBox, QTableWidget {
            background-color: #ffffff;
            color: #202124;
            border: 1px solid #cfcfcf;
        }
        QTableWidget {
            gridline-color: #bfc5cc;
            alternate-background-color: #fafafa;
            selection-background-color: #dbe7ff;
            selection-color: #202124;
        }
        QHeaderView::section {
            background-color: #efefef;
            color: #202124;
            border: 1px solid #d8d8d8;
            padding: 4px;
        }
        QPushButton {
            background-color: #ffffff;
            border: 1px solid #c5c5c5;
            padding: 4px 10px;
        }
        QPushButton:hover {
            background-color: #f0f0f0;
        }
        QGroupBox {
            border: 1px solid #d0d0d0;
            margin-top: 10px;
        }
        QGroupBox::title {
            subcontrol-origin: margin;
            left: 10px;
            padding: 0 4px;
        }
        """
    )


class TableCellTabNavigator(QObject):
    def __init__(self, table: QTableWidget, parent_window: "GeneratedUiPreviewWindow | None" = None) -> None:
        super().__init__(table)
        self.table = table
        self.parent_window = parent_window
        self.column_order = [column for column in MIDDLE_TAB_COLUMN_ORDER if column < table.columnCount()]

    def redirect_x_cell(self, row: int) -> bool:
        if row < 0 or row >= self.table.rowCount():
            return False
        if not self._is_focusable_cell(row, LENGTH_COLUMN_INDEX):
            return False

        self.table.setCurrentCell(row, LENGTH_COLUMN_INDEX)
        QTimer.singleShot(0, lambda: self._focus_cell(row, LENGTH_COLUMN_INDEX))
        return True

    def register_widget(self, widget: QWidget, row: int, column: int) -> None:
        widget.setProperty("table_row", row)
        widget.setProperty("table_column", column)
        widget.installEventFilter(self)

        if widget.focusPolicy() == Qt.FocusPolicy.NoFocus and column not in TABLE_SKIP_FOCUS_COLUMN_INDEXES:
            widget.setFocusPolicy(Qt.FocusPolicy.StrongFocus)

        if isinstance(widget, QComboBox):
            widget.setFocusPolicy(Qt.FocusPolicy.StrongFocus)
            line_edit = widget.lineEdit()
            if line_edit is not None:
                line_edit.setFocusPolicy(Qt.FocusPolicy.StrongFocus)
                self.register_widget(line_edit, row, column)

    def eventFilter(self, watched: QObject, event: QEvent) -> bool:
        if event.type() != QEvent.Type.KeyPress or not isinstance(event, QKeyEvent):
            return super().eventFilter(watched, event)

        if event.key() == Qt.Key.Key_Backtab:
            return self._focus_relative_cell(watched, forward=False)
        if event.key() == Qt.Key.Key_Tab:
            forward = not bool(event.modifiers() & Qt.KeyboardModifier.ShiftModifier)
            return self._focus_relative_cell(watched, forward=forward)

        return super().eventFilter(watched, event)

    def _focus_relative_cell(self, watched: QObject, *, forward: bool) -> bool:
        row = watched.property("table_row")
        column = watched.property("table_column")
        if row is None or column is None:
            return False

        row = int(row)
        column = int(column)
        if column == X_COLUMN_INDEX:
            return self.redirect_x_cell(row)

        next_cell = self._find_next_editable_cell(row, column, forward=forward)
        if next_cell is None:
            if self.parent_window is not None:
                self.parent_window.focus_after_table(forward=forward)
                return True
            return False

        next_row, next_column = next_cell
        self.table.setCurrentCell(next_row, next_column)
        QTimer.singleShot(0, lambda: self._focus_cell(next_row, next_column))
        return True

    def _find_next_editable_cell(self, row: int, column: int, *, forward: bool) -> tuple[int, int] | None:
        row_count = self.table.rowCount()
        if row_count <= 0 or not self.column_order:
            return None

        if forward:
            return self._find_next_cell_forward(row, column)
        return self._find_next_cell_backward(row, column)

    def _find_next_cell_forward(self, row: int, column: int) -> tuple[int, int] | None:
        start_row = max(row, 0)
        if start_row >= self.table.rowCount():
            return None

        current_position = self._column_position(column)
        for row_index in range(start_row, self.table.rowCount()):
            columns = self.column_order
            if row_index == start_row and current_position is not None:
                columns = self.column_order[current_position + 1 :]
            for candidate_column in columns:
                if self._is_focusable_cell(row_index, candidate_column):
                    return row_index, candidate_column
            current_position = None
        return None

    def _find_next_cell_backward(self, row: int, column: int) -> tuple[int, int] | None:
        start_row = min(row, self.table.rowCount() - 1)
        if start_row < 0:
            return None

        current_position = self._column_position(column)
        for row_index in range(start_row, -1, -1):
            columns = list(reversed(self.column_order))
            if row_index == start_row and current_position is not None:
                columns = list(reversed(self.column_order[:current_position]))
            for candidate_column in columns:
                if self._is_focusable_cell(row_index, candidate_column):
                    return row_index, candidate_column
            current_position = None
        return None

    def _column_position(self, column: int) -> int | None:
        try:
            return self.column_order.index(column)
        except ValueError:
            return None

    def _is_focusable_cell(self, row: int, column: int) -> bool:
        if column in TABLE_SKIP_FOCUS_COLUMN_INDEXES:
            return False

        cell_widget = self.table.cellWidget(row, column)
        if cell_widget is not None:
            return cell_widget.isEnabled() and cell_widget.focusPolicy() != Qt.FocusPolicy.NoFocus

        item = self.table.item(row, column)
        if item is None:
            return False
        flags = item.flags()
        return bool((flags & Qt.ItemFlag.ItemIsEnabled) and (flags & Qt.ItemFlag.ItemIsEditable))

    def _focus_cell(self, row: int, column: int) -> None:
        cell_widget = self.table.cellWidget(row, column)
        if cell_widget is not None:
            focus_target = self._preferred_focus_target(cell_widget)
            if focus_target is not None:
                focus_target.setFocus(Qt.FocusReason.TabFocusReason)
            return

        self.table.editItem(self.table.item(row, column))

    def _preferred_focus_target(self, widget: QWidget) -> QWidget | None:
        if isinstance(widget, QComboBox):
            line_edit = widget.lineEdit()
            if line_edit is not None:
                line_edit.selectAll()
            # Table 內 editable QComboBox 若直接把焦點代理到內部 lineEdit，
            # Qt 會把整個 combo 視為已經「進入又離開」，導致下一次 Tab
            # 直接跳過後續同列的 combo 欄位。這裡必須停在 combo 本體，
            # 才能讓材質/冷裱加工/板材種類/板材厚度/其他備料都成為可停留點。
            return widget

        if isinstance(widget, QLineEdit):
            widget.selectAll()
            return widget

        focus_widget = widget.focusProxy()
        if isinstance(focus_widget, QWidget):
            return focus_widget
        if widget.focusPolicy() != Qt.FocusPolicy.NoFocus:
            return widget
        return widget.findChild(QWidget)


class TableItemDelegate(QStyledItemDelegate):
    def __init__(self, navigator: TableCellTabNavigator, parent: QTableWidget) -> None:
        super().__init__(parent)
        self.navigator = navigator

    def createEditor(self, parent: QWidget, option, index):
        if not index.isValid() or index.column() in TABLE_SKIP_FOCUS_COLUMN_INDEXES:
            return None

        editor = super().createEditor(parent, option, index)
        if isinstance(editor, QWidget):
            self.navigator.register_widget(editor, index.row(), index.column())
        return editor


def normalize_line_edit_value(widget: object) -> str | None:
    if isinstance(widget, QLineEdit):
        value = widget.text().strip()
        return value or None
    return None


def normalize_text_edit_value(widget: object) -> str | None:
    value = widget.toPlainText().strip() if widget is not None and hasattr(widget, "toPlainText") else ""
    return value or None


def parse_decimal_or_none(raw_value: str | None) -> Decimal | None:
    if raw_value is None:
        return None
    normalized = raw_value.strip().replace(",", "")
    if not normalized:
        return None
    try:
        return Decimal(normalized)
    except (InvalidOperation, ValueError) as exc:
        raise ValueError(f"數值格式錯誤：{raw_value}") from exc


def parse_int_or_none(raw_value: str | None) -> int | None:
    decimal_value = parse_decimal_or_none(raw_value)
    if decimal_value is None:
        return None
    if decimal_value != decimal_value.to_integral_value():
        raise ValueError(f"整數欄位不可輸入小數：{raw_value}")
    return int(decimal_value)


def extract_first_int(value: object) -> int | None:
    text = unicodedata.normalize("NFKC", str(value or "")).strip()
    if not text:
        return None
    direct = parse_int_or_none(text)
    if direct is not None:
        return direct
    match = re.search(r"\d+", text)
    return int(match.group(0)) if match else None


def format_decimal_for_ui(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, Decimal):
        normalized = format(value.normalize(), "f")
        return "0" if normalized == "-0" else normalized
    return str(value)


def ceil_decimal(value: Decimal) -> Decimal:
    return value.to_integral_value(rounding=ROUND_CEILING)


def round_decimal(value: Decimal) -> Decimal:
    return value.to_integral_value(rounding=ROUND_HALF_UP)


def load_option_item_ids_from_v2() -> dict[str, dict[str, int]]:
    lookup = {group: {} for group in COMBO_COLUMN_GROUPS.values()}
    if pymysql is None:
        return lookup

    query = """
        SELECT id, option_group, item_name
        FROM option_items
        WHERE is_active = 1
          AND option_group IN (%s, %s, %s, %s, %s, %s)
        ORDER BY option_group, sort_order, id
    """
    try:
        with pymysql.connect(**_connect_kwargs()) as conn:
            with conn.cursor() as cur:
                cur.execute(query, tuple(COMBO_COLUMN_GROUPS.values()))
                for row in cur.fetchall():
                    group = str(row["option_group"])
                    item_name = (row.get("item_name") or "").strip()
                    item_id = row.get("id")
                    if group and item_name and item_id is not None:
                        lookup.setdefault(group, {})[item_name] = int(item_id)
    except Exception:
        return lookup

    return lookup


def normalize_token_text(value: str) -> str:
    normalized = unicodedata.normalize("NFKC", value or "")
    normalized = normalized.replace(" ", "").strip()
    return normalized.upper()


def split_path_parts(path_text: str) -> list[str]:
    return [part for part in re.split(r"[\\/]+", path_text) if part]


def looks_like_size_text(value: str) -> bool:
    candidate = unicodedata.normalize("NFKC", (value or "").strip())
    return bool(SIZE_TEXT_PATTERN.fullmatch(candidate) or PAPER_SIZE_PATTERN.fullmatch(candidate))


def looks_like_thickness_text(value: str) -> bool:
    candidate = unicodedata.normalize("NFKC", (value or "").strip())
    return bool(THICKNESS_PATTERN.fullmatch(candidate))


def format_import_dimension(value: str) -> str:
    number = Decimal(value)
    return format_decimal_for_ui(number)


def extract_inline_wh_size_from_import_stem(stem: str) -> tuple[str, str]:
    """Extract inline W/H size markers such as w99_h50cm from a filename stem."""
    normalized = unicodedata.normalize("NFKC", stem or "").strip()
    match = INLINE_WH_SIZE_PATTERN.search(normalized)
    if not match:
        return "", normalized

    width = format_import_dimension(match.group(2))
    height = format_import_dimension(match.group(4))
    unit = (match.group(3) or match.group(5) or "").lower()
    size_text = f"{width}X{height}{unit}"
    remaining = (normalized[: match.start()] + " " + normalized[match.end() :]).strip(" _-")
    remaining = re.sub(r"[ _-]+$", "", remaining).strip()
    return size_text, remaining


def extract_inline_material_from_import_text(text: str) -> tuple[str, str]:
    """Extract material aliases that are glued to the production item, e.g. AC檯面PVC."""
    remaining = unicodedata.normalize("NFKC", text or "").strip()
    for alias, material in sorted(MATERIAL_ALIASES.items(), key=lambda item: len(item[0]), reverse=True):
        pattern = re.compile(rf"(?<![A-Za-z0-9]){re.escape(alias)}(?![A-Za-z0-9])", re.IGNORECASE)
        match = pattern.search(remaining)
        if match:
            remaining = (remaining[: match.start()] + " " + remaining[match.end() :]).strip(" _-")
            remaining = re.sub(r"\s{2,}", " ", remaining).strip()
            return material, remaining
    return "", remaining


def extract_dimension_token_number(value: str) -> Decimal | None:
    candidate = unicodedata.normalize("NFKC", (value or "").strip())
    if not candidate:
        return None
    match = re.fullmatch(r"(\d+(?:\.\d+)?)(?:mm|cm)?", candidate, re.IGNORECASE)
    if not match:
        return None
    return Decimal(match.group(1))


def split_size_text_for_table(size_text: str) -> tuple[str, str]:
    candidate = unicodedata.normalize("NFKC", (size_text or "").strip())
    if not SIZE_TEXT_PATTERN.fullmatch(candidate):
        return (candidate, "") if candidate else ("", "")

    left_raw, right_raw = re.split(r"[xX]", candidate, maxsplit=1)
    left_value = extract_dimension_token_number(left_raw)
    right_value = extract_dimension_token_number(right_raw)
    if left_value is None or right_value is None:
        return candidate, ""
    return format_decimal_for_ui(left_value), format_decimal_for_ui(right_value)


def strip_case_prefix(case_folder: str) -> str:
    normalized = unicodedata.normalize("NFKC", case_folder or "").strip()
    return CASE_PREFIX_PATTERN.sub("", normalized).strip() or normalized


def parse_import_path_context(folder_path: str) -> dict[str, str]:
    parts = split_path_parts(folder_path)
    customer_name = ""
    case_folder = Path(folder_path).name
    if "1682輸出" in parts:
        idx = parts.index("1682輸出")
        tail = parts[idx + 1 :]
        if len(tail) >= 2 and tail[0] == "000客戶":
            customer_name = tail[1]
            if len(tail) >= 3:
                case_folder = tail[2]
        elif len(tail) >= 2:
            customer_name = tail[0]
            case_folder = tail[1]
    return {
        "source_folder": folder_path,
        "customer_name": customer_name.strip(),
        "case_folder": case_folder.strip(),
        "case_name": strip_case_prefix(case_folder),
    }


def compact_compare_text(value: str) -> str:
    normalized = normalize_token_text(value)
    for noise in ("板", "版", "型", "鐵", "角", "架"):
        normalized = normalized.replace(noise, "")
    return normalized


def match_candidate(token: str, candidates: list[str]) -> tuple[str | None, float]:
    normalized_token = compact_compare_text(token)
    if not normalized_token:
        return None, 0.0
    best_name = None
    best_score = 0.0
    for candidate in candidates:
        normalized_candidate = compact_compare_text(candidate)
        if not normalized_candidate:
            continue
        if normalized_token == normalized_candidate:
            return candidate, 1.0
        if normalized_token in normalized_candidate or normalized_candidate in normalized_token:
            score = 0.88
        else:
            score = difflib.SequenceMatcher(None, normalized_token, normalized_candidate).ratio()
        if score > best_score:
            best_score = score
            best_name = candidate
    return best_name, best_score


def extract_json_text(raw_text: str) -> str:
    cleaned = raw_text.strip()
    if cleaned.startswith("```"):
        cleaned = re.sub(r"^```(?:json)?", "", cleaned).strip()
        if cleaned.endswith("```"):
            cleaned = cleaned[:-3].strip()
    return cleaned


def ensure_str_list(value: object) -> list[str]:
    if value is None:
        return []
    if isinstance(value, str):
        text = value.strip()
        return [text] if text else []
    if isinstance(value, (list, tuple, set)):
        items: list[str] = []
        for item in value:
            text = str(item).strip()
            if text:
                items.append(text)
        return items
    text = str(value).strip()
    return [text] if text else []


def deepseek_configured() -> bool:
    return bool(os.environ.get("OPENROUTER_API_KEY") or os.environ.get("DEEPSEEK_API_KEY"))


def default_import_lamination(customer_name: str) -> str:
    return "亮面" if (customer_name or "").strip() == "冠銘專用" else "霧面"


class ComboBoxWheelBlocker(QObject):
    def __init__(self, combo: QComboBox) -> None:
        super().__init__(combo)
        self.combo = combo

    def eventFilter(self, watched: QObject, event: QEvent) -> bool:
        if event.type() == QEvent.Type.Wheel and not self.combo.view().isVisible():
            event.ignore()
            return True
        return super().eventFilter(watched, event)


def disable_combobox_mouse_wheel(combo: QComboBox) -> None:
    blocker = ComboBoxWheelBlocker(combo)
    combo.installEventFilter(blocker)
    combo._wheel_blocker = blocker  # type: ignore[attr-defined]


def configure_combo_autocomplete(combo: QComboBox) -> None:
    combo.setEditable(True)
    combo.setInsertPolicy(QComboBox.InsertPolicy.NoInsert)
    completer = QCompleter(combo.model(), combo)
    completer.setCaseSensitivity(Qt.CaseSensitivity.CaseInsensitive)
    completer.setFilterMode(Qt.MatchFlag.MatchContains)
    completer.setCompletionMode(QCompleter.CompletionMode.PopupCompletion)
    combo.setCompleter(completer)


class ImportHeaderConfirmDialog(QDialog):
    def __init__(self, folder_path: str, header_payload: dict[str, str], parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self.setWindowTitle("導入：確認資料夾與表頭")
        self.setModal(True)
        self.resize(760, 0)
        self.folder_path = folder_path

        layout = QVBoxLayout(self)
        form = QFormLayout()
        form.setFieldGrowthPolicy(QFormLayout.FieldGrowthPolicy.ExpandingFieldsGrow)
        layout.addLayout(form)

        self.folder_edit = QLineEdit(folder_path, self)
        self.folder_edit.setReadOnly(True)
        self.customer_edit = QLineEdit(header_payload.get("customer_name", ""), self)
        self.case_name_edit = QLineEdit(header_payload.get("case_name", ""), self)
        self.work_number_edit = QLineEdit(header_payload.get("work_number", ""), self)

        form.addRow("資料夾路徑", self.folder_edit)
        form.addRow("客戶名稱", self.customer_edit)
        form.addRow("案件名稱", self.case_name_edit)
        form.addRow("工單編號", self.work_number_edit)

        buttons = QDialogButtonBox(parent=self)
        self.reselect_button = buttons.addButton("重新選擇資料夾", QDialogButtonBox.ButtonRole.ActionRole)
        self.confirm_button = buttons.addButton("確認", QDialogButtonBox.ButtonRole.AcceptRole)
        self.cancel_button = buttons.addButton("取消", QDialogButtonBox.ButtonRole.RejectRole)
        self.reselect_button.clicked.connect(self._handle_reselect_clicked)
        self.confirm_button.clicked.connect(self._handle_accept_clicked)
        self.cancel_button.clicked.connect(self.reject)
        layout.addWidget(buttons)

    def _handle_reselect_clicked(self) -> None:
        start_dir = self.folder_path or IMPORT_ROOT_HINT
        selected = QFileDialog.getExistingDirectory(self, "選擇導入資料夾", start_dir)
        if not selected:
            return
        self.folder_path = selected
        self.folder_edit.setText(selected)
        parsed = parse_import_path_context(selected)
        self.customer_edit.setText(parsed.get("customer_name", ""))
        self.case_name_edit.setText(parsed.get("case_name", ""))

    def _handle_accept_clicked(self) -> None:
        if not self.folder_edit.text().strip():
            QMessageBox.warning(self, "資料不足", "請先選擇資料夾。")
            return
        if not self.customer_edit.text().strip():
            QMessageBox.warning(self, "資料不足", "客戶名稱不可為空。")
            return
        if not self.case_name_edit.text().strip():
            QMessageBox.warning(self, "資料不足", "案件名稱不可為空。")
            return
        self.accept()

    def payload(self) -> dict[str, str]:
        return {
            "source_folder": self.folder_edit.text().strip(),
            "customer_name": self.customer_edit.text().strip(),
            "case_name": self.case_name_edit.text().strip(),
            "work_number": self.work_number_edit.text().strip(),
        }


class ImportExtraMaterialsDialog(QDialog):
    def __init__(
        self,
        options: list[str],
        selected_values: list[str] | None = None,
        unresolved_tokens: list[str] | None = None,
        parent: QWidget | None = None,
    ) -> None:
        super().__init__(parent)
        self.setWindowTitle("編輯其他備料")
        self.setModal(True)
        self.resize(420, 0)
        self._checkboxes: list[QCheckBox] = []
        selected = {value.strip() for value in (selected_values or []) if value and value.strip()}
        unresolved = [value.strip() for value in (unresolved_tokens or []) if value and value.strip()]

        layout = QVBoxLayout(self)
        if unresolved:
            hint = QLabel(f"未解 token：{'+'.join(unresolved)}", self)
            hint.setWordWrap(True)
            layout.addWidget(hint)

        for option in options:
            checkbox = QCheckBox(option, self)
            checkbox.setChecked(option in selected)
            layout.addWidget(checkbox)
            self._checkboxes.append(checkbox)

        custom_values = [value for value in selected if value not in set(options)]
        self.custom_edit = QLineEdit("+".join(custom_values), self)
        self.custom_edit.setPlaceholderText("其他自訂備料，用 + 串接")
        layout.addWidget(self.custom_edit)

        buttons = QDialogButtonBox(parent=self)
        self.confirm_button = buttons.addButton("確認", QDialogButtonBox.ButtonRole.AcceptRole)
        self.cancel_button = buttons.addButton("取消", QDialogButtonBox.ButtonRole.RejectRole)
        self.confirm_button.clicked.connect(self.accept)
        self.cancel_button.clicked.connect(self.reject)
        layout.addWidget(buttons)

    def payload(self) -> list[str]:
        values = [checkbox.text().strip() for checkbox in self._checkboxes if checkbox.isChecked() and checkbox.text().strip()]
        for part in self.custom_edit.text().split("+"):
            value = part.strip()
            if value and value not in values:
                values.append(value)
        return values


class ImportLineItemsConfirmDialog(QDialog):
    COLUMN_KEYS = [
        "production_item",
        "size_text",
        "material",
        "lamination",
        "board_type",
        "board_thickness",
        "extra_materials",
        "quantity",
        "unresolved_tokens",
        "confidence",
    ]
    COLUMN_LABELS = [
        "製作項目",
        "尺寸",
        "材質",
        "冷裱",
        "板材",
        "板厚",
        "其他備料",
        "數量",
        "未解 token",
        "信心",
    ]
    COMBO_COLUMN_KEYS = {
        2: "material",
        3: "lamination",
        4: "board_type",
        5: "board_thickness",
    }

    def __init__(self, lines: list[dict[str, object]], option_choices: dict[str, list[str]] | None = None, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self.setWindowTitle("導入：確認明細")
        self.setModal(True)
        self.resize(1180, 640)
        self._lines = lines
        self.option_choices = option_choices or {}
        self.go_back = False

        layout = QVBoxLayout(self)
        self.table = QTableWidget(len(lines), len(self.COLUMN_KEYS), self)
        self.table.setHorizontalHeaderLabels(self.COLUMN_LABELS)
        self.table.verticalHeader().setVisible(False)
        self.table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table.setEditTriggers(
            QAbstractItemView.EditTrigger.DoubleClicked
            | QAbstractItemView.EditTrigger.SelectedClicked
            | QAbstractItemView.EditTrigger.EditKeyPressed
        )
        header = self.table.horizontalHeader()
        if header is not None:
            header.setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
            header.setStretchLastSection(True)
        layout.addWidget(self.table)

        buttons = QDialogButtonBox(parent=self)
        self.back_button = buttons.addButton("返回上一步", QDialogButtonBox.ButtonRole.ActionRole)
        self.confirm_button = buttons.addButton("確認匯入", QDialogButtonBox.ButtonRole.AcceptRole)
        self.cancel_button = buttons.addButton("取消", QDialogButtonBox.ButtonRole.RejectRole)
        self.back_button.clicked.connect(self._handle_back_clicked)
        self.confirm_button.clicked.connect(self.accept)
        self.cancel_button.clicked.connect(self.reject)
        layout.addWidget(buttons)

        self.table.cellDoubleClicked.connect(self._handle_table_cell_double_clicked)
        self._populate_rows()

    def _make_combo_widget(self, key: str, value: str) -> QComboBox:
        combo = QComboBox(self.table)
        options = [option for option in self.option_choices.get(key, []) if option]
        for option in options:
            combo.addItem(option)
        configure_combo_autocomplete(combo)
        found_index = combo.findText(value)
        if found_index >= 0:
            combo.setCurrentIndex(found_index)
        else:
            combo.setEditText(value)
        return combo

    def _populate_rows(self) -> None:
        for row_index, line in enumerate(self._lines):
            values = {
                "production_item": str(line.get("production_item") or ""),
                "size_text": str(line.get("size_text") or ""),
                "material": str(line.get("material") or ""),
                "lamination": str(line.get("lamination") or ""),
                "board_type": str(line.get("board_type") or ""),
                "board_thickness": str(line.get("board_thickness") or ""),
                "extra_materials": "+".join(line.get("extra_materials") or []),
                "quantity": str(line.get("quantity") or 1),
                "unresolved_tokens": "+".join(line.get("unresolved_tokens") or []),
                "confidence": f"{float(line.get('confidence') or 0):.2f}",
            }
            for column, key in enumerate(self.COLUMN_KEYS):
                if column in self.COMBO_COLUMN_KEYS:
                    combo = self._make_combo_widget(self.COMBO_COLUMN_KEYS[column], values[key])
                    self.table.setCellWidget(row_index, column, combo)
                    self.table.setItem(row_index, column, QTableWidgetItem(values[key]))
                    continue
                item = QTableWidgetItem(values[key])
                if key in {"extra_materials", "unresolved_tokens", "confidence"}:
                    item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)
                self.table.setItem(row_index, column, item)

    def _handle_table_cell_double_clicked(self, row_index: int, column: int) -> None:
        if column != 6:
            return
        current_values = [part.strip() for part in self._cell_text(row_index, column).split("+") if part.strip()]
        unresolved_tokens = [part.strip() for part in self._cell_text(row_index, 8).split("+") if part.strip()]
        dialog = ImportExtraMaterialsDialog(
            list(self.option_choices.get("extra_materials", [])),
            current_values,
            unresolved_tokens,
            self,
        )
        if dialog.exec() != QDialog.DialogCode.Accepted:
            return
        item = self.table.item(row_index, column)
        if item is None:
            item = QTableWidgetItem()
            item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)
            self.table.setItem(row_index, column, item)
        item.setText("+".join(dialog.payload()))

    def _handle_back_clicked(self) -> None:
        self.go_back = True
        self.reject()

    def payload(self) -> list[dict[str, object]]:
        updated_lines: list[dict[str, object]] = []
        for row_index, source_line in enumerate(self._lines):
            line = dict(source_line)
            line["production_item"] = self._cell_text(row_index, 0)
            line["size_text"] = self._cell_text(row_index, 1)
            line["material"] = self._cell_text(row_index, 2)
            line["lamination"] = self._cell_text(row_index, 3)
            line["board_type"] = self._cell_text(row_index, 4)
            line["board_thickness"] = self._cell_text(row_index, 5)
            line["extra_materials"] = [part.strip() for part in self._cell_text(row_index, 6).split("+") if part.strip()]
            line["quantity"] = parse_int_or_none(self._cell_text(row_index, 7)) or 1
            updated_lines.append(line)
        return updated_lines

    def _cell_text(self, row_index: int, column: int) -> str:
        cell_widget = self.table.cellWidget(row_index, column)
        if isinstance(cell_widget, QComboBox):
            return cell_widget.currentText().strip()
        item = self.table.item(row_index, column)
        return item.text().strip() if item is not None else ""


class GeneratedUiPreviewWindow(QMainWindow):
    TOP_TAB_ORDER = [
        "le_worknum",
        "cb_customerName",
        "le_contactName",
        "le_startTime",
        "le_caseName",
        "le_phone",
        "lle_address",
        "le_endTime",
    ]
    BOTTOM_TAB_ORDER = [
        "te_remark",
        "btn_open",
        "btn_save",
        "btn_reset",
        "btn_billing",
        "btn_subtotal",
        "btn_calcuate",
        "btn_import",
        "btn_invoice",
        "btn_add_customer",
        "btn_delete_work_order",
    ]

    def __init__(self) -> None:
        super().__init__()
        self.combo_column_options = load_combo_options_from_v2()
        self.option_item_ids_by_group = load_option_item_ids_from_v2()
        self.option_item_names_by_group_and_id = self._invert_option_item_lookup(self.option_item_ids_by_group)
        self.clients = load_clients_from_v2()
        self.client_rows_by_short_name = {
            str(row["short_name"]): row for row in self.clients if row.get("short_name")
        }
        self._last_auto_filled_phone = ""
        self._last_auto_filled_address = ""
        self._loaded_work_order_number = ""
        self._import_review_rows: dict[int, dict[str, object]] = {}
        self.table_tab_navigator: TableCellTabNavigator | None = None
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)
        self.build_label: QLabel | None = None
        self._suspend_auto_append_checks = False
        self._inject_customer_management_buttons()
        self._configure_customer_name_combo()
        self._apply_input_font_defaults()
        self._tune_generated_layout()
        self._initialize_blank_work_order()
        self._configure_save_flow()
        self._configure_open_flow()
        self._configure_reset_flow()
        self._configure_calculation_actions()
        self._configure_billing_action()
        self._configure_import_action()
        self._configure_customer_management_actions()
        self._configure_delete_work_order_action()
        self._configure_line_row_actions()
        self._configure_focus_chain()

    def _apply_input_font_defaults(self) -> None:
        top_font = build_input_font(TOP_FORM_FONT_POINT_SIZE)
        for name in [
            "le_worknum",
            "cb_customerName",
            "le_contactName",
            "le_startTime",
            "le_caseName",
            "le_phone",
            "lle_address",
            "le_endTime",
            "te_remark",
        ]:
            widget = getattr(self.ui, name, None)
            if isinstance(widget, (QLineEdit, QComboBox, QTextEdit)):
                widget.setFont(top_font)
                if isinstance(widget, QComboBox):
                    if widget.lineEdit() is not None:
                        widget.lineEdit().setFont(top_font)
                    if widget.view() is not None:
                        widget.view().setFont(top_font)

        table = getattr(self.ui, "tbl_lineItems", None)
        if isinstance(table, QTableWidget):
            table.setFont(build_input_font(TABLE_NUMBER_FONT_POINT_SIZE))

    def _configure_customer_name_combo(self) -> None:
        combo = getattr(self.ui, "cb_customerName", None)
        if not isinstance(combo, QComboBox):
            return

        disable_combobox_mouse_wheel(combo)
        configure_combo_autocomplete(combo)
        combo.setMaxVisibleItems(8)
        combo.clear()
        for row in self.clients:
            combo.addItem(str(row["short_name"]), row)

        combo.currentTextChanged.connect(self._handle_customer_name_changed)

    def _handle_customer_name_changed(self, customer_name: str) -> None:
        row = self.client_rows_by_short_name.get(customer_name.strip())
        if row is None:
            return
        self._apply_customer_contact_defaults(row)

    def _inject_customer_management_buttons(self) -> None:
        lower_buttons_layout = getattr(self.ui, "horizontalLayout_15", None)
        button_parent = getattr(self.ui, "widget14", self)
        if lower_buttons_layout is None:
            return

        add_customer_button = QPushButton("新增客戶", button_parent)
        add_customer_button.setObjectName("btn_add_customer")
        lower_buttons_layout.addWidget(add_customer_button)
        self.ui.btn_add_customer = add_customer_button

        delete_work_order_button = QPushButton("刪除工單", button_parent)
        delete_work_order_button.setObjectName("btn_delete_work_order")
        lower_buttons_layout.addWidget(delete_work_order_button)
        self.ui.btn_delete_work_order = delete_work_order_button

    @staticmethod
    def _invert_option_item_lookup(option_item_ids_by_group: dict[str, dict[str, int]]) -> dict[str, dict[int, str]]:
        return {
            group: {item_id: item_name for item_name, item_id in items.items()}
            for group, items in option_item_ids_by_group.items()
        }

    def _configure_save_flow(self) -> None:
        save_button = getattr(self.ui, "btn_save", None)
        if save_button is not None:
            save_button.clicked.connect(self._handle_save_clicked)

    def _configure_open_flow(self) -> None:
        open_button = getattr(self.ui, "btn_open", None)
        if open_button is not None:
            open_button.clicked.connect(self._handle_open_clicked)

    def _configure_reset_flow(self) -> None:
        reset_button = getattr(self.ui, "btn_reset", None)
        if reset_button is not None:
            reset_button.clicked.connect(self._handle_reset_clicked)

    def _configure_calculation_actions(self) -> None:
        subtotal_button = getattr(self.ui, "btn_subtotal", None)
        if subtotal_button is not None:
            subtotal_button.clicked.connect(self._handle_subtotal_clicked)

        calculate_button = getattr(self.ui, "btn_calcuate", None)
        if calculate_button is not None:
            calculate_button.clicked.connect(self._handle_calculate_clicked)

    def _configure_billing_action(self) -> None:
        billing_button = getattr(self.ui, "btn_billing", None)
        if billing_button is not None:
            billing_button.clicked.connect(self._handle_billing_clicked)

    def _configure_import_action(self) -> None:
        import_button = getattr(self.ui, "btn_import", None)
        if import_button is not None:
            import_button.clicked.connect(self._handle_import_clicked)

    def _configure_customer_management_actions(self) -> None:
        add_customer_button = getattr(self.ui, "btn_add_customer", None)
        if add_customer_button is not None:
            add_customer_button.clicked.connect(self._handle_add_customer_clicked)

    def _configure_delete_work_order_action(self) -> None:
        delete_button = getattr(self.ui, "btn_delete_work_order", None)
        if delete_button is not None:
            delete_button.clicked.connect(self._handle_delete_work_order_clicked)

    def _configure_line_row_actions(self) -> None:
        table = getattr(self.ui, "tbl_lineItems", None)
        if table is None:
            return
        table.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        table.customContextMenuRequested.connect(self._show_line_row_context_menu)

    def _set_status_message(self, message: str, timeout_ms: int = 8000) -> None:
        self.statusBar().showMessage(message, timeout_ms)

    def _header_payload(self) -> dict[str, str | int | None]:
        customer_name = ""
        customer_combo = getattr(self.ui, "cb_customerName", None)
        if isinstance(customer_combo, QComboBox):
            customer_name = customer_combo.currentText().strip()

        client_row = self.client_rows_by_short_name.get(customer_name) if customer_name else None

        work_number = normalize_line_edit_value(getattr(self.ui, "le_worknum", None))
        if not work_number:
            raise ValueError("工單號不可為空。")

        return {
            "work_number": work_number,
            "case_name": normalize_line_edit_value(getattr(self.ui, "le_caseName", None)),
            "client_id": client_row.get("id") if client_row else None,
            "customer_name_text": customer_name or None,
            "company_phone": normalize_line_edit_value(getattr(self.ui, "le_phone", None)),
            "contact_name": normalize_line_edit_value(getattr(self.ui, "le_contactName", None)),
            "work_time": normalize_line_edit_value(getattr(self.ui, "le_startTime", None)),
            "cleanup_time": normalize_line_edit_value(getattr(self.ui, "le_endTime", None)),
            "work_address": normalize_line_edit_value(getattr(self.ui, "lle_address", None)),
            "production_amount": parse_decimal_or_none(normalize_line_edit_value(getattr(self.ui, "le_productionAmount", None))),
            "tax_amount": parse_decimal_or_none(normalize_line_edit_value(getattr(self.ui, "le_taxAmount", None))),
            "total_amount": parse_decimal_or_none(normalize_line_edit_value(getattr(self.ui, "le_totalAmount", None))),
            "remark": normalize_text_edit_value(getattr(self.ui, "te_remark", None)),
            "status": "draft",
        }

    def _table_cell_text(self, row: int, column: int) -> str:
        table = getattr(self.ui, "tbl_lineItems", None)
        if table is None:
            return ""

        cell_widget = table.cellWidget(row, column)
        if isinstance(cell_widget, QComboBox):
            return cell_widget.currentText().strip()
        if isinstance(cell_widget, QLineEdit):
            return cell_widget.text().strip()

        item = table.item(row, column)
        return item.text().strip() if item is not None else ""

    def _line_row_has_meaningful_data(self, row: int) -> bool:
        meaningful_columns = [0, 1, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14]
        for column in meaningful_columns:
            if self._table_cell_text(row, column):
                return True
        return False

    def _resolve_option_item_storage(self, option_group: str, item_name: str) -> tuple[int | None, str | None]:
        normalized_name = item_name.strip()
        if not normalized_name:
            return None, None

        item_id = self.option_item_ids_by_group.get(option_group, {}).get(normalized_name)
        if item_id is not None:
            return item_id, None
        return None, normalized_name

    def _parse_table_decimal(self, row: int, column: int, *, blank_as_zero: bool = True, invalid_as_zero: bool = True) -> Decimal:
        raw_value = self._table_cell_text(row, column)
        if not raw_value:
            return Decimal("0") if blank_as_zero else Decimal("0")
        try:
            return parse_decimal_or_none(raw_value) or Decimal("0")
        except ValueError:
            if invalid_as_zero:
                return Decimal("0")
            raise

    def _write_table_value(self, row: int, column: int, text: str) -> None:
        table = getattr(self.ui, "tbl_lineItems", None)
        if table is None:
            return

        cell_widget = table.cellWidget(row, column)
        if isinstance(cell_widget, QLineEdit):
            cell_widget.setText(text)
            return
        if isinstance(cell_widget, QComboBox):
            found_index = cell_widget.findText(text)
            if found_index >= 0:
                cell_widget.setCurrentIndex(found_index)
            else:
                cell_widget.setEditText(text)
            return

        item = table.item(row, column)
        if item is not None:
            item.setText(text)

    def _set_summary_field_value(self, field_name: str, value: Decimal) -> None:
        widget = getattr(self.ui, field_name, None)
        if isinstance(widget, QLineEdit):
            widget.setText(format_decimal_for_ui(value))

    def calculate_line_subtotals(self) -> dict[str, object]:
        table = getattr(self.ui, "tbl_lineItems", None)
        if table is None:
            return {"updated_rows": 0, "skipped_rows": 0, "rows": []}

        results: list[dict[str, str]] = []
        updated_rows = 0
        skipped_rows = 0
        for row in range(table.rowCount()):
            if not self._line_row_has_meaningful_data(row):
                skipped_rows += 1
                continue

            width = self._parse_table_decimal(row, 1)
            length = self._parse_table_decimal(row, 3)
            quantity = self._parse_table_decimal(row, 4)
            unit_price = self._parse_table_decimal(row, 12)

            if width <= 0 or length <= 0 or quantity <= 0:
                cbm = Decimal("0")
            else:
                per_item_cbm_raw = (width * length) / Decimal("900")
                per_item_cbm = per_item_cbm_raw.to_integral_value(rounding=ROUND_CEILING)
                if per_item_cbm < 1:
                    per_item_cbm = Decimal("1")
                cbm = per_item_cbm * quantity
            line_total = cbm * unit_price

            self._write_table_value(row, 11, format_decimal_for_ui(cbm))
            self._write_table_value(row, 13, format_decimal_for_ui(line_total))
            updated_rows += 1
            results.append(
                {
                    "row": str(row + 1),
                    "cbm": format_decimal_for_ui(cbm),
                    "line_total": format_decimal_for_ui(line_total),
                }
            )

        return {"updated_rows": updated_rows, "skipped_rows": skipped_rows, "rows": results}

    def _compute_document_totals(self) -> dict[str, Decimal]:
        table = getattr(self.ui, "tbl_lineItems", None)
        production_amount = Decimal("0")
        if table is not None:
            for row in range(table.rowCount()):
                if not self._line_row_has_meaningful_data(row):
                    continue
                production_amount += self._parse_table_decimal(row, 13)
                production_amount += self._parse_table_decimal(row, 14)

        tax_amount = round_decimal(production_amount * Decimal("0.05"))
        total_amount = production_amount + tax_amount
        return {
            "production_amount": production_amount,
            "tax_amount": tax_amount,
            "total_amount": total_amount,
        }

    def _apply_document_totals_to_summary_fields(self, totals: dict[str, Decimal]) -> dict[str, str]:
        self._set_summary_field_value("le_productionAmount", totals["production_amount"])
        self._set_summary_field_value("le_taxAmount", totals["tax_amount"])
        self._set_summary_field_value("le_totalAmount", totals["total_amount"])
        return {
            "production_amount": format_decimal_for_ui(totals["production_amount"]),
            "tax_amount": format_decimal_for_ui(totals["tax_amount"]),
            "total_amount": format_decimal_for_ui(totals["total_amount"]),
        }

    def calculate_document_totals(self) -> dict[str, str]:
        return self._apply_document_totals_to_summary_fields(self._compute_document_totals())

    def _handle_subtotal_clicked(self) -> None:
        try:
            result = self.calculate_line_subtotals()
        except Exception as exc:
            message = str(exc) or exc.__class__.__name__
            QMessageBox.warning(self, "小計失敗", message)
            self._set_status_message(f"小計失敗：{message}")
            return

        self._set_status_message(
            f"已更新 {result['updated_rows']} 列小計；略過 {result['skipped_rows']} 列空白列。"
        )

    def _handle_calculate_clicked(self) -> None:
        try:
            result = self.calculate_document_totals()
        except Exception as exc:
            message = str(exc) or exc.__class__.__name__
            QMessageBox.warning(self, "計算失敗", message)
            self._set_status_message(f"計算失敗：{message}")
            return

        self._set_status_message(
            f"已完成整單計算：製作金額 {result['production_amount']} / 稅額 {result['tax_amount']} / 總計 {result['total_amount']}"
        )

    def _current_client_row(self) -> dict[str, str | int | None] | None:
        customer_combo = getattr(self.ui, "cb_customerName", None)
        if not isinstance(customer_combo, QComboBox):
            return None
        customer_name = customer_combo.currentText().strip()
        return self.client_rows_by_short_name.get(customer_name) if customer_name else None

    def _billing_header_snapshot(self) -> dict[str, str]:
        client_row = self._current_client_row() or {}
        customer_combo = getattr(self.ui, "cb_customerName", None)
        customer_short_name = customer_combo.currentText().strip() if isinstance(customer_combo, QComboBox) else ""
        return {
            "work_number": normalize_line_edit_value(getattr(self.ui, "le_worknum", None)) or "",
            "customer_short_name": customer_short_name,
            "customer_full_name": str(client_row.get("full_name") or customer_short_name or ""),
            "case_name": normalize_line_edit_value(getattr(self.ui, "le_caseName", None)) or "",
            "contact_name": normalize_line_edit_value(getattr(self.ui, "le_contactName", None)) or "",
            "phone": normalize_line_edit_value(getattr(self.ui, "le_phone", None)) or "",
            "work_address": normalize_line_edit_value(getattr(self.ui, "lle_address", None)) or "",
            "work_time": normalize_line_edit_value(getattr(self.ui, "le_startTime", None)) or "",
            "cleanup_time": normalize_line_edit_value(getattr(self.ui, "le_endTime", None)) or "",
            "company_tax_id": str(client_row.get("tax_id") or ""),
            "company_address": str(client_row.get("address") or ""),
        }

    def _billing_line_items(self) -> list[dict[str, str]]:
        table = getattr(self.ui, "tbl_lineItems", None)
        if table is None:
            return []
        rows: list[dict[str, str]] = []
        for row in range(table.rowCount()):
            if not self._line_row_has_meaningful_data(row):
                continue
            width = self._table_cell_text(row, 1)
            length = self._table_cell_text(row, 3)
            size = " x ".join(part for part in [width, length] if part) if (width or length) else ""
            board_text = f"{self._table_cell_text(row, 8)}{self._table_cell_text(row, 7)}".strip()
            rows.append(
                {
                    "production_item": self._table_cell_text(row, 0),
                    "size": size,
                    "quantity": self._table_cell_text(row, 4),
                    "material": self._table_cell_text(row, 5),
                    "board": board_text,
                    "extra_material": self._table_cell_text(row, 9),
                    "extra_material_quantity": self._table_cell_text(row, 10),
                    "cbm": self._table_cell_text(row, 11),
                    "line_total": self._table_cell_text(row, 13),
                    "extra_material_total": self._table_cell_text(row, 14),
                }
            )
        return rows

    def export_billing_excels(self, output_dir: Path | None = None) -> list[Path]:
        if load_workbook is None:
            raise RuntimeError("缺少 openpyxl，無法匯出請款 Excel。")
        template_path = resolve_billing_template_path()
        if not template_path.exists():
            raise FileNotFoundError(f"找不到請款樣板：{template_path}")

        totals = self.calculate_document_totals()
        header = self._billing_header_snapshot()
        if not header["work_number"]:
            raise ValueError("工單號不可為空。")
        if not header["customer_short_name"]:
            raise ValueError("請先選擇客戶，再匯出請款 Excel。")
        line_items = self._billing_line_items()
        if not line_items:
            raise ValueError("目前沒有可匯出的明細列。")

        export_dir = output_dir.expanduser().resolve() if output_dir is not None else Path.cwd()
        export_dir.mkdir(parents=True, exist_ok=True)
        page_chunks = chunked(line_items, BILLING_MAX_ROWS_PER_PAGE)
        exported_paths: list[Path] = []
        for page_index, page_rows in enumerate(page_chunks, start=1):
            workbook = load_workbook(template_path)
            sheet = workbook.active
            apply_billing_header_to_sheet(sheet, header)
            apply_billing_rows_to_sheet(sheet, page_rows)
            if page_index == len(page_chunks):
                apply_billing_totals_to_sheet(sheet, totals)
            else:
                clear_billing_totals_on_sheet(sheet)

            output_path = export_dir / build_billing_output_filename(
                header["work_number"],
                header["customer_short_name"],
                header["case_name"],
                page_index,
                len(page_chunks),
            )
            workbook.save(output_path)
            exported_paths.append(output_path)
        return exported_paths

    def _handle_billing_clicked(self) -> None:
        output_dir = resolve_billing_output_dir()
        try:
            exported_paths = self.export_billing_excels(output_dir)
        except Exception as exc:
            message = str(exc) or exc.__class__.__name__
            QMessageBox.warning(self, "請款匯出失敗", message)
            self._set_status_message(f"請款匯出失敗：{message}")
            return

        files_text = "\n".join(str(path) for path in exported_paths)
        QMessageBox.information(self, "請款匯出完成", f"已輸出 {len(exported_paths)} 個 Excel：\n{files_text}")
        self._set_status_message(f"請款 Excel 已輸出 {len(exported_paths)} 個檔案。")

    def _collect_line_payloads(self) -> list[dict[str, object | None]]:
        table = getattr(self.ui, "tbl_lineItems", None)
        if table is None:
            return []

        line_payloads: list[dict[str, object | None]] = []
        for row in range(table.rowCount()):
            row_number = row + 1
            if not self._line_row_has_meaningful_data(row):
                continue

            production_item = self._table_cell_text(row, 0)
            width_mm = self._table_cell_text(row, 1)
            height_mm = self._table_cell_text(row, 3)
            quantity = self._table_cell_text(row, 4)
            material = self._table_cell_text(row, 5)
            lamination = self._table_cell_text(row, 6)
            board_type = self._table_cell_text(row, 7)
            board_thickness = self._table_cell_text(row, 8)
            extra_material = self._table_cell_text(row, 9)
            extra_material_quantity = self._table_cell_text(row, 10)
            cbm = self._table_cell_text(row, 11)
            cbm_unit_price = self._table_cell_text(row, 12)
            line_total = self._table_cell_text(row, 13)
            extra_material_total = self._table_cell_text(row, 14)

            production_item_id, production_item_text = self._resolve_option_item_storage("production_item", production_item)
            material_id, material_text = self._resolve_option_item_storage("material", material)
            lamination_id, lamination_text = self._resolve_option_item_storage("lamination", lamination)
            board_type_id, board_type_text = self._resolve_option_item_storage("board_type", board_type)
            board_thickness_id, board_thickness_text = self._resolve_option_item_storage("board_thickness", board_thickness)
            extra_material_id, extra_material_text = self._resolve_option_item_storage("extra_material", extra_material)

            line_payloads.append(
                {
                    "line_no": len(line_payloads) + 1,
                    "production_item_id": production_item_id,
                    "production_item_text": production_item_text,
                    "width_mm": parse_decimal_or_none(width_mm),
                    "height_mm": parse_decimal_or_none(height_mm),
                    "quantity": parse_int_or_none(quantity),
                    "material_id": material_id,
                    "material_text": material_text,
                    "lamination_id": lamination_id,
                    "lamination_text": lamination_text,
                    "board_type_id": board_type_id,
                    "board_type_text": board_type_text,
                    "board_thickness_id": board_thickness_id,
                    "board_thickness_text": board_thickness_text,
                    "extra_material_id": extra_material_id,
                    "extra_material_text": extra_material_text,
                    "extra_material_quantity": parse_int_or_none(extra_material_quantity),
                    "cbm": parse_decimal_or_none(cbm),
                    "cbm_unit_price": parse_decimal_or_none(cbm_unit_price),
                    "line_total": parse_decimal_or_none(line_total),
                    "extra_material_total": parse_decimal_or_none(extra_material_total),
                }
            )

        return line_payloads

    def _upsert_work_order_header(self, cur, payload: dict[str, str | int | None]) -> int:
        insert_sql = """
            INSERT INTO work_orders (
                work_number, case_name, client_id, customer_name_text, company_phone,
                contact_name, work_time, cleanup_time, work_address,
                production_amount, tax_amount, total_amount, remark, status
            ) VALUES (
                %(work_number)s, %(case_name)s, %(client_id)s, %(customer_name_text)s, %(company_phone)s,
                %(contact_name)s, %(work_time)s, %(cleanup_time)s, %(work_address)s,
                %(production_amount)s, %(tax_amount)s, %(total_amount)s, %(remark)s, %(status)s
            )
            ON DUPLICATE KEY UPDATE
                case_name = VALUES(case_name),
                client_id = VALUES(client_id),
                customer_name_text = VALUES(customer_name_text),
                company_phone = VALUES(company_phone),
                contact_name = VALUES(contact_name),
                work_time = VALUES(work_time),
                cleanup_time = VALUES(cleanup_time),
                work_address = VALUES(work_address),
                production_amount = VALUES(production_amount),
                tax_amount = VALUES(tax_amount),
                total_amount = VALUES(total_amount),
                remark = VALUES(remark),
                status = VALUES(status),
                id = LAST_INSERT_ID(id)
        """
        cur.execute(insert_sql, payload)
        return int(cur.lastrowid)

    def _replace_work_order_lines(self, cur, work_order_id: int, line_payloads: list[dict[str, object | None]]) -> None:
        cur.execute("DELETE FROM work_order_lines WHERE work_order_id = %s", (work_order_id,))
        if not line_payloads:
            return

        insert_sql = """
            INSERT INTO work_order_lines (
                work_order_id, line_no, production_item_id, production_item_text, width_mm, height_mm, quantity,
                material_id, material_text, lamination_id, lamination_text, board_type_id, board_type_text,
                board_thickness_id, board_thickness_text, extra_material_id, extra_material_text,
                extra_material_quantity, cbm, cbm_unit_price, line_total, extra_material_total
            ) VALUES (
                %(work_order_id)s, %(line_no)s, %(production_item_id)s, %(production_item_text)s, %(width_mm)s, %(height_mm)s, %(quantity)s,
                %(material_id)s, %(material_text)s, %(lamination_id)s, %(lamination_text)s, %(board_type_id)s, %(board_type_text)s,
                %(board_thickness_id)s, %(board_thickness_text)s, %(extra_material_id)s, %(extra_material_text)s,
                %(extra_material_quantity)s, %(cbm)s, %(cbm_unit_price)s, %(line_total)s, %(extra_material_total)s
            )
        """
        rows = []
        for payload in line_payloads:
            row_payload = dict(payload)
            row_payload["work_order_id"] = work_order_id
            rows.append(row_payload)
        cur.executemany(insert_sql, rows)

    def save_header_to_work_orders(self) -> int:
        if pymysql is None:
            raise RuntimeError("缺少 PyMySQL，無法儲存到 work_order_v2。")

        payload = self._header_payload()
        with pymysql.connect(**_connect_kwargs()) as conn:
            with conn.cursor() as cur:
                work_order_id = self._upsert_work_order_header(cur, payload)
            conn.commit()
        return work_order_id

    def _fetch_work_order_bundle(self, work_number: str) -> tuple[dict[str, object], list[dict[str, object]]]:
        if pymysql is None:
            raise RuntimeError("缺少 PyMySQL，無法讀取 work_order_v2。")
        normalized_work_number = work_number.strip()
        if not normalized_work_number:
            raise ValueError("請先輸入工單號再開啟。")

        header_sql = """
            SELECT
                wo.id,
                wo.work_number,
                wo.case_name,
                wo.client_id,
                wo.customer_name_text,
                c.short_name AS client_short_name,
                wo.company_phone,
                wo.contact_name,
                wo.work_time,
                wo.cleanup_time,
                wo.work_address,
                wo.production_amount,
                wo.tax_amount,
                wo.total_amount,
                wo.remark,
                wo.status
            FROM work_orders wo
            LEFT JOIN clients c ON c.id = wo.client_id
            WHERE wo.work_number = %s
            LIMIT 1
        """
        line_sql = """
            SELECT
                wol.line_no,
                wol.width_mm,
                wol.height_mm,
                wol.quantity,
                wol.extra_material_quantity,
                wol.cbm,
                wol.cbm_unit_price,
                wol.line_total,
                wol.extra_material_total,
                wol.production_item_text,
                wol.material_text,
                wol.lamination_text,
                wol.board_type_text,
                wol.board_thickness_text,
                wol.extra_material_text,
                pi.item_name AS production_item_name,
                mi.item_name AS material_name,
                li.item_name AS lamination_name,
                bti.item_name AS board_type_name,
                bthi.item_name AS board_thickness_name,
                emi.item_name AS extra_material_name
            FROM work_order_lines wol
            LEFT JOIN option_items pi ON pi.id = wol.production_item_id
            LEFT JOIN option_items mi ON mi.id = wol.material_id
            LEFT JOIN option_items li ON li.id = wol.lamination_id
            LEFT JOIN option_items bti ON bti.id = wol.board_type_id
            LEFT JOIN option_items bthi ON bthi.id = wol.board_thickness_id
            LEFT JOIN option_items emi ON emi.id = wol.extra_material_id
            WHERE wol.work_order_id = %s
            ORDER BY wol.line_no, wol.id
        """

        with pymysql.connect(**_connect_kwargs()) as conn:
            with conn.cursor() as cur:
                cur.execute(header_sql, (normalized_work_number,))
                header_row = cur.fetchone()
                if header_row is None:
                    raise LookupError(f"找不到工單號：{normalized_work_number}")
                cur.execute(line_sql, (header_row["id"],))
                line_rows = cur.fetchall()
        return header_row, line_rows

    def _coerce_option_item_name(self, option_group: str, value: object) -> str:
        if value is None:
            return ""
        if isinstance(value, int):
            return self.option_item_names_by_group_and_id.get(option_group, {}).get(value, "")
        if isinstance(value, str):
            return value.strip()
        try:
            return self.option_item_names_by_group_and_id.get(option_group, {}).get(int(value), "")
        except (TypeError, ValueError):
            return str(value).strip()

    def _build_table_rows_from_line_payloads(self, line_rows: list[dict[str, object]]) -> list[list[str]]:
        target_row_count = max(DEFAULT_LINE_ITEM_ROW_COUNT, len(line_rows) + 1)
        built_rows: list[list[str]] = []
        for line_row in line_rows:
            built_rows.append(
                [
                    self._coerce_option_item_name("production_item", line_row.get("production_item_name") or line_row.get("production_item_text")),
                    format_decimal_for_ui(line_row.get("width_mm")),
                    "x",
                    format_decimal_for_ui(line_row.get("height_mm")),
                    format_decimal_for_ui(line_row.get("quantity")),
                    self._coerce_option_item_name("material", line_row.get("material_name") or line_row.get("material_text")),
                    self._coerce_option_item_name("lamination", line_row.get("lamination_name") or line_row.get("lamination_text")),
                    self._coerce_option_item_name("board_type", line_row.get("board_type_name") or line_row.get("board_type_text")),
                    self._coerce_option_item_name("board_thickness", line_row.get("board_thickness_name") or line_row.get("board_thickness_text")),
                    self._coerce_option_item_name("extra_material", line_row.get("extra_material_name") or line_row.get("extra_material_text")),
                    format_decimal_for_ui(line_row.get("extra_material_quantity")),
                    format_decimal_for_ui(line_row.get("cbm")),
                    format_decimal_for_ui(line_row.get("cbm_unit_price")),
                    format_decimal_for_ui(line_row.get("line_total")),
                    format_decimal_for_ui(line_row.get("extra_material_total")),
                ]
            )

        while len(built_rows) < target_row_count:
            built_rows.append(list(BLANK_LINE_ROW_TEMPLATE))
        return built_rows

    def load_work_order_by_number(self, work_number: str) -> tuple[int, int]:
        header_row, line_rows = self._fetch_work_order_bundle(work_number)

        customer_combo = getattr(self.ui, "cb_customerName", None)
        if isinstance(customer_combo, QComboBox):
            customer_name = str(header_row.get("client_short_name") or header_row.get("customer_name_text") or "").strip()
            customer_index = customer_combo.findText(customer_name) if customer_name else -1
            if customer_index >= 0:
                customer_combo.setCurrentIndex(customer_index)
            else:
                customer_combo.setCurrentIndex(-1)
                customer_combo.setEditText(customer_name)

        field_mappings = {
            "le_worknum": header_row.get("work_number"),
            "le_caseName": header_row.get("case_name"),
            "le_phone": header_row.get("company_phone"),
            "le_contactName": header_row.get("contact_name"),
            "le_startTime": header_row.get("work_time"),
            "le_endTime": header_row.get("cleanup_time"),
            "lle_address": header_row.get("work_address"),
            "le_productionAmount": format_decimal_for_ui(header_row.get("production_amount")),
            "le_taxAmount": format_decimal_for_ui(header_row.get("tax_amount")),
            "le_totalAmount": format_decimal_for_ui(header_row.get("total_amount")),
        }
        for attr, value in field_mappings.items():
            widget = getattr(self.ui, attr, None)
            if isinstance(widget, QLineEdit):
                widget.setText(str(value or ""))

        if hasattr(self.ui, "te_remark"):
            self.ui.te_remark.setPlainText(str(header_row.get("remark") or ""))

        table_rows = self._build_table_rows_from_line_payloads(line_rows)
        self._populate_line_items_table_with_rows(table_rows)
        if not (field_mappings["le_productionAmount"] and field_mappings["le_taxAmount"] and field_mappings["le_totalAmount"]):
            self.calculate_document_totals()
        self._last_auto_filled_phone = str(header_row.get("company_phone") or "")
        self._last_auto_filled_address = str(header_row.get("work_address") or "")
        self._loaded_work_order_number = str(header_row.get("work_number") or "").strip()
        return int(header_row["id"]), len(line_rows)

    def _has_loaded_content_on_screen(self) -> bool:
        return self._screen_has_meaningful_content(include_lookup_work_number=True)

    def _screen_has_meaningful_content(self, *, include_lookup_work_number: bool) -> bool:
        if self._loaded_work_order_number:
            return True

        customer_combo = getattr(self.ui, "cb_customerName", None)
        if isinstance(customer_combo, QComboBox) and customer_combo.currentText().strip():
            return True

        line_edit_attrs = [
            "le_contactName",
            "le_phone",
            "le_caseName",
            "le_startTime",
            "le_endTime",
            "lle_address",
            "le_productionAmount",
            "le_taxAmount",
            "le_totalAmount",
        ]
        if include_lookup_work_number:
            line_edit_attrs.insert(0, "le_worknum")

        for attr in line_edit_attrs:
            widget = getattr(self.ui, attr, None)
            if isinstance(widget, QLineEdit) and widget.text().strip():
                return True

        if hasattr(self.ui, "te_remark") and self.ui.te_remark.toPlainText().strip():
            return True

        table = getattr(self.ui, "tbl_lineItems", None)
        if table is not None:
            for row in range(table.rowCount()):
                if self._line_row_has_meaningful_data(row):
                    return True
        return False

    def reset_work_order_to_blank(self) -> None:
        self._initialize_blank_work_order()
        table = getattr(self.ui, "tbl_lineItems", None)
        if table is not None and table.rowCount() > 0:
            table.setCurrentCell(0, 0)
        self._set_status_message("已重置為空白新工單。")

    def _handle_reset_clicked(self) -> None:
        if self._has_loaded_content_on_screen():
            confirm = QMessageBox.question(
                self,
                "重置工單",
                "目前畫面已有內容；重置會清空整張工單並回到空白新工單狀態。\n\n要繼續嗎？",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No,
            )
            if confirm != QMessageBox.StandardButton.Yes:
                self._set_status_message("已取消重置工單。")
                return

        self.reset_work_order_to_blank()

    def _handle_open_clicked(self) -> None:
        work_number = normalize_line_edit_value(getattr(self.ui, "le_worknum", None)) or ""
        if self._screen_has_meaningful_content(include_lookup_work_number=False):
            confirm = QMessageBox.question(
                self,
                "開啟工單",
                "目前畫面內容會直接被載入的工單覆蓋；尚未做完整 dirty-check。\n\n要繼續開啟嗎？",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No,
            )
            if confirm != QMessageBox.StandardButton.Yes:
                self._set_status_message("已取消開啟工單。")
                return

        try:
            work_order_id, line_count = self.load_work_order_by_number(work_number)
        except Exception as exc:
            message = str(exc) or exc.__class__.__name__
            QMessageBox.warning(self, "開啟失敗", message)
            self._set_status_message(f"開啟失敗：{message}")
            return

        success_message = f"已開啟 work_orders #{work_order_id}（工單 {work_number}，明細 {line_count} 筆）"
        QMessageBox.information(self, "開啟成功", success_message)
        self._set_status_message(success_message)

    def _handle_import_clicked(self) -> None:
        initial_dir = IMPORT_ROOT_HINT if Path(IMPORT_ROOT_HINT).exists() else str(Path.home())
        selected_folder = QFileDialog.getExistingDirectory(self, "選擇導入資料夾", initial_dir)
        if not selected_folder:
            self._set_status_message("已取消導入。")
            return

        try:
            preview_payload = self._build_import_preview_payload(selected_folder)
        except Exception as exc:
            message = str(exc) or exc.__class__.__name__
            QMessageBox.warning(self, "導入失敗", message)
            self._set_status_message(f"導入失敗：{message}")
            return

        if not self._confirm_replace_current_screen_if_needed():
            return

        self._apply_import_preview_to_screen(preview_payload)
        imported_count = len(preview_payload["lines"])
        review_count = sum(1 for line in preview_payload["lines"] if line.get("needs_review"))
        if review_count:
            message = f"已導入 {imported_count} 筆明細；其中 {review_count} 筆需要你再看一下。"
        else:
            message = f"已導入 {imported_count} 筆明細。"
        self._set_status_message(message)
        QMessageBox.information(self, "導入完成", message)

    def _build_import_preview_payload(self, selected_folder: str) -> dict[str, object]:
        loading = QProgressDialog("正在讀取資料夾與解析檔名…", None, 0, 0, self)
        loading.setWindowTitle("導入中")
        loading.setWindowModality(Qt.WindowModality.ApplicationModal)
        loading.setCancelButton(None)
        loading.setMinimumDuration(0)
        loading.show()
        QApplication.processEvents()
        try:
            context = parse_import_path_context(selected_folder)
            image_paths = sorted(
                [
                    path for path in Path(selected_folder).iterdir()
                    if path.is_file() and path.suffix.lower() in IMPORT_IMAGE_SUFFIXES
                ],
                key=lambda p: p.name.lower(),
            )
            if not image_paths:
                raise ValueError("選到的資料夾內沒有 .jpg / .jpeg 檔案。")
            lines = [self._parse_import_file(path, context) for path in image_paths]
            loading.setLabelText("正在交給 AI 輔助解析規格…")
            QApplication.processEvents()
            self._apply_ai_resolution_to_import_lines(lines, context)
            return {
                "source_folder": context["source_folder"],
                "customer_name": context["customer_name"],
                "case_folder": context["case_folder"],
                "case_name": context["case_name"],
                "work_number": "",
                "lines": lines,
            }
        finally:
            loading.close()

    def _parse_import_file(self, file_path: Path, context: dict[str, str]) -> dict[str, object]:
        stem = unicodedata.normalize("NFKC", file_path.stem)
        size_text, stem_without_inline_size = extract_inline_wh_size_from_import_stem(stem)
        inline_material, stem_without_inline_material = extract_inline_material_from_import_text(stem_without_inline_size)
        parts = [part.strip(" _") for part in stem_without_inline_material.split("-") if part.strip(" _")]
        production_item = parts[0] if parts else stem_without_inline_material.strip(" _-") or stem
        rest_parts = parts[1:]
        if not size_text and rest_parts and looks_like_size_text(rest_parts[0]):
            size_text = rest_parts[0]
            rest_parts = rest_parts[1:]
        raw_spec_text = "-".join(rest_parts)
        raw_spec_tokens = [token.strip() for token in raw_spec_text.split("+") if token.strip()]
        spec_payload = self._resolve_import_spec_tokens(raw_spec_tokens, context)
        if inline_material and not spec_payload.get("material"):
            spec_payload["material"] = inline_material
        return {
            "source_file": file_path.name,
            "original_filename": stem,
            "filename_parts": parts,
            "production_item": production_item,
            "size_text": size_text,
            "raw_spec_text": raw_spec_text,
            "raw_spec_tokens": raw_spec_tokens,
            **spec_payload,
        }

    def _resolve_import_spec_tokens(self, raw_tokens: list[str], context: dict[str, str]) -> dict[str, object]:
        material = ""
        lamination = default_import_lamination(context.get("customer_name", ""))
        board_type = ""
        board_thickness = ""
        extra_materials: list[str] = []
        ignored_tokens: list[str] = []
        unresolved_tokens: list[str] = []
        pending_thickness_tokens: list[tuple[int, str]] = []
        notes: list[str] = []
        quantity = 1
        consumed_indexes: set[int] = set()
        candidate_extra_materials = self.combo_column_options.get(9, [])

        for index, raw_token in enumerate(raw_tokens):
            token = unicodedata.normalize("NFKC", raw_token).strip()
            normalized = normalize_token_text(token)
            if not normalized:
                consumed_indexes.add(index)
                continue

            quantity_match = QUANTITY_TOKEN_PATTERN.fullmatch(normalized)
            if quantity_match:
                quantity = max(1, int(quantity_match.group(1)))
                consumed_indexes.add(index)
                continue

            if normalized in {"亮面", "霧面"}:
                lamination = normalized
                consumed_indexes.add(index)
                continue

            material_name = MATERIAL_ALIASES.get(normalized)
            if material_name:
                material = material_name
                consumed_indexes.add(index)
                continue

            board_name = BOARD_TYPE_ALIASES.get(token) or BOARD_TYPE_ALIASES.get(normalized)
            if board_name:
                board_type = board_name
                consumed_indexes.add(index)
                continue

            if looks_like_thickness_text(token):
                pending_thickness_tokens.append((index, token))
                continue

            if token.startswith("右折") or token.startswith("左折"):
                ignored_tokens.append(token)
                consumed_indexes.add(index)
                continue

            matched_name, score = match_candidate(token, candidate_extra_materials)
            if matched_name and score >= 0.58:
                extra_materials.append(matched_name)
                if score < 0.85:
                    notes.append(f"{token} 近似配對為 {matched_name}")
                consumed_indexes.add(index)
                continue

        if board_type:
            for index, token in pending_thickness_tokens:
                if not board_thickness:
                    board_thickness = token
                consumed_indexes.add(index)
            if not board_thickness:
                board_thickness = BOARD_TYPE_DEFAULT_THICKNESS.get(board_type, "")

        for index, raw_token in enumerate(raw_tokens):
            if index not in consumed_indexes:
                unresolved_tokens.append(unicodedata.normalize("NFKC", raw_token).strip())

        confidence = 0.95
        if notes:
            confidence -= 0.12
        if unresolved_tokens:
            confidence -= min(0.4, 0.12 * len(unresolved_tokens))
        confidence = max(0.1, round(confidence, 2))

        return {
            "material": material,
            "lamination": lamination,
            "board_type": board_type,
            "board_thickness": board_thickness,
            "extra_materials": extra_materials,
            "quantity": quantity,
            "ignored_tokens": ignored_tokens,
            "unresolved_tokens": unresolved_tokens,
            "notes": notes,
            "confidence": confidence,
        }

    def _apply_ai_resolution_to_import_lines(self, lines: list[dict[str, object]], context: dict[str, str]) -> None:
        ai_results = self._call_import_ai_for_lines(lines, context)
        for index, line in enumerate(lines):
            ai_line = ai_results.get(index, {})
            merged = self._merge_ai_line_resolution(line, ai_line)
            lines[index] = self._finalize_import_line_resolution(merged, context)

    def _build_import_ai_candidates(self) -> dict[str, list[str]]:
        return {
            "materials": sorted({*self.combo_column_options.get(5, []), *MATERIAL_ALIASES.values()}),
            "laminations": sorted({*self.combo_column_options.get(6, []), "亮面", "霧面"}),
            "board_types": sorted({*self.combo_column_options.get(7, []), *BOARD_TYPE_DEFAULT_THICKNESS.keys()}),
            "board_thicknesses": sorted({*self.combo_column_options.get(8, []), *[value for value in BOARD_TYPE_DEFAULT_THICKNESS.values() if value]}),
            "extra_materials": list(self.combo_column_options.get(9, [])),
        }

    def _call_import_ai_for_lines(self, lines: list[dict[str, object]], context: dict[str, str]) -> dict[int, dict[str, object]]:
        if not deepseek_configured() or not lines:
            return {}

        payload_items = []
        for index, line in enumerate(lines):
            payload_items.append(
                {
                    "index": index,
                    "customer_name": context.get("customer_name", ""),
                    "case_name": context.get("case_name", ""),
                    "file_name": line.get("source_file", ""),
                    "original_filename": str(line.get("original_filename") or Path(str(line.get("source_file", ""))).stem),
                    "filename_parts": list(line.get("filename_parts") or []),
                    "production_item": line.get("production_item", ""),
                    "size_text": line.get("size_text", ""),
                    "raw_spec_text": line.get("raw_spec_text", ""),
                    "raw_spec_tokens": line.get("raw_spec_tokens", []),
                    "current_guess": {
                        "material": line.get("material", ""),
                        "lamination": line.get("lamination", ""),
                        "board_type": line.get("board_type", ""),
                        "board_thickness": line.get("board_thickness", ""),
                        "extra_materials": line.get("extra_materials", []),
                        "ignored_tokens": line.get("ignored_tokens", []),
                        "unresolved_tokens": line.get("unresolved_tokens", []),
                        "quantity": line.get("quantity", 1),
                    },
                }
            )

        prompt_payload = {
            "rules": {
                "lamination_default": "冠銘專用=亮面，其餘=霧面；若 token 明寫亮面/霧面，以 token 為準",
                "board_defaults": BOARD_TYPE_DEFAULT_THICKNESS,
                "ignore_patterns": ["右折*", "左折*"],
                "notes": [
                    "請直接根據 customer_name、case_name、original_filename、filename_parts、raw_spec_text、raw_spec_tokens 做語意抽取，不要假設第幾段一定是名稱或尺寸。",
                    "current_guess 只是 fallback 參考，不是最終答案；若你更有把握，可以覆蓋 production_item、size_text、material、lamination、board_type、board_thickness、extra_materials、quantity。",
                    "若檔名某一段混合了名稱與尺寸（例如 頭卡30x20cm），請自行拆出語意，不要原樣照抄到 production_item。",
                    "板材厚度只能在同一檔名已有板材 token（例如 H、合成板、發泡板、瓦楞板）時才判定；沒有板材就不要把 30MM/5MM 之類解析成 board_thickness，應保留在 production_item 或 notes。",
                    "若同時出現多個尺寸或數量線索，quantity 只放工單總數；其他像大版尺寸、印幾片、排版資訊，放進 notes，不要硬塞進 quantity。",
                    "若 token 與既有候選值很像，優先對應到候選值，不要發明新名字。",
                    "若仍無法判斷，欄位留空，並把原因寫進 review_reason；不要硬猜。",
                ],
            },
            "candidates": self._build_import_ai_candidates(),
            "items": payload_items,
        }
        try:
            response_json = self._chat_complete_import_ai(prompt_payload)
        except Exception as exc:
            self._set_status_message(f"AI 解析失敗，改用本地規則：{exc}")
            return {}

        items = response_json.get("items") if isinstance(response_json, dict) else None
        if not isinstance(items, list):
            return {}

        results: dict[int, dict[str, object]] = {}
        for item in items:
            if not isinstance(item, dict):
                continue
            try:
                index = int(item.get("index"))
            except (TypeError, ValueError):
                continue
            results[index] = item
        return results

    def _chat_complete_import_ai(self, prompt_payload: dict[str, object]) -> dict[str, object]:
        system_prompt = (
            "你是工單檔名解析助手。"
            "你的工作是從不穩定、格式不固定的檔名中做語意抽取，而不是套固定模板。"
            "請根據 candidates、customer_name、case_name、original_filename、filename_parts 與 current_guess，直接輸出完整工單列解析結果。"
            "輸出必須是 JSON 物件，格式為 {\"items\": [...]}。"
            "每個 item 必須包含 index, production_item, size_text, material, lamination, board_type, board_thickness, extra_materials, quantity, needs_review, review_reason, unresolved_tokens, notes, confidence。"
            "quantity 只代表工單總數；其他像大版尺寸、切割尺寸、印幾片、拼版資訊請放在 notes。"
            "board_thickness 只有在 board_type 有板材時才可填；沒有 H/合成板/發泡板/瓦楞板等板材線索時，30MM/5MM 應視為名稱內容，不可當板厚。"
            "能從 candidates 選的就選 candidates；不確定就留空並 needs_review=true。"
            "不要輸出 markdown，不要輸出額外說明。"
        )
        user_prompt = json.dumps(prompt_payload, ensure_ascii=False)
        api_key = os.environ.get("OPENROUTER_API_KEY") or os.environ.get("DEEPSEEK_API_KEY")
        if not api_key:
            raise RuntimeError("未設定 OPENROUTER_API_KEY 或 DEEPSEEK_API_KEY")
        api_url = OPENROUTER_API_URL if os.environ.get("OPENROUTER_API_KEY") else DEEPSEEK_API_URL
        request_body = {
            "model": IMPORT_AI_MODEL,
            "temperature": 0.1,
            "response_format": {"type": "json_object"},
            "messages": [
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": user_prompt},
            ],
        }
        headers = {
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json",
        }
        if api_url == OPENROUTER_API_URL:
            headers["HTTP-Referer"] = "https://github.com/jackey10055206/work-order"
            headers["X-Title"] = "work-order import"

        request = urllib_request.Request(
            api_url,
            data=json.dumps(request_body).encode("utf-8"),
            headers=headers,
            method="POST",
        )
        try:
            with urllib_request.urlopen(request, timeout=IMPORT_AI_TIMEOUT_SECONDS) as response:
                body = response.read().decode("utf-8")
        except urllib_error.HTTPError as exc:
            detail = exc.read().decode("utf-8", errors="replace") if exc.fp else str(exc)
            raise RuntimeError(f"AI HTTP {exc.code}: {detail[:240]}") from exc
        except urllib_error.URLError as exc:
            raise RuntimeError(f"AI 連線失敗：{exc.reason}") from exc

        payload = json.loads(body)
        choices = payload.get("choices") if isinstance(payload, dict) else None
        if not choices:
            raise RuntimeError("AI 回傳格式不正確：缺少 choices")
        message = choices[0].get("message", {})
        content = message.get("content", "")
        if isinstance(content, list):
            content = "".join(str(part.get("text", "")) if isinstance(part, dict) else str(part) for part in content)
        parsed = json.loads(extract_json_text(str(content)))
        if not isinstance(parsed, dict):
            raise RuntimeError("AI JSON 解析失敗：不是物件")
        return parsed

    def _merge_ai_line_resolution(self, line: dict[str, object], ai_line: dict[str, object]) -> dict[str, object]:
        merged = dict(line)
        for key in ("production_item", "size_text", "material", "lamination", "board_type", "board_thickness"):
            ai_value = str(ai_line.get(key) or "").strip()
            if ai_value:
                merged[key] = ai_value
        try:
            ai_quantity = int(ai_line.get("quantity")) if ai_line.get("quantity") is not None else None
        except (TypeError, ValueError):
            ai_quantity = None
        if ai_quantity and ai_quantity > 0:
            merged["quantity"] = ai_quantity

        extra_materials = ensure_str_list(merged.get("extra_materials"))
        for value in ensure_str_list(ai_line.get("extra_materials")):
            if value not in extra_materials:
                extra_materials.append(value)
        merged["extra_materials"] = extra_materials

        ignored_tokens = ensure_str_list(merged.get("ignored_tokens"))
        for value in ensure_str_list(ai_line.get("ignored_tokens")):
            if value not in ignored_tokens:
                ignored_tokens.append(value)
        merged["ignored_tokens"] = ignored_tokens

        merged["unresolved_tokens"] = ensure_str_list(ai_line.get("unresolved_tokens"))
        merged["review_reason"] = ensure_str_list(ai_line.get("review_reason"))
        merged["needs_review"] = bool(ai_line.get("needs_review"))

        notes = ensure_str_list(merged.get("notes"))
        for value in ensure_str_list(ai_line.get("notes")):
            if value not in notes:
                notes.append(value)
        merged["notes"] = notes

        try:
            confidence = float(ai_line.get("confidence"))
        except (TypeError, ValueError):
            confidence = float(merged.get("confidence") or 0)
        merged["confidence"] = max(0.1, min(1.0, round(confidence, 2)))
        return merged

    def _normalize_import_candidate_value(self, value: object, candidates: list[str]) -> str:
        text = str(value or "").strip()
        if not text:
            return ""
        if text in candidates:
            return text
        matched_name, score = match_candidate(text, candidates)
        if matched_name and score >= 0.7:
            return matched_name
        return ""

    def _finalize_import_line_resolution(self, line: dict[str, object], context: dict[str, str]) -> dict[str, object]:
        finalized = dict(line)
        candidates = self._build_import_ai_candidates()
        review_reason = ensure_str_list(finalized.get("review_reason"))
        notes = ensure_str_list(finalized.get("notes"))
        unresolved_tokens = ensure_str_list(finalized.get("unresolved_tokens"))

        finalized["production_item"] = str(finalized.get("production_item") or "").strip()
        finalized["size_text"] = str(finalized.get("size_text") or "").strip()
        finalized["material"] = self._normalize_import_candidate_value(finalized.get("material"), candidates["materials"])
        finalized["lamination"] = self._normalize_import_candidate_value(finalized.get("lamination"), candidates["laminations"])
        finalized["board_type"] = self._normalize_import_candidate_value(finalized.get("board_type"), candidates["board_types"])
        finalized["board_thickness"] = self._normalize_import_candidate_value(finalized.get("board_thickness"), candidates["board_thicknesses"])

        if not finalized["lamination"]:
            finalized["lamination"] = default_import_lamination(context.get("customer_name", ""))

        if finalized["board_type"] and not finalized["board_thickness"]:
            default_thickness = BOARD_TYPE_DEFAULT_THICKNESS.get(finalized["board_type"], "")
            if default_thickness:
                finalized["board_thickness"] = default_thickness
        if not finalized["board_type"]:
            finalized["board_thickness"] = ""

        normalized_extra_materials: list[str] = []
        unknown_extra_materials: list[str] = []
        for value in ensure_str_list(finalized.get("extra_materials")):
            normalized_value = self._normalize_import_candidate_value(value, candidates["extra_materials"])
            if normalized_value:
                if normalized_value not in normalized_extra_materials:
                    normalized_extra_materials.append(normalized_value)
            elif value not in unknown_extra_materials:
                unknown_extra_materials.append(value)
        finalized["extra_materials"] = normalized_extra_materials

        quantity = extract_first_int(finalized.get("quantity"))
        if quantity is None:
            quantity = 1
            if finalized.get("quantity") not in (None, "", 1, "1"):
                review_reason.append("數量格式有問題，已回退成 1")
        finalized["quantity"] = max(1, quantity)

        if unknown_extra_materials:
            review_reason.append(f"其他備料未完全對上：{'+'.join(unknown_extra_materials)}")
        if unresolved_tokens:
            review_reason.append(f"仍有未解規格：{'+'.join(unresolved_tokens)}")
        if not finalized["production_item"]:
            review_reason.append("製作項目未判定")
        if finalized["size_text"] and not looks_like_size_text(finalized["size_text"]):
            review_reason.append(f"尺寸格式待確認：{finalized['size_text']}")
        if not finalized["material"]:
            review_reason.append("材質未判定")
        if not finalized["board_type"] and finalized["board_thickness"]:
            review_reason.append("只有板厚、沒有板材")

        unique_review_reason: list[str] = []
        for value in review_reason:
            text = str(value).strip()
            if text and text not in unique_review_reason:
                unique_review_reason.append(text)
        finalized["review_reason"] = unique_review_reason
        finalized["notes"] = ensure_str_list(notes)
        finalized["needs_review"] = bool(finalized.get("needs_review")) or bool(unique_review_reason)

        try:
            confidence = float(finalized.get("confidence") or 0)
        except (TypeError, ValueError):
            confidence = 0.35
        if finalized["needs_review"] and confidence > 0.74:
            confidence = 0.74
        finalized["confidence"] = max(0.1, min(1.0, round(confidence, 2)))
        return finalized

    def _confirm_replace_current_screen_if_needed(self) -> bool:
        if not self._has_loaded_content_on_screen():
            return True
        confirm = QMessageBox.question(
            self,
            "覆蓋目前內容",
            "目前畫面已有資料；導入會覆蓋目前工單內容。\n\n要繼續嗎？",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No,
        )
        if confirm != QMessageBox.StandardButton.Yes:
            self._set_status_message("已取消覆蓋目前內容。")
            return False
        return True

    def _apply_import_preview_to_screen(self, preview_payload: dict[str, object]) -> None:
        self._initialize_blank_work_order()
        self._import_review_rows = {}
        customer_name = str(preview_payload.get("customer_name") or "")
        customer_combo = getattr(self.ui, "cb_customerName", None)
        if isinstance(customer_combo, QComboBox):
            customer_index = customer_combo.findText(customer_name)
            if customer_index >= 0:
                customer_combo.setCurrentIndex(customer_index)
            else:
                customer_combo.setEditText(customer_name)

        field_values = {
            "le_worknum": str(preview_payload.get("work_number") or ""),
            "le_caseName": str(preview_payload.get("case_name") or ""),
        }
        for attr, value in field_values.items():
            widget = getattr(self.ui, attr, None)
            if isinstance(widget, QLineEdit):
                widget.setText(value)

        table_rows = [self._build_table_row_from_import_line(line) for line in list(preview_payload.get("lines") or [])]
        while len(table_rows) < DEFAULT_LINE_ITEM_ROW_COUNT:
            table_rows.append(list(BLANK_LINE_ROW_TEMPLATE))
        if not table_rows or self._row_values_have_meaningful_content(table_rows[-1]):
            table_rows.append(list(BLANK_LINE_ROW_TEMPLATE))
        self._populate_line_items_table_with_rows(table_rows)
        self._apply_import_review_decorations(list(preview_payload.get("lines") or []))

    def _apply_import_review_decorations(self, lines: list[dict[str, object]]) -> None:
        table = getattr(self.ui, "tbl_lineItems", None)
        if table is None:
            return

        self._import_review_rows = {}
        for row_index, line in enumerate(lines):
            if row_index >= table.rowCount():
                break
            reasons = ensure_str_list(line.get("review_reason"))
            needs_review = bool(line.get("needs_review"))
            if not needs_review and not reasons:
                continue

            is_high_severity = not line.get("production_item") or not line.get("material")
            marker_color = "#dc2626" if is_high_severity else "#d97706"
            tooltip_parts = [f"來源檔名：{line.get('source_file', '')}"]
            if reasons:
                tooltip_parts.append("需要確認：" + "；".join(reasons))
            notes = ensure_str_list(line.get("notes"))
            if notes:
                tooltip_parts.append("AI 備註：" + "；".join(notes))
            tooltip = "\n".join(part for part in tooltip_parts if part)
            self._import_review_rows[row_index] = {"tooltip": tooltip, "line": line}

            for column in range(table.columnCount()):
                item = table.item(row_index, column)
                if item is not None:
                    item.setToolTip(tooltip)
                widget = table.cellWidget(row_index, column)
                if widget is not None:
                    widget.setToolTip(tooltip)

            marker_widget = table.cellWidget(row_index, 0)
            marker_style = (
                "QLineEdit {background-color: #fef3c7; border: 1px solid #f59e0b; "
                f"border-left: 4px solid {marker_color}; border-radius: 4px; padding-left: 6px;"
                "}"
            )
            if isinstance(marker_widget, QLineEdit):
                marker_widget.setStyleSheet(marker_style)
            elif isinstance(marker_widget, QComboBox) and marker_widget.lineEdit() is not None:
                marker_widget.lineEdit().setStyleSheet(marker_style)

    def _build_table_row_from_import_line(self, line: dict[str, object]) -> list[str]:
        size_text = str(line.get("size_text") or "")
        width_text, height_text = split_size_text_for_table(size_text)
        return [
            str(line.get("production_item") or ""),
            width_text,
            "x",
            height_text,
            str(line.get("quantity") or 1),
            str(line.get("material") or ""),
            str(line.get("lamination") or ""),
            str(line.get("board_type") or ""),
            str(line.get("board_thickness") or ""),
            "+".join(line.get("extra_materials") or []),
            "",
            "",
            "",
            "",
            "",
        ]

    def _row_values_have_meaningful_content(self, row_values: list[str]) -> bool:
        for column in [0, 1, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14]:
            if column < len(row_values) and str(row_values[column]).strip():
                return True
        return False

    def save_work_order_with_lines(self) -> tuple[int, int]:
        if pymysql is None:
            raise RuntimeError("缺少 PyMySQL，無法儲存到 work_order_v2。")

        self.calculate_document_totals()
        header_payload = self._header_payload()
        line_payloads = self._collect_line_payloads()
        with pymysql.connect(**_connect_kwargs()) as conn:
            with conn.cursor() as cur:
                work_order_id = self._upsert_work_order_header(cur, header_payload)
                self._replace_work_order_lines(cur, work_order_id, line_payloads)
            conn.commit()
        self._loaded_work_order_number = str(header_payload.get("work_number") or "").strip()
        return work_order_id, len(line_payloads)

    def _find_existing_work_order_id(self, work_number: str) -> int | None:
        if pymysql is None:
            return None

        normalized_work_number = work_number.strip()
        if not normalized_work_number:
            return None

        with pymysql.connect(**_connect_kwargs()) as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT id FROM work_orders WHERE work_number = %s LIMIT 1", (normalized_work_number,))
                row = cur.fetchone()
        return int(row["id"]) if row else None

    def _confirm_overwrite_existing_work_order(self, work_number: str) -> bool:
        existing_work_order_id = self._find_existing_work_order_id(work_number)
        if existing_work_order_id is None:
            return True

        confirm = QMessageBox.question(
            self,
            "覆蓋既有工單",
            f"工單號 {work_number} 已經有既有資料（work_orders #{existing_work_order_id}）。\n\n是否要覆蓋？",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No,
        )
        if confirm != QMessageBox.StandardButton.Yes:
            self._set_status_message("已取消覆蓋既有工單。")
            return False
        return True

    def _handle_save_clicked(self) -> None:
        work_number = normalize_line_edit_value(getattr(self.ui, "le_worknum", None)) or ""
        try:
            if not self._confirm_overwrite_existing_work_order(work_number):
                return
            work_order_id, line_count = self.save_work_order_with_lines()
        except Exception as exc:
            message = str(exc) or exc.__class__.__name__
            QMessageBox.warning(self, "儲存失敗", message)
            self._set_status_message(f"儲存失敗：{message}")
            return

        work_number = work_number or "-"
        success_message = f"已儲存 work_orders #{work_order_id}（工單 {work_number}，明細 {line_count} 筆）"
        QMessageBox.information(self, "儲存成功", success_message)
        self._set_status_message(success_message)

    def _refresh_client_cache(self, *, preferred_short_name: str | None = None) -> None:
        current_text = ""
        combo = getattr(self.ui, "cb_customerName", None)
        if isinstance(combo, QComboBox):
            current_text = combo.currentText().strip()

        self.clients = load_clients_from_v2()
        self.client_rows_by_short_name = {
            str(row["short_name"]): row for row in self.clients if row.get("short_name")
        }

        if not isinstance(combo, QComboBox):
            return

        selected_text = (preferred_short_name or current_text).strip()
        combo.blockSignals(True)
        combo.clear()
        for row in self.clients:
            combo.addItem(str(row["short_name"]), row)
        combo.blockSignals(False)

        if selected_text:
            target_index = combo.findText(selected_text)
            if target_index >= 0:
                combo.setCurrentIndex(target_index)
            else:
                combo.setEditText(selected_text)
            self._handle_customer_name_changed(selected_text)

    def _build_add_customer_dialog(self) -> tuple[QDialog, dict[str, QLineEdit]]:
        dialog = QDialog(self)
        dialog.setWindowTitle("新增客戶")
        dialog.setModal(True)
        dialog.resize(420, 0)

        layout = QVBoxLayout(dialog)
        form = QFormLayout()
        form.setFieldGrowthPolicy(QFormLayout.FieldGrowthPolicy.ExpandingFieldsGrow)
        form.setLabelAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        layout.addLayout(form)

        combo = getattr(self.ui, "cb_customerName", None)
        initial_short_name = combo.currentText().strip() if isinstance(combo, QComboBox) else ""
        initial_phone = normalize_line_edit_value(getattr(self.ui, "le_phone", None))
        initial_address = normalize_line_edit_value(getattr(self.ui, "lle_address", None))

        fields = {
            "short_name": QLineEdit(initial_short_name, dialog),
            "full_name": QLineEdit(initial_short_name, dialog),
            "phone": QLineEdit(initial_phone, dialog),
            "address": QLineEdit(initial_address, dialog),
            "tax_id": QLineEdit("", dialog),
        }
        labels = {
            "short_name": "客戶簡稱",
            "full_name": "客戶全名",
            "phone": "電話",
            "address": "地址",
            "tax_id": "統編",
        }
        for name, widget in fields.items():
            widget.setFont(build_input_font(12))
            form.addRow(labels[name], widget)

        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel, parent=dialog)
        buttons.button(QDialogButtonBox.StandardButton.Ok).setText("儲存")
        buttons.button(QDialogButtonBox.StandardButton.Cancel).setText("取消")
        buttons.accepted.connect(dialog.accept)
        buttons.rejected.connect(dialog.reject)
        layout.addWidget(buttons)
        return dialog, fields

    def _insert_client_record(self, payload: dict[str, str | None]) -> None:
        if pymysql is None:
            raise RuntimeError("缺少 PyMySQL，無法新增客戶資料。")

        short_name = (payload.get("short_name") or "").strip()
        if not short_name:
            raise ValueError("客戶簡稱不可為空。")

        with pymysql.connect(**_connect_kwargs()) as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT id FROM clients WHERE short_name = %s LIMIT 1", (short_name,))
                existing = cur.fetchone()
                if existing is not None:
                    raise ValueError(f"客戶簡稱已存在：{short_name}")
                cur.execute(
                    """
                    INSERT INTO clients (short_name, full_name, phone, address, tax_id)
                    VALUES (%s, %s, %s, %s, %s)
                    """,
                    (
                        short_name,
                        payload.get("full_name") or None,
                        payload.get("phone") or None,
                        payload.get("address") or None,
                        payload.get("tax_id") or None,
                    ),
                )
            conn.commit()

    def _handle_add_customer_clicked(self) -> None:
        dialog, fields = self._build_add_customer_dialog()
        short_name_widget = fields["short_name"]
        short_name_widget.setFocus(Qt.FocusReason.ActiveWindowFocusReason)
        short_name_widget.selectAll()
        if dialog.exec() != QDialog.DialogCode.Accepted:
            self._set_status_message("已取消新增客戶。")
            return

        payload = {name: widget.text().strip() or None for name, widget in fields.items()}
        if payload.get("short_name") and not payload.get("full_name"):
            payload["full_name"] = payload["short_name"]
        try:
            self._insert_client_record(payload)
            self._refresh_client_cache(preferred_short_name=str(payload.get("short_name") or ""))
        except Exception as exc:
            message = str(exc) or exc.__class__.__name__
            QMessageBox.warning(self, "新增客戶失敗", message)
            self._set_status_message(f"新增客戶失敗：{message}")
            return

        short_name = str(payload.get("short_name") or "")
        QMessageBox.information(self, "新增客戶成功", f"已新增客戶：{short_name}")
        self._set_status_message(f"已新增客戶：{short_name}")

    def _delete_work_order_record(self, work_number: str) -> int:
        if pymysql is None:
            raise RuntimeError("缺少 PyMySQL，無法刪除工單。")

        normalized_work_number = work_number.strip()
        if not normalized_work_number:
            raise ValueError("請先輸入工單號，再刪除工單。")

        with pymysql.connect(**_connect_kwargs()) as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT id FROM work_orders WHERE work_number = %s LIMIT 1", (normalized_work_number,))
                row = cur.fetchone()
                if row is None:
                    raise LookupError(f"找不到工單號：{normalized_work_number}")
                work_order_id = int(row["id"])
                cur.execute("DELETE FROM work_orders WHERE id = %s", (work_order_id,))
            conn.commit()
        return work_order_id

    def _handle_delete_work_order_clicked(self) -> None:
        work_number = normalize_line_edit_value(getattr(self.ui, "le_worknum", None)) or ""
        existing_work_order_id = self._find_existing_work_order_id(work_number)
        if existing_work_order_id is None:
            message = "請先輸入或開啟有效工單號，再刪除工單。"
            QMessageBox.warning(self, "刪除工單失敗", message)
            self._set_status_message(message)
            return

        confirm = QMessageBox.question(
            self,
            "刪除工單",
            f"確定要刪除工單 {work_number}（work_orders #{existing_work_order_id}）嗎？\n\n這會一併刪掉所有明細列，而且無法復原。",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No,
        )
        if confirm != QMessageBox.StandardButton.Yes:
            self._set_status_message("已取消刪除工單。")
            return

        try:
            deleted_work_order_id = self._delete_work_order_record(work_number)
        except Exception as exc:
            message = str(exc) or exc.__class__.__name__
            QMessageBox.warning(self, "刪除工單失敗", message)
            self._set_status_message(f"刪除工單失敗：{message}")
            return

        self.reset_work_order_to_blank()
        self._loaded_work_order_number = ""
        success_message = f"已刪除工單 {work_number}（work_orders #{deleted_work_order_id}）。"
        QMessageBox.information(self, "刪除工單成功", success_message)
        self._set_status_message(success_message)

    def _apply_customer_contact_defaults(self, row: dict[str, str | int | None]) -> None:
        phone_widget = getattr(self.ui, "le_phone", None)
        if not isinstance(phone_widget, QLineEdit):
            return

        next_phone = str(row.get("phone") or "")
        current_phone = phone_widget.text().strip()
        can_fill_phone = not current_phone or current_phone == self._last_auto_filled_phone

        if next_phone and can_fill_phone:
            phone_widget.setText(next_phone)
            self._last_auto_filled_phone = next_phone
        elif not next_phone and current_phone == self._last_auto_filled_phone:
            phone_widget.clear()
            self._last_auto_filled_phone = ""

    def _tune_generated_layout(self) -> None:
        main_layout = getattr(self.ui, "verticalLayout_2", None)
        line_items_section = getattr(self.ui, "wdg_lineItems", None)
        line_items_layout = getattr(self.ui, "verticalLayout", None)
        bottom_section = getattr(self.ui, "wdg_bottomSection", None)
        bottom_container = getattr(self.ui, "widget10", None)
        bottom_layout = getattr(self.ui, "horizontalLayout_17", None)
        remark_group = getattr(self.ui, "grp_remark", None)
        summary_actions = getattr(self.ui, "wdg_summaryActions", None)
        remark_editor = getattr(self.ui, "te_remark", None)
        if bottom_layout is None or remark_group is None or summary_actions is None:
            return

        self._tune_line_items_table()

        if main_layout is not None:
            main_layout.setSpacing(0)
            main_layout.setStretch(0, 0)
            main_layout.setStretch(1, 1)
            main_layout.setStretch(2, 0)

        if line_items_section is not None:
            line_items_section.setMinimumHeight(0)
            line_items_section.setSizePolicy(QSizePolicy.Policy.Preferred, QSizePolicy.Policy.Expanding)

        top_content = getattr(self.ui, "wdg_topContent", None)
        if top_content is not None:
            top_content.setMinimumHeight(81)
            top_content.setMaximumHeight(81)
        if line_items_layout is not None:
            line_items_layout.setContentsMargins(0, 0, 0, 0)
            line_items_layout.setSpacing(0)

        if bottom_section is not None:
            bottom_section.setMinimumHeight(156)
            bottom_section.setMaximumHeight(156)
        if bottom_container is not None:
            bottom_container.setGeometry(20, 0, 1481, 156)

        remark_group.setMaximumWidth(16777215)
        remark_group.setMinimumWidth(600)
        remark_group.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Preferred)
        remark_group.setMinimumHeight(152)
        remark_group.setMaximumHeight(152)
        remark_layout = getattr(self.ui, "horizontalLayout_9", None)
        if remark_layout is not None:
            remark_layout.setContentsMargins(6, 6, 6, 6)
        if remark_editor is not None:
            remark_editor.setMinimumHeight(114)
            remark_editor.setMaximumHeight(114)
        self._ensure_build_label(remark_group)
        self._position_build_label()

        amount_summary = getattr(self.ui, "wdg_amountSummary", None)
        action_buttons = getattr(self.ui, "wdg_actionButtons", None)
        amount_layout = getattr(self.ui, "horizontalLayout_13", None)
        top_buttons_layout = getattr(self.ui, "horizontalLayout_14", None)
        lower_buttons_layout = getattr(self.ui, "horizontalLayout_15", None)
        button_rows_layout = getattr(self.ui, "horizontalLayout_16", None)
        top_group = getattr(self.ui, "widget14", None)

        summary_outer_width = 718
        summary_inner_width = 706
        button_group_width = 674

        if amount_summary is not None:
            amount_summary.setGeometry(10, 4, summary_inner_width, 54)
        if action_buttons is not None:
            action_buttons.setGeometry(10, 58, summary_inner_width, 82)
        if top_group is not None:
            top_group.setGeometry(4, 0, button_group_width, 80)

        if amount_layout is not None:
            amount_layout.setContentsMargins(0, 0, 0, 0)
            amount_layout.setSpacing(6)

        summary_specs = [
            ("wdg_productionAmountField", "widget11", "lbl_productionAmount", "le_productionAmount", 84, 130, 220),
            ("wdg_taxAmountField", "widget12", "lbl_taxAmount", "le_taxAmount", 56, 152, 218),
            ("wdg_totalAmountField", "widget13", "lbl_totalAmount", "le_totalAmount", 56, 152, 218),
        ]
        for field_name, inner_name, label_name, lineedit_name, label_width, lineedit_width, inner_width in summary_specs:
            field = getattr(self.ui, field_name, None)
            inner = getattr(self.ui, inner_name, None)
            label = getattr(self.ui, label_name, None)
            lineedit = getattr(self.ui, lineedit_name, None)
            if field is not None:
                field.setMinimumWidth(inner_width + 20)
                field.setMaximumWidth(inner_width + 20)
                field.setSizePolicy(QSizePolicy.Policy.Fixed, QSizePolicy.Policy.Preferred)
            if inner is not None:
                inner.setGeometry(10, 10, inner_width, 30)
            if label is not None:
                label.setMinimumWidth(label_width)
                label.setMaximumWidth(label_width)
                label.setSizePolicy(QSizePolicy.Policy.Fixed, QSizePolicy.Policy.Preferred)
                label.setStyleSheet(
                    f"background-color: {SUMMARY_LABEL_BG}; color: {SUMMARY_LABEL_FG}; "
                    f"border: 1px solid {SUMMARY_LABEL_BORDER}; border-radius: 4px; "
                    "font-weight: 600; padding: 0 10px;"
                )
                label.setAlignment(Qt.AlignmentFlag.AlignCenter)
            if lineedit is not None:
                lineedit.setMinimumWidth(lineedit_width)
                lineedit.setMaximumWidth(lineedit_width)
                lineedit.setMinimumHeight(30)
                lineedit.setTextMargins(8, 0, 10, 0)
                lineedit.setSizePolicy(QSizePolicy.Policy.Fixed, QSizePolicy.Policy.Fixed)
                lineedit.setAlignment(Qt.AlignmentFlag.AlignRight)
                if lineedit_name in SUMMARY_LOCKED_FIELD_NAMES:
                    lineedit.setReadOnly(True)
                    lineedit.setFocusPolicy(Qt.FocusPolicy.NoFocus)

        for layout_name in ("horizontalLayout_10", "horizontalLayout_11", "horizontalLayout_12"):
            field_layout = getattr(self.ui, layout_name, None)
            if field_layout is not None:
                field_layout.setContentsMargins(0, 0, 0, 0)
                field_layout.setSpacing(6)

        button_specs = {
            "btn_open": 106,
            "btn_save": 106,
            "btn_reset": 106,
            "btn_billing": 106,
            "btn_subtotal": 106,
            "btn_calcuate": 106,
            "btn_import": 106,
            "btn_invoice": 106,
            "btn_add_customer": 106,
            "btn_delete_work_order": 106,
        }
        for button_name, width in button_specs.items():
            button = getattr(self.ui, button_name, None)
            if button is not None:
                button.setMinimumWidth(width)
                button.setMaximumWidth(width)
                button.setMinimumHeight(38)
                button.setMaximumHeight(38)
                button.setSizePolicy(QSizePolicy.Policy.Fixed, QSizePolicy.Policy.Fixed)

        if top_buttons_layout is not None:
            top_buttons_layout.setContentsMargins(0, 0, 0, 0)
            top_buttons_layout.setSpacing(4)
        if lower_buttons_layout is not None:
            lower_buttons_layout.setContentsMargins(2, 2, 0, 0)
            lower_buttons_layout.setSpacing(6)
            lower_buttons_layout.addStretch(1)
        if button_rows_layout is not None:
            button_rows_layout.setContentsMargins(0, 0, 0, 0)
            button_rows_layout.setSpacing(0)
            button_rows_layout.setStretch(0, 0)
            button_rows_layout.setStretch(1, 1)

        summary_actions.setMinimumWidth(summary_outer_width)
        summary_actions.setMaximumWidth(summary_outer_width)
        summary_actions.setMinimumHeight(152)
        summary_actions.setMaximumHeight(152)
        summary_actions.setSizePolicy(QSizePolicy.Policy.Preferred, QSizePolicy.Policy.Fixed)
        bottom_layout.setAlignment(summary_actions, Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignTop)
        bottom_layout.setSpacing(2)
        bottom_layout.setStretch(0, 5)
        bottom_layout.setStretch(1, 0)

    def _ensure_build_label(self, remark_group: QWidget) -> None:
        if self.build_label is not None:
            return

        self.build_label = QLabel(BUILD_LABEL_TEXT, remark_group)
        self.build_label.setObjectName("lbl_buildVersion")
        self.build_label.setFocusPolicy(Qt.FocusPolicy.NoFocus)
        self.build_label.setAttribute(Qt.WidgetAttribute.WA_TransparentForMouseEvents, True)
        self.build_label.setStyleSheet(
            "color: rgba(32, 33, 36, 0.55); "
            "background-color: rgba(245, 245, 245, 0.88); "
            "border: 1px solid rgba(201, 177, 234, 0.65); "
            "border-radius: 4px; "
            "padding: 1px 6px;"
        )
        self.build_label.adjustSize()
        self.build_label.raise_()

    def _position_build_label(self) -> None:
        if self.build_label is None:
            return

        label_margin_left = 14
        label_margin_bottom = 8
        x = label_margin_left
        y = max(24, self.ui.grp_remark.height() - self.build_label.sizeHint().height() - label_margin_bottom)
        self.build_label.move(x, y)

    def resizeEvent(self, event) -> None:
        super().resizeEvent(event)
        self._position_build_label()

    def _ordered_focus_widgets(self, names: list[str]) -> list[QWidget]:
        widgets: list[QWidget] = []
        for name in names:
            widget = getattr(self.ui, name, None)
            if isinstance(widget, QWidget) and widget.focusPolicy() != Qt.FocusPolicy.NoFocus:
                widgets.append(widget)
        return widgets

    def _tab_target_widget(self, widget: QWidget) -> QWidget:
        if isinstance(widget, QComboBox):
            widget.setFocusPolicy(Qt.FocusPolicy.StrongFocus)
            line_edit = widget.lineEdit()
            if line_edit is not None:
                line_edit.setFocusPolicy(Qt.FocusPolicy.StrongFocus)
            # Editable QComboBox 會把焦點代理到內部 lineEdit。
            # Tab chain 必須綁在 combo 本體上，Qt 才會正確停在這個欄位，
            # 並在離開後接到下一個外部 widget，而不是把 customerName 略過。
            return widget
        return widget

    def _redirect_x_focus(self, row: int) -> bool:
        if self.table_tab_navigator is None:
            return False
        return self.table_tab_navigator.redirect_x_cell(row)

    def _sanitize_table_current_cell(self, row: int, column: int) -> None:
        if column != X_COLUMN_INDEX:
            return
        self._redirect_x_focus(row)

    def _sanitize_current_table_focus(self) -> bool:
        table = getattr(self.ui, "tbl_lineItems", None)
        if table is None:
            return False
        if table.currentColumn() != X_COLUMN_INDEX:
            return False
        return self._redirect_x_focus(table.currentRow())

    def _configure_focus_chain(self) -> None:
        top_widgets = self._ordered_focus_widgets(self.TOP_TAB_ORDER)
        bottom_widgets = self._ordered_focus_widgets(self.BOTTOM_TAB_ORDER)
        table = getattr(self.ui, "tbl_lineItems", None)
        ordered_widgets = [self._tab_target_widget(widget) for widget in top_widgets]
        if isinstance(table, QWidget) and table.focusPolicy() != Qt.FocusPolicy.NoFocus:
            ordered_widgets.append(table)
        ordered_widgets.extend(self._tab_target_widget(widget) for widget in bottom_widgets)

        for first, second in zip(ordered_widgets, ordered_widgets[1:]):
            QWidget.setTabOrder(first, second)

        for widget in top_widgets:
            widget.installEventFilter(self)

        if table is not None:
            table.installEventFilter(self)
            table.currentCellChanged.connect(self._sanitize_table_current_cell)

    def eventFilter(self, watched: QObject, event: QEvent) -> bool:
        if event.type() == QEvent.Type.KeyPress and isinstance(event, QKeyEvent):
            if watched is getattr(self.ui, "le_endTime", None) and event.key() in (Qt.Key.Key_Tab, Qt.Key.Key_Backtab):
                forward = event.key() == Qt.Key.Key_Tab and not bool(event.modifiers() & Qt.KeyboardModifier.ShiftModifier)
                if forward:
                    self.focus_first_table_cell()
                    return True

            if watched is getattr(self.ui, "tbl_lineItems", None) and event.key() in (Qt.Key.Key_Tab, Qt.Key.Key_Backtab):
                if self._sanitize_current_table_focus():
                    return True
                forward = event.key() == Qt.Key.Key_Tab and not bool(event.modifiers() & Qt.KeyboardModifier.ShiftModifier)
                return self._focus_table_boundary_cell(forward=forward)

        return super().eventFilter(watched, event)

    def _first_focusable_table_cell(self) -> tuple[int, int] | None:
        if self.table_tab_navigator is None:
            return None
        return self.table_tab_navigator._find_next_editable_cell(0, -1, forward=True)

    def _last_focusable_table_cell(self) -> tuple[int, int] | None:
        if self.table_tab_navigator is None:
            return None
        table = getattr(self.ui, "tbl_lineItems", None)
        if table is None:
            return None
        return self.table_tab_navigator._find_next_editable_cell(table.rowCount() - 1, table.columnCount(), forward=False)

    def _focus_table_boundary_cell(self, *, forward: bool) -> bool:
        table = getattr(self.ui, "tbl_lineItems", None)
        navigator = self.table_tab_navigator
        if table is None or navigator is None:
            return False

        current_row = table.currentRow()
        current_column = table.currentColumn()
        has_current_cell = current_row >= 0 and current_column >= 0

        if has_current_cell and current_column == X_COLUMN_INDEX:
            return navigator.redirect_x_cell(current_row)

        if has_current_cell and navigator._is_focusable_cell(current_row, current_column):
            target_cell = navigator._find_next_editable_cell(current_row, current_column, forward=forward)
            if target_cell is not None:
                row, column = target_cell
                table.setCurrentCell(row, column)
                QTimer.singleShot(0, lambda: navigator._focus_cell(row, column))
                return True

            self.focus_after_table(forward=forward)
            return True

        boundary_cell = self._first_focusable_table_cell() if forward else self._last_focusable_table_cell()
        if boundary_cell is None:
            self.focus_after_table(forward=forward)
            return True

        row, column = boundary_cell
        table.setFocus(Qt.FocusReason.TabFocusReason if forward else Qt.FocusReason.BacktabFocusReason)
        table.setCurrentCell(row, column)
        QTimer.singleShot(0, lambda: navigator._focus_cell(row, column))
        return True

    def focus_first_table_cell(self) -> None:
        table = getattr(self.ui, "tbl_lineItems", None)
        first_cell = self._first_focusable_table_cell()
        if table is None or first_cell is None or self.table_tab_navigator is None:
            return

        row, column = first_cell
        table.setFocus(Qt.FocusReason.TabFocusReason)
        table.setCurrentCell(row, column)
        QTimer.singleShot(0, lambda: self.table_tab_navigator._focus_cell(row, column))

    def focus_last_top_field(self) -> None:
        top_widgets = self._ordered_focus_widgets(self.TOP_TAB_ORDER)
        if top_widgets:
            top_widgets[-1].setFocus(Qt.FocusReason.BacktabFocusReason)

    def focus_after_table(self, *, forward: bool) -> None:
        if not forward:
            self.focus_last_top_field()
            return

        bottom_widgets = self._ordered_focus_widgets(self.BOTTOM_TAB_ORDER)
        if bottom_widgets:
            bottom_widgets[0].setFocus(Qt.FocusReason.TabFocusReason)

    def _tune_line_items_table(self) -> None:
        table = getattr(self.ui, "tbl_lineItems", None)
        if table is None:
            return

        table.clear()
        table.setColumnCount(len(TABLE_HEADERS))
        self.table_tab_navigator = TableCellTabNavigator(table, self)
        table.setItemDelegate(TableItemDelegate(self.table_tab_navigator, table))
        table.setHorizontalHeader(BandHeaderView(Qt.Orientation.Horizontal, table))
        table.setHorizontalHeaderLabels(TABLE_HEADERS)
        self._apply_line_items_header_colors(table)
        table.setRowCount(DEFAULT_LINE_ITEM_ROW_COUNT)
        table.setAlternatingRowColors(True)
        table.setWordWrap(False)
        table.setShowGrid(True)
        table.setCornerButtonEnabled(False)
        table.setSelectionBehavior(table.SelectionBehavior.SelectItems)
        table.setSelectionMode(table.SelectionMode.SingleSelection)
        table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        table.setTabKeyNavigation(False)
        table.setMinimumHeight(418)
        table.setStyleSheet(
            """
            QTableWidget {
                border: 1px solid #b9c0c7;
                background: #ffffff;
                gridline-color: #bcc3ca;
                alternate-background-color: #fafafa;
                selection-background-color: #dbe7ff;
                selection-color: #202124;
            }
            QTableWidget::item {
                padding: 4px 6px;
                border-right: 1px solid #d7dce1;
                border-bottom: 1px solid #d7dce1;
            }
            QHeaderView::section {
                background: #ececec;
                color: #202124;
                border-top: 1px solid #c4c8cc;
                border-left: 1px solid #c4c8cc;
                border-right: 1px solid #c4c8cc;
                border-bottom: 1px solid #adb3b9;
                padding: 6px 4px;
                font-weight: 700;
            }
            """
        )

        header = table.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        header.setMinimumSectionSize(28)
        header.setDefaultAlignment(Qt.AlignmentFlag.AlignCenter)
        header.setStretchLastSection(False)
        header.setHighlightSections(False)
        header.setFixedHeight(34)

        v_header = table.verticalHeader()
        v_header.setVisible(False)
        v_header.setDefaultSectionSize(36)
        v_header.setMinimumSectionSize(36)

        for index, width in enumerate(TABLE_COLUMN_WIDTHS):
            table.setColumnWidth(index, width)

        for row in range(table.rowCount()):
            table.setRowHeight(row, 36)

    def _apply_line_items_header_colors(self, table) -> None:
        for column in HEADER_YELLOW_COLUMNS:
            item = table.horizontalHeaderItem(column)
            if item is not None:
                item.setBackground(HEADER_YELLOW_BG)
                item.setForeground(HEADER_YELLOW_FG)
                item.setFont(build_table_header_font(column))

        for column in HEADER_PURPLE_COLUMNS:
            item = table.horizontalHeaderItem(column)
            if item is not None:
                item.setBackground(HEADER_PURPLE_BG)
                item.setForeground(HEADER_PURPLE_FG)
                item.setFont(build_table_header_font(column))

    def _make_table_item(self, text: str, column: int) -> QTableWidgetItem:
        item = QTableWidgetItem(text)
        item.setForeground(QColor("#202124"))
        item.setFont(build_table_cell_font(column))

        if column == X_COLUMN_INDEX:
            item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            item.setFlags(Qt.ItemFlag.NoItemFlags)
            item.setBackground(QColor("#efefef"))
            item.setForeground(QColor("#6b7280"))
        elif column in TABLE_LOCKED_COLUMN_INDEXES:
            item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)
            item.setBackground(QColor("#f5f5f5"))
            item.setForeground(QColor("#6b7280"))
        elif column in NUMERIC_COLUMN_INDEXES:
            item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
        elif column in SELECT_LIKE_COLUMN_INDEXES:
            item.setTextAlignment(Qt.AlignmentFlag.AlignVCenter | Qt.AlignmentFlag.AlignLeft)
        else:
            item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)

        return item

    def _make_combo_box(self, column: int, current_text: str) -> QComboBox:
        combo = QComboBox()
        disable_combobox_mouse_wheel(combo)
        combo.addItems(self.combo_column_options[column])
        configure_combo_autocomplete(combo)
        combo.setFrame(True)
        combo.setMaxVisibleItems(8)
        combo.setMinimumHeight(28)
        combo.setMaximumHeight(28)
        combo_font = build_table_cell_font(column)
        combo.setFont(combo_font)
        if combo.lineEdit() is not None:
            combo.lineEdit().setFont(combo_font)
        if combo.view() is not None:
            combo.view().setFont(combo_font)

        current_index = combo.findText(current_text)
        if current_index >= 0:
            combo.setCurrentIndex(current_index)
        elif current_text:
            combo.setEditText(current_text)
        else:
            combo.setCurrentIndex(-1)
            combo.clearEditText()
        return combo

    def _make_line_edit(self, column: int, text: str) -> QLineEdit:
        line_edit = QLineEdit(text)
        line_edit.setFrame(False)
        line_edit.setMinimumHeight(28)
        line_edit.setMaximumHeight(28)
        line_edit.setTextMargins(6, 0, 6, 0)
        line_edit.setFont(build_table_cell_font(column))
        if column in NUMERIC_COLUMN_INDEXES:
            line_edit.setAlignment(Qt.AlignmentFlag.AlignCenter)
        else:
            line_edit.setAlignment(Qt.AlignmentFlag.AlignVCenter | Qt.AlignmentFlag.AlignLeft)
        return line_edit

    def _make_widget_backing_item(self, column: int) -> QTableWidgetItem:
        item = self._make_table_item("", column)
        item.setFlags(Qt.ItemFlag.ItemIsSelectable | Qt.ItemFlag.ItemIsEnabled)
        return item

    def _register_table_cell_widget(self, widget: QWidget, row: int, column: int) -> QWidget:
        if self.table_tab_navigator is not None:
            self.table_tab_navigator.register_widget(widget, row, column)
        return widget

    def _populate_single_line_item_row(self, row_idx: int, row: list[str] | None = None) -> None:
        table = getattr(self.ui, "tbl_lineItems", None)
        if table is None:
            return

        row_values = list(row) if row is not None else list(BLANK_LINE_ROW_TEMPLATE)
        if len(row_values) < len(TABLE_HEADERS):
            row_values.extend([""] * (len(TABLE_HEADERS) - len(row_values)))
        if len(row_values) > len(TABLE_HEADERS):
            row_values = row_values[: len(TABLE_HEADERS)]
        row_values[X_COLUMN_INDEX] = "x"

        for col_idx, value in enumerate(row_values):
            if col_idx in COMBO_COLUMN_OPTIONS:
                widget = self._make_combo_box(col_idx, value)
                self._connect_auto_append_signal(widget)
                table.setCellWidget(row_idx, col_idx, self._register_table_cell_widget(widget, row_idx, col_idx))
                table.setItem(row_idx, col_idx, self._make_widget_backing_item(col_idx))
            elif col_idx in TABLE_SKIP_FOCUS_COLUMN_INDEXES:
                table.setItem(row_idx, col_idx, self._make_table_item(value, col_idx))
            else:
                widget = self._make_line_edit(col_idx, value)
                self._connect_auto_append_signal(widget)
                table.setCellWidget(row_idx, col_idx, self._register_table_cell_widget(widget, row_idx, col_idx))
                table.setItem(row_idx, col_idx, self._make_widget_backing_item(col_idx))
        table.setRowHeight(row_idx, 36)

    def _connect_auto_append_signal(self, widget: QWidget) -> None:
        if isinstance(widget, QComboBox):
            widget.currentTextChanged.connect(self._handle_line_item_widget_changed)
            line_edit = widget.lineEdit()
            if line_edit is not None:
                line_edit.textChanged.connect(self._handle_line_item_widget_changed)
        elif isinstance(widget, QLineEdit):
            widget.textChanged.connect(self._handle_line_item_widget_changed)

    def _append_blank_line_item_row(self) -> int:
        table = getattr(self.ui, "tbl_lineItems", None)
        if table is None:
            return -1

        row_idx = table.rowCount()
        table.insertRow(row_idx)
        self._populate_single_line_item_row(row_idx, list(BLANK_LINE_ROW_TEMPLATE))
        return row_idx

    def _ensure_trailing_blank_line(self) -> bool:
        table = getattr(self.ui, "tbl_lineItems", None)
        if table is None or table.rowCount() <= 0:
            return False

        last_row = table.rowCount() - 1
        if self._line_row_has_meaningful_data(last_row):
            self._append_blank_line_item_row()
            return True
        return False

    def _handle_line_item_widget_changed(self, _value: str) -> None:
        if self._suspend_auto_append_checks:
            return
        self._ensure_trailing_blank_line()

    def _table_all_rows_snapshot(self) -> list[list[str]]:
        table = getattr(self.ui, "tbl_lineItems", None)
        if table is None:
            return []
        return [
            [self._table_cell_text(row, column) for column in range(table.columnCount())]
            for row in range(table.rowCount())
        ]

    def _apply_row_values(self, row_index: int, row_values: list[str]) -> None:
        table = getattr(self.ui, "tbl_lineItems", None)
        if table is None or row_index < 0 or row_index >= table.rowCount():
            return

        normalized_values = list(row_values[: len(TABLE_HEADERS)])
        if len(normalized_values) < len(TABLE_HEADERS):
            normalized_values.extend([""] * (len(TABLE_HEADERS) - len(normalized_values)))
        normalized_values[X_COLUMN_INDEX] = "x"

        for column, value in enumerate(normalized_values):
            if column == X_COLUMN_INDEX:
                item = table.item(row_index, column)
                if item is not None:
                    item.setText("x")
                continue

            cell_widget = table.cellWidget(row_index, column)
            if isinstance(cell_widget, QComboBox):
                found_index = cell_widget.findText(value)
                if found_index >= 0:
                    cell_widget.setCurrentIndex(found_index)
                else:
                    cell_widget.setEditText(value)
            elif isinstance(cell_widget, QLineEdit):
                cell_widget.setText(value)
            else:
                item = table.item(row_index, column)
                if item is not None:
                    item.setText(value)

    def clear_line_item_row(self, row_index: int) -> bool:
        table = getattr(self.ui, "tbl_lineItems", None)
        if table is None or row_index < 0 or row_index >= table.rowCount():
            return False

        if not self._line_row_has_meaningful_data(row_index):
            self._set_status_message(f"第 {row_index + 1} 列目前是空白列，無需清列。")
            return False

        self._suspend_auto_append_checks = True
        try:
            self._apply_row_values(row_index, list(BLANK_LINE_ROW_TEMPLATE))
        finally:
            self._suspend_auto_append_checks = False
        self._ensure_trailing_blank_line()
        table.setCurrentCell(row_index, 0)
        self._set_status_message(f"已清空第 {row_index + 1} 列；此空列不會存入 DB。")
        return True

    def delete_line_item_row(self, row_index: int) -> bool:
        table = getattr(self.ui, "tbl_lineItems", None)
        if table is None or row_index < 0 or row_index >= table.rowCount():
            return False

        total_rows = table.rowCount()
        is_meaningful_row = self._line_row_has_meaningful_data(row_index)
        is_last_row = row_index == total_rows - 1
        if not is_meaningful_row and is_last_row and total_rows <= DEFAULT_LINE_ITEM_ROW_COUNT:
            self._set_status_message("最後預備列會保留 1 列空白，不再往下刪。")
            return False

        rows = self._table_all_rows_snapshot()
        del rows[row_index]
        if not rows:
            rows = [list(BLANK_LINE_ROW_TEMPLATE)]

        while len(rows) < DEFAULT_LINE_ITEM_ROW_COUNT:
            rows.append(list(BLANK_LINE_ROW_TEMPLATE))

        meaningful_count = sum(1 for row in rows if any((row[column] or "").strip() for column in [0, 1, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14]))
        if not rows or any((rows[-1][column] or "").strip() for column in [0, 1, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14]):
            rows.append(list(BLANK_LINE_ROW_TEMPLATE))

        self._populate_line_items_table_with_rows(rows)
        next_row = min(row_index, max(table.rowCount() - 1, 0))
        table.setCurrentCell(next_row, 0)
        if is_meaningful_row:
            self._set_status_message(f"已刪除第 {row_index + 1} 列；其後資料已往上遞補，line_no 會依有效列重排。")
        else:
            self._set_status_message(f"已移除第 {row_index + 1} 列空白列，目前保留 {meaningful_count} 列有效資料與 1 列預備列。")
        return True

    def _show_line_row_context_menu(self, pos: QPoint) -> None:
        table = getattr(self.ui, "tbl_lineItems", None)
        if table is None:
            return

        item = table.itemAt(pos)
        row_index = item.row() if item is not None else table.rowAt(pos.y())
        if row_index < 0:
            return

        table.setCurrentCell(row_index, 0)
        is_meaningful_row = self._line_row_has_meaningful_data(row_index)
        is_last_row = row_index == table.rowCount() - 1
        can_delete = is_meaningful_row or table.rowCount() > DEFAULT_LINE_ITEM_ROW_COUNT or not is_last_row

        menu = QMenu(table)
        clear_action = QAction("清除此列", menu)
        clear_action.setEnabled(is_meaningful_row)
        clear_action.triggered.connect(lambda: self.clear_line_item_row(row_index))
        menu.addAction(clear_action)

        delete_action = QAction("刪除此列", menu)
        delete_action.setEnabled(can_delete)
        delete_action.triggered.connect(lambda: self.delete_line_item_row(row_index))
        menu.addAction(delete_action)

        menu.exec(table.viewport().mapToGlobal(pos))

    def _initialize_blank_work_order(self) -> None:
        self.setWindowTitle("project.ui preview (generated)")

        self._last_auto_filled_phone = ""
        self._last_auto_filled_address = ""
        self._loaded_work_order_number = ""

        customer_combo = getattr(self.ui, "cb_customerName", None)
        if isinstance(customer_combo, QComboBox):
            customer_combo.setCurrentIndex(-1)
            customer_combo.clearEditText()

        for attr in (
            "le_worknum",
            "le_contactName",
            "le_phone",
            "le_caseName",
            "le_startTime",
            "le_endTime",
            "lle_address",
            "le_productionAmount",
            "le_taxAmount",
            "le_totalAmount",
        ):
            widget = getattr(self.ui, attr, None)
            if widget is not None:
                widget.clear()

        if hasattr(self.ui, "te_remark"):
            self.ui.te_remark.clear()

        self._populate_line_items_table_with_rows(
            [list(BLANK_LINE_ROW_TEMPLATE) for _ in range(DEFAULT_LINE_ITEM_ROW_COUNT)]
        )

    def _populate_line_items_table_with_rows(self, rows: list[list[str]]) -> None:
        if not hasattr(self.ui, "tbl_lineItems"):
            return

        table = self.ui.tbl_lineItems
        desired_rows = max(DEFAULT_LINE_ITEM_ROW_COUNT, len(rows))
        if desired_rows <= 0:
            desired_rows = DEFAULT_LINE_ITEM_ROW_COUNT
        table.setRowCount(desired_rows)
        table.clearContents()
        self.table_tab_navigator = TableCellTabNavigator(table, self)
        self._suspend_auto_append_checks = True
        try:
            for row_idx in range(table.rowCount()):
                row = rows[row_idx] if row_idx < len(rows) else list(BLANK_LINE_ROW_TEMPLATE)
                self._populate_single_line_item_row(row_idx, row)
        finally:
            self._suspend_auto_append_checks = False
        self._ensure_trailing_blank_line()

    def _seed_demo_values(self) -> None:
        self._initialize_blank_work_order()

        customer_combo = getattr(self.ui, "cb_customerName", None)
        if isinstance(customer_combo, QComboBox):
            if self.clients:
                customer_combo.setCurrentIndex(0)
            else:
                customers = ["采月廣告有限公司", "萬榮國際", "KING", "就肆電競"]
                customer_combo.addItems(customers)
                customer_combo.setCurrentIndex(0)

        defaults = {
            "le_worknum": "26-03-29-01",
            "le_contactName": "王小姐",
            "le_phone": "02-8647-8840",
            "le_caseName": "南港展場春季活動主視覺",
            "le_startTime": "2026/03/31 09:00",
            "le_endTime": "2026/03/31 18:00",
            "lle_address": "台北市南港區經貿二路 186 號 4 樓",
            "le_productionAmount": "13600",
            "le_taxAmount": "680",
            "le_totalAmount": "14280",
        }
        for attr, value in defaults.items():
            widget = getattr(self.ui, attr, None)
            if widget is None:
                continue
            if attr in {"le_phone", "lle_address"} and widget.text().strip():
                continue
            widget.setText(value)

        if hasattr(self.ui, "te_remark"):
            self.ui.te_remark.setPlainText("現場施工前 30 分鐘需與窗口聯絡；材料依樓層分批搬運。")

        def option_at(column: int, index: int, fallback: str = "") -> str:
            options = self.combo_column_options.get(column, [])
            if options:
                return options[min(index, len(options) - 1)]
            return fallback

        sample_rows = [
            [option_at(0, 0, "大圖輸出"), "500", "x", "240", "2", option_at(5, 0, "PVC 貼圖"), option_at(6, 0, "冷裱 + 修邊"), option_at(7, 0, "KT 板"), option_at(8, 0, "5mm"), option_at(9, 0, "黑膠帶收邊"), "2", "80", "2800", "5600", "300"],
            [option_at(0, 1, "展場貼圖"), "320", "x", "260", "1", option_at(5, 1, "防水帆布"), option_at(6, 1, "雙面對裱"), option_at(7, 1, "塑鋁板"), option_at(8, 1, "3mm"), option_at(9, 1, "無"), "1", "22", "6500", "6500", "0"],
            [option_at(0, 2, "立牌製作"), "90", "x", "180", "6", option_at(5, 2, "PP 相紙"), option_at(6, 2, "冷裱"), option_at(7, 2, "豪卡板"), option_at(8, 2, "10mm"), option_at(9, 2, "腳架孔位"), "6", "67.5", "950", "5700", "480"],
            [option_at(0, 3, "桌裙布置"), "240", "x", "75", "1", option_at(5, 3, "單透布"), option_at(6, 3, "包邊處理"), option_at(7, 3, "珍珠板"), option_at(8, 3, "2mm"), option_at(9, 3, "魔鬼氈"), "1", "12.5", "1800", "1800", "180"],
        ]
        self._populate_line_items_table_with_rows(sample_rows)


def _window_header_snapshot(window: GeneratedUiPreviewWindow) -> dict[str, str]:
    customer_combo = getattr(window.ui, "cb_customerName", None)
    return {
        "work_number": normalize_line_edit_value(getattr(window.ui, "le_worknum", None)) or "",
        "case_name": normalize_line_edit_value(getattr(window.ui, "le_caseName", None)) or "",
        "customer_name": customer_combo.currentText().strip() if isinstance(customer_combo, QComboBox) else "",
        "phone": normalize_line_edit_value(getattr(window.ui, "le_phone", None)) or "",
        "contact_name": normalize_line_edit_value(getattr(window.ui, "le_contactName", None)) or "",
        "start_time": normalize_line_edit_value(getattr(window.ui, "le_startTime", None)) or "",
        "end_time": normalize_line_edit_value(getattr(window.ui, "le_endTime", None)) or "",
        "address": normalize_line_edit_value(getattr(window.ui, "lle_address", None)) or "",
        "production_amount": normalize_line_edit_value(getattr(window.ui, "le_productionAmount", None)) or "",
        "tax_amount": normalize_line_edit_value(getattr(window.ui, "le_taxAmount", None)) or "",
        "total_amount": normalize_line_edit_value(getattr(window.ui, "le_totalAmount", None)) or "",
        "remark": normalize_text_edit_value(getattr(window.ui, "te_remark", None)) or "",
    }


def _window_line_snapshot(window: GeneratedUiPreviewWindow) -> list[list[str]]:
    table = getattr(window.ui, "tbl_lineItems", None)
    if table is None:
        return []

    rows: list[list[str]] = []
    for row in range(table.rowCount()):
        if not window._line_row_has_meaningful_data(row):
            continue
        rows.append([window._table_cell_text(row, column) for column in range(table.columnCount())])
    return rows


@contextmanager
def _mock_message_boxes(
    *,
    question_response: QMessageBox.StandardButton = QMessageBox.StandardButton.Yes,
) -> dict[str, list[dict[str, str]]]:
    events: dict[str, list[dict[str, str]]] = {"question": [], "information": [], "warning": []}
    original_question = QMessageBox.question
    original_information = QMessageBox.information
    original_warning = QMessageBox.warning

    def fake_question(parent, title, text, buttons, default_button):
        events["question"].append({"title": str(title), "text": str(text)})
        return question_response

    def fake_information(parent, title, text):
        events["information"].append({"title": str(title), "text": str(text)})
        return QMessageBox.StandardButton.Ok

    def fake_warning(parent, title, text):
        events["warning"].append({"title": str(title), "text": str(text)})
        return QMessageBox.StandardButton.Ok

    QMessageBox.question = fake_question
    QMessageBox.information = fake_information
    QMessageBox.warning = fake_warning
    try:
        yield events
    finally:
        QMessageBox.question = original_question
        QMessageBox.information = original_information
        QMessageBox.warning = original_warning


def _db_bundle_snapshot(work_number: str) -> tuple[dict[str, str], list[list[str]], dict[str, int]]:
    with pymysql.connect(**_connect_kwargs()) as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT wo.id, wo.work_number, wo.case_name, c.short_name AS customer_name,
                       wo.company_phone, wo.contact_name, wo.work_time, wo.cleanup_time,
                       wo.work_address, wo.production_amount, wo.tax_amount, wo.total_amount, wo.remark
                FROM work_orders wo
                LEFT JOIN clients c ON c.id = wo.client_id
                WHERE wo.work_number = %s
                """,
                (work_number,),
            )
            header_rows = cur.fetchall()
            if len(header_rows) != 1:
                raise AssertionError(f"expected exactly 1 work_orders row for {work_number}, got {len(header_rows)}")
            header_row = header_rows[0]

            cur.execute(
                """
                SELECT wol.line_no,
                       pi.item_name AS production_item_name,
                       wol.width_mm,
                       wol.height_mm,
                       wol.quantity,
                       mi.item_name AS material_name,
                       li.item_name AS lamination_name,
                       bti.item_name AS board_type_name,
                       bthi.item_name AS board_thickness_name,
                       emi.item_name AS extra_material_name,
                       wol.extra_material_quantity,
                       wol.cbm,
                       wol.cbm_unit_price,
                       wol.line_total,
                       wol.extra_material_total,
                       wol.production_item_text,
                       wol.material_text,
                       wol.lamination_text,
                       wol.board_type_text,
                       wol.board_thickness_text,
                       wol.extra_material_text
                FROM work_order_lines wol
                LEFT JOIN option_items pi ON pi.id = wol.production_item_id
                LEFT JOIN option_items mi ON mi.id = wol.material_id
                LEFT JOIN option_items li ON li.id = wol.lamination_id
                LEFT JOIN option_items bti ON bti.id = wol.board_type_id
                LEFT JOIN option_items bthi ON bthi.id = wol.board_thickness_id
                LEFT JOIN option_items emi ON emi.id = wol.extra_material_id
                WHERE wol.work_order_id = %s
                ORDER BY wol.line_no, wol.id
                """,
                (header_row["id"],),
            )
            line_rows = cur.fetchall()
            cur.execute("SELECT COUNT(*) AS c FROM work_orders WHERE work_number = %s", (work_number,))
            header_count = int(cur.fetchone()["c"])
            cur.execute(
                "SELECT COUNT(*) AS c FROM (SELECT line_no, COUNT(*) AS dup_count FROM work_order_lines WHERE work_order_id = %s GROUP BY line_no HAVING COUNT(*) > 1) AS dup",
                (header_row["id"],),
            )
            duplicated_line_nos = int(cur.fetchone()["c"])

    header_snapshot = {
        "work_number": str(header_row.get("work_number") or ""),
        "case_name": str(header_row.get("case_name") or ""),
        "customer_name": str(header_row.get("customer_name") or ""),
        "phone": str(header_row.get("company_phone") or ""),
        "contact_name": str(header_row.get("contact_name") or ""),
        "start_time": str(header_row.get("work_time") or ""),
        "end_time": str(header_row.get("cleanup_time") or ""),
        "address": str(header_row.get("work_address") or ""),
        "production_amount": format_decimal_for_ui(header_row.get("production_amount")),
        "tax_amount": format_decimal_for_ui(header_row.get("tax_amount")),
        "total_amount": format_decimal_for_ui(header_row.get("total_amount")),
        "remark": str(header_row.get("remark") or ""),
    }
    line_snapshot = [
        [
            str(line_row.get("production_item_name") or line_row.get("production_item_text") or ""),
            format_decimal_for_ui(line_row.get("width_mm")),
            "x",
            format_decimal_for_ui(line_row.get("height_mm")),
            format_decimal_for_ui(line_row.get("quantity")),
            str(line_row.get("material_name") or line_row.get("material_text") or ""),
            str(line_row.get("lamination_name") or line_row.get("lamination_text") or ""),
            str(line_row.get("board_type_name") or line_row.get("board_type_text") or ""),
            str(line_row.get("board_thickness_name") or line_row.get("board_thickness_text") or ""),
            str(line_row.get("extra_material_name") or line_row.get("extra_material_text") or ""),
            format_decimal_for_ui(line_row.get("extra_material_quantity")),
            format_decimal_for_ui(line_row.get("cbm")),
            format_decimal_for_ui(line_row.get("cbm_unit_price")),
            format_decimal_for_ui(line_row.get("line_total")),
            format_decimal_for_ui(line_row.get("extra_material_total")),
        ]
        for line_row in line_rows
    ]
    meta = {
        "work_order_id": int(header_row["id"]),
        "header_count": header_count,
        "line_count": len(line_rows),
        "duplicated_line_nos": duplicated_line_nos,
    }
    return header_snapshot, line_snapshot, meta


def _verify_table_navigation(window: GeneratedUiPreviewWindow) -> list[tuple[int, int]]:
    navigator = window.table_tab_navigator
    table = getattr(window.ui, "tbl_lineItems", None)
    if navigator is None or table is None:
        raise AssertionError("table navigator is not ready")

    sequence: list[tuple[int, int]] = []
    cell = navigator._find_next_editable_cell(0, -1, forward=True)
    for _ in range(len(navigator.column_order) * 2):
        if cell is None:
            break
        sequence.append(cell)
        cell = navigator._find_next_editable_cell(cell[0], cell[1], forward=True)

    expected_first_row = [(0, column) for column in navigator.column_order]
    if sequence[: len(expected_first_row)] != expected_first_row:
        raise AssertionError(f"unexpected first-row tab order: {sequence[:len(expected_first_row)]}")
    if len(sequence) <= len(expected_first_row) or sequence[len(expected_first_row)] != (1, navigator.column_order[0]):
        raise AssertionError(f"unexpected row wrap after tab order: {sequence}")
    return sequence[: len(expected_first_row) + 1]


def run_roundtrip_verification(window: GeneratedUiPreviewWindow, work_number: str) -> dict[str, object]:
    work_order_id, original_line_count = window.load_work_order_by_number(work_number)

    case_widget = getattr(window.ui, "le_caseName", None)
    remark_widget = getattr(window.ui, "te_remark", None)
    contact_widget = getattr(window.ui, "le_contactName", None)
    if isinstance(case_widget, QLineEdit):
        case_widget.setText(f"{case_widget.text().strip()}｜RT".strip("｜"))
    if isinstance(contact_widget, QLineEdit):
        contact_widget.setText("陳小美")
    if hasattr(remark_widget, "setPlainText"):
        remark_widget.setPlainText("round-trip verify / edit-save-open / no duplicate lines")

    table = window.ui.tbl_lineItems

    def set_row_values(row_index: int, row_values: list[str]) -> None:
        for column, value in enumerate(row_values):
            if column == X_COLUMN_INDEX:
                continue
            cell_widget = table.cellWidget(row_index, column)
            if isinstance(cell_widget, QComboBox):
                found_index = cell_widget.findText(value)
                if found_index >= 0:
                    cell_widget.setCurrentIndex(found_index)
                else:
                    cell_widget.setEditText(value)
            elif isinstance(cell_widget, QLineEdit):
                cell_widget.setText(value)
            else:
                item = table.item(row_index, column)
                if item is not None:
                    item.setText(value)

    first_row_values = [
        window._table_cell_text(0, 0) or (window.combo_column_options.get(0) or ["大圖輸出"])[0],
        "777",
        "x",
        "333",
        "9",
        window._table_cell_text(0, 5) or (window.combo_column_options.get(5) or [""])[0],
        window._table_cell_text(0, 6) or (window.combo_column_options.get(6) or [""])[0],
        window._table_cell_text(0, 7) or (window.combo_column_options.get(7) or [""])[0],
        window._table_cell_text(0, 8) or (window.combo_column_options.get(8) or [""])[0],
        window._table_cell_text(0, 9) or (window.combo_column_options.get(9) or [""])[0],
        "3",
        "55.5",
        "1200",
        "10800",
        "240",
    ]
    set_row_values(0, first_row_values)

    before_rows = table.rowCount()
    last_row = table.rowCount() - 1
    new_row_values = [
        (window.combo_column_options.get(0) or ["大圖輸出"])[0],
        "111",
        "x",
        "222",
        "1",
        (window.combo_column_options.get(5) or [""])[0],
        (window.combo_column_options.get(6) or [""])[0],
        (window.combo_column_options.get(7) or [""])[0],
        (window.combo_column_options.get(8) or [""])[0],
        (window.combo_column_options.get(9) or [""])[0],
        "1",
        "8.5",
        "700",
        "700",
        "50",
    ]
    set_row_values(last_row, new_row_values)
    app = QApplication.instance()
    if app is not None:
        app.processEvents()
    after_rows = table.rowCount()
    if after_rows != before_rows + 1:
        raise AssertionError(f"auto append failed: before={before_rows}, after={after_rows}")

    tab_sequence = _verify_table_navigation(window)
    saved_work_order_id, saved_line_count = window.save_work_order_with_lines()
    if saved_work_order_id != work_order_id:
        raise AssertionError(f"work_order_id changed unexpectedly: {work_order_id} -> {saved_work_order_id}")

    reloaded_work_order_id, reloaded_line_count = window.load_work_order_by_number(work_number)
    ui_header = _window_header_snapshot(window)
    ui_lines = _window_line_snapshot(window)
    db_header, db_lines, db_meta = _db_bundle_snapshot(work_number)

    if reloaded_work_order_id != work_order_id:
        raise AssertionError(f"reload returned different work_order_id: {work_order_id} vs {reloaded_work_order_id}")
    if ui_header != db_header:
        raise AssertionError(f"UI/DB header mismatch: ui={ui_header} db={db_header}")
    if ui_lines != db_lines:
        raise AssertionError(f"UI/DB lines mismatch: ui={ui_lines} db={db_lines}")
    if db_meta["header_count"] != 1:
        raise AssertionError(f"duplicate work_orders header detected: {db_meta}")
    if db_meta["duplicated_line_nos"] != 0:
        raise AssertionError(f"duplicate line_no detected: {db_meta}")

    return {
        "work_order_id": work_order_id,
        "original_line_count": original_line_count,
        "saved_line_count": saved_line_count,
        "reloaded_line_count": reloaded_line_count,
        "db_line_count": db_meta["line_count"],
        "header_count": db_meta["header_count"],
        "duplicated_line_nos": db_meta["duplicated_line_nos"],
        "auto_append_rows_before": before_rows,
        "auto_append_rows_after": after_rows,
        "tab_sequence": tab_sequence,
        "build_label": BUILD_LABEL_TEXT,
    }


def run_calculation_verification(window: GeneratedUiPreviewWindow) -> dict[str, object]:
    window._initialize_blank_work_order()
    table = window.ui.tbl_lineItems

    calc_rows = [
        [
            (window.combo_column_options.get(0) or ["大圖輸出"])[0],
            "100",
            "x",
            "100",
            "1",
            (window.combo_column_options.get(5) or [""])[0],
            (window.combo_column_options.get(6) or [""])[0],
            (window.combo_column_options.get(7) or [""])[0],
            (window.combo_column_options.get(8) or [""])[0],
            (window.combo_column_options.get(9) or [""])[0],
            "1",
            "",
            "10",
            "",
            "50",
        ],
        [
            (window.combo_column_options.get(0) or ["大圖輸出"])[0],
            "100",
            "x",
            "100",
            "5",
            (window.combo_column_options.get(5) or [""])[0],
            (window.combo_column_options.get(6) or [""])[0],
            (window.combo_column_options.get(7) or [""])[0],
            (window.combo_column_options.get(8) or [""])[0],
            (window.combo_column_options.get(9) or [""])[0],
            "1",
            "",
            "20",
            "",
            "",
        ],
        [
            (window.combo_column_options.get(0) or ["大圖輸出"])[0],
            "200",
            "x",
            "150",
            "1",
            (window.combo_column_options.get(5) or [""])[0],
            (window.combo_column_options.get(6) or [""])[0],
            (window.combo_column_options.get(7) or [""])[0],
            (window.combo_column_options.get(8) or [""])[0],
            (window.combo_column_options.get(9) or [""])[0],
            "1",
            "",
            "15",
            "",
            "abc",
        ],
        [
            (window.combo_column_options.get(0) or ["大圖輸出"])[0],
            "90",
            "x",
            "90",
            "1",
            (window.combo_column_options.get(5) or [""])[0],
            (window.combo_column_options.get(6) or [""])[0],
            (window.combo_column_options.get(7) or [""])[0],
            (window.combo_column_options.get(8) or [""])[0],
            (window.combo_column_options.get(9) or [""])[0],
            "1",
            "",
            "10",
            "",
            "1",
        ],
    ]
    window._populate_line_items_table_with_rows(calc_rows)

    subtotal_result = window.calculate_line_subtotals()
    row_results = []
    expected_row_values = [
        {"cbm": "12", "line_total": "120"},
        {"cbm": "56", "line_total": "1120"},
        {"cbm": "34", "line_total": "510"},
        {"cbm": "9", "line_total": "90"},
    ]
    for row_index, expected in enumerate(expected_row_values):
        actual = {
            "cbm": window._table_cell_text(row_index, 11),
            "line_total": window._table_cell_text(row_index, 13),
        }
        if actual != expected:
            raise AssertionError(f"subtotal mismatch at row {row_index + 1}: actual={actual} expected={expected}")
        row_results.append({"row": row_index + 1, **actual})

    try:
        window.calculate_document_totals()
    except Exception as exc:
        raise AssertionError(f"calculate_document_totals should tolerate blank/non-numeric extra material total: {exc}") from exc

    totals = {
        "production_amount": normalize_line_edit_value(getattr(window.ui, "le_productionAmount", None)) or "",
        "tax_amount": normalize_line_edit_value(getattr(window.ui, "le_taxAmount", None)) or "",
        "total_amount": normalize_line_edit_value(getattr(window.ui, "le_totalAmount", None)) or "",
    }
    expected_totals = {
        "production_amount": "1891",
        "tax_amount": "95",
        "total_amount": "1986",
    }
    if totals != expected_totals:
        raise AssertionError(f"document totals mismatch: actual={totals} expected={expected_totals}")

    blank_row_before = table.rowCount()
    last_row = table.rowCount() - 1
    widget = table.cellWidget(last_row, 1)
    if not isinstance(widget, QLineEdit):
        raise AssertionError("expected width cell to be editable line edit")
    widget.setText("10")
    app = QApplication.instance()
    if app is not None:
        app.processEvents()
    blank_row_after = table.rowCount()
    if blank_row_after != blank_row_before + 1:
        raise AssertionError(f"auto append broke during calc verify: before={blank_row_before} after={blank_row_after}")

    tab_sequence = _verify_table_navigation(window)
    return {
        "subtotal_result": subtotal_result,
        "row_results": row_results,
        "totals": totals,
        "auto_append_rows_before": blank_row_before,
        "auto_append_rows_after": blank_row_after,
        "tab_sequence": tab_sequence,
        "build_label": BUILD_LABEL_TEXT,
    }


def run_row_action_verification(window: GeneratedUiPreviewWindow, work_number: str) -> dict[str, object]:
    window._seed_demo_values()
    worknum_widget = getattr(window.ui, "le_worknum", None)
    if isinstance(worknum_widget, QLineEdit):
        worknum_widget.setText(work_number)

    baseline_work_order_id, baseline_line_count = window.save_work_order_with_lines()
    baseline_table_rows = window.ui.tbl_lineItems.rowCount()

    cleared = window.clear_line_item_row(1)
    if not cleared:
        raise AssertionError("failed to clear row 2")
    cleared_work_order_id, cleared_line_count = window.save_work_order_with_lines()
    if cleared_work_order_id != baseline_work_order_id:
        raise AssertionError("work_order_id changed after clear-row save")
    if cleared_line_count != baseline_line_count - 1:
        raise AssertionError(f"clear-row should reduce line count by 1: {cleared_line_count} vs {baseline_line_count}")

    reloaded_work_order_id, reloaded_line_count = window.load_work_order_by_number(work_number)
    if reloaded_work_order_id != baseline_work_order_id:
        raise AssertionError("work_order_id changed after clear-row reload")
    if reloaded_line_count != cleared_line_count:
        raise AssertionError("reloaded line count mismatch after clear-row flow")
    if len(_window_line_snapshot(window)) != cleared_line_count:
        raise AssertionError("cleared blank row was unexpectedly preserved in UI meaningful snapshot")

    window._seed_demo_values()
    if isinstance(worknum_widget, QLineEdit):
        worknum_widget.setText(work_number)
    window.save_work_order_with_lines()

    deleted_middle = window.delete_line_item_row(1)
    if not deleted_middle:
        raise AssertionError("failed to delete middle valid row")
    middle_tab_sequence = _verify_table_navigation(window)
    middle_work_order_id, middle_line_count = window.save_work_order_with_lines()
    if middle_work_order_id != baseline_work_order_id:
        raise AssertionError("work_order_id changed after middle delete")

    window.load_work_order_by_number(work_number)
    ui_lines_after_middle_delete = _window_line_snapshot(window)
    db_header, db_lines_after_middle_delete, db_meta = _db_bundle_snapshot(work_number)
    if ui_lines_after_middle_delete != db_lines_after_middle_delete:
        raise AssertionError("UI/DB mismatch after middle delete round-trip")
    if db_meta["duplicated_line_nos"] != 0:
        raise AssertionError(f"duplicate line_no after middle delete: {db_meta}")

    deleted_last = window.delete_line_item_row(len(ui_lines_after_middle_delete) - 1)
    if not deleted_last:
        raise AssertionError("failed to delete last valid row")
    final_tab_sequence = _verify_table_navigation(window)
    last_delete_work_order_id, last_delete_line_count = window.save_work_order_with_lines()
    if last_delete_work_order_id != baseline_work_order_id:
        raise AssertionError("work_order_id changed after last-line delete")

    window.load_work_order_by_number(work_number)
    table = window.ui.tbl_lineItems
    if window._line_row_has_meaningful_data(table.rowCount() - 1):
        raise AssertionError("last row should remain a blank preparatory row after delete+reload")
    final_ui_lines = _window_line_snapshot(window)
    _final_db_header, final_db_lines, final_db_meta = _db_bundle_snapshot(work_number)
    if final_ui_lines != final_db_lines:
        raise AssertionError("UI/DB mismatch after last valid row delete round-trip")
    if final_db_meta["duplicated_line_nos"] != 0:
        raise AssertionError(f"duplicate line_no after last delete: {final_db_meta}")

    return {
        "work_order_id": baseline_work_order_id,
        "baseline_line_count": baseline_line_count,
        "baseline_table_rows": baseline_table_rows,
        "after_clear_line_count": cleared_line_count,
        "after_middle_delete_line_count": middle_line_count,
        "after_last_delete_line_count": last_delete_line_count,
        "reloaded_line_count": len(final_ui_lines),
        "db_line_count": final_db_meta["line_count"],
        "header_customer_name": db_header["customer_name"],
        "middle_delete_tab_sequence": middle_tab_sequence,
        "final_tab_sequence": final_tab_sequence,
        "build_label": BUILD_LABEL_TEXT,
    }


def run_reset_verification(window: GeneratedUiPreviewWindow, work_number: str) -> dict[str, object]:
    window._initialize_blank_work_order()
    blank_row_count_before = window.ui.tbl_lineItems.rowCount()
    with _mock_message_boxes(question_response=QMessageBox.StandardButton.Yes) as blank_events:
        window._handle_reset_clicked()
    if blank_events["question"]:
        raise AssertionError("blank reset should not ask for confirmation")
    if window.ui.tbl_lineItems.rowCount() != DEFAULT_LINE_ITEM_ROW_COUNT:
        raise AssertionError("blank reset should keep 15 rows")

    window._seed_demo_values()
    with _mock_message_boxes(question_response=QMessageBox.StandardButton.No) as cancel_events:
        window._handle_reset_clicked()
    if len(cancel_events["question"]) != 1:
        raise AssertionError("filled reset should ask exactly one confirmation")
    if not normalize_line_edit_value(getattr(window.ui, "le_worknum", None)):
        raise AssertionError("cancelled reset should keep existing content")

    with _mock_message_boxes(question_response=QMessageBox.StandardButton.Yes) as confirm_events:
        window._handle_reset_clicked()
    if len(confirm_events["question"]) != 1:
        raise AssertionError("confirmed reset should ask exactly one confirmation")
    if window._has_loaded_content_on_screen():
        raise AssertionError("reset should clear the whole work order")
    if window.ui.tbl_lineItems.rowCount() != DEFAULT_LINE_ITEM_ROW_COUNT:
        raise AssertionError("reset should restore line table to 15 rows")

    window._seed_demo_values()
    worknum_widget = getattr(window.ui, "le_worknum", None)
    if isinstance(worknum_widget, QLineEdit):
        worknum_widget.setText(work_number)
    saved_work_order_id, saved_line_count = window.save_work_order_with_lines()
    window.reset_work_order_to_blank()
    if window.ui.tbl_lineItems.rowCount() != DEFAULT_LINE_ITEM_ROW_COUNT:
        raise AssertionError("direct reset should restore line table to 15 rows")
    window._seed_demo_values()
    if isinstance(worknum_widget, QLineEdit):
        worknum_widget.setText(work_number)
    resaved_work_order_id, resaved_line_count = window.save_work_order_with_lines()
    reloaded_work_order_id, reloaded_line_count = window.load_work_order_by_number(work_number)
    if saved_work_order_id != resaved_work_order_id or saved_work_order_id != reloaded_work_order_id:
        raise AssertionError("save/reset/re-save should keep same work_order_id")

    return {
        "blank_reset_row_count_before": blank_row_count_before,
        "blank_reset_row_count_after": window.ui.tbl_lineItems.rowCount(),
        "confirmations_on_cancel": len(cancel_events["question"]),
        "confirmations_on_confirm": len(confirm_events["question"]),
        "saved_work_order_id": saved_work_order_id,
        "saved_line_count": saved_line_count,
        "resaved_line_count": resaved_line_count,
        "reloaded_line_count": reloaded_line_count,
        "build_label": BUILD_LABEL_TEXT,
    }


def run_open_dirty_check_verification(window: GeneratedUiPreviewWindow, work_number: str) -> dict[str, object]:
    window._seed_demo_values()
    worknum_widget = getattr(window.ui, "le_worknum", None)
    case_widget = getattr(window.ui, "le_caseName", None)
    remark_widget = getattr(window.ui, "te_remark", None)
    if isinstance(worknum_widget, QLineEdit):
        worknum_widget.setText(work_number)
    seeded_work_order_id, seeded_line_count = window.save_work_order_with_lines()

    window._initialize_blank_work_order()
    if isinstance(worknum_widget, QLineEdit):
        worknum_widget.setText(work_number)
    with _mock_message_boxes(question_response=QMessageBox.StandardButton.Yes) as lookup_only_events:
        window._handle_open_clicked()
    if lookup_only_events["question"]:
        raise AssertionError("lookup-only open should not ask for overwrite confirmation")
    if len(lookup_only_events["information"]) != 1:
        raise AssertionError("lookup-only open should still finish the open flow")

    window.reset_work_order_to_blank()
    if isinstance(worknum_widget, QLineEdit):
        worknum_widget.setText(work_number)
    if isinstance(case_widget, QLineEdit):
        case_widget.setText("DIRTY CHECK SHOULD ASK")
    with _mock_message_boxes(question_response=QMessageBox.StandardButton.No) as typed_content_events:
        window._handle_open_clicked()
    if len(typed_content_events["question"]) != 1:
        raise AssertionError("open with edited content should ask exactly one confirmation")
    if typed_content_events["information"]:
        raise AssertionError("cancelled open should not show success information")

    window.load_work_order_by_number(work_number)
    if not window._loaded_work_order_number:
        raise AssertionError("loaded work order marker was not recorded")
    with _mock_message_boxes(question_response=QMessageBox.StandardButton.No) as loaded_screen_events:
        window._handle_open_clicked()
    if len(loaded_screen_events["question"]) != 1:
        raise AssertionError("loaded work order should still ask before opening another one")

    window.reset_work_order_to_blank()
    if isinstance(worknum_widget, QLineEdit):
        worknum_widget.setText(work_number)
    if remark_widget is not None:
        remark_widget.setPlainText("remark makes screen dirty")
    with _mock_message_boxes(question_response=QMessageBox.StandardButton.No) as remark_events:
        window._handle_open_clicked()
    if len(remark_events["question"]) != 1:
        raise AssertionError("remark content should trigger overwrite confirmation")

    return {
        "seeded_work_order_id": seeded_work_order_id,
        "seeded_line_count": seeded_line_count,
        "lookup_only_confirmations": len(lookup_only_events["question"]),
        "typed_content_confirmations": len(typed_content_events["question"]),
        "loaded_screen_confirmations": len(loaded_screen_events["question"]),
        "remark_confirmations": len(remark_events["question"]),
        "loaded_work_number_marker": window._loaded_work_order_number,
        "build_label": BUILD_LABEL_TEXT,
    }


def run_save_overwrite_verification(window: GeneratedUiPreviewWindow, work_number: str) -> dict[str, object]:
    window._seed_demo_values()
    worknum_widget = getattr(window.ui, "le_worknum", None)
    if isinstance(worknum_widget, QLineEdit):
        worknum_widget.setText(work_number)
    baseline_work_order_id, baseline_line_count = window.save_work_order_with_lines()
    baseline_header, baseline_lines, baseline_meta = _db_bundle_snapshot(work_number)

    case_widget = getattr(window.ui, "le_caseName", None)
    if isinstance(case_widget, QLineEdit):
        case_widget.setText("SAVE UX SHOULD NOT OVERWRITE WHEN CANCELLED")

    with _mock_message_boxes(question_response=QMessageBox.StandardButton.No) as cancel_events:
        window._handle_save_clicked()
    after_cancel_header, after_cancel_lines, after_cancel_meta = _db_bundle_snapshot(work_number)
    if len(cancel_events["question"]) != 1:
        raise AssertionError("existing work number save should ask before overwrite")
    if after_cancel_header != baseline_header or after_cancel_lines != baseline_lines or after_cancel_meta != baseline_meta:
        raise AssertionError("cancelled overwrite should not change DB contents")

    if isinstance(case_widget, QLineEdit):
        case_widget.setText("SAVE UX OVERWRITE CONFIRMED")
    with _mock_message_boxes(question_response=QMessageBox.StandardButton.Yes) as confirm_events:
        window._handle_save_clicked()
    after_confirm_header, _after_confirm_lines, after_confirm_meta = _db_bundle_snapshot(work_number)
    if len(confirm_events["question"]) != 1:
        raise AssertionError("confirmed overwrite should ask exactly one confirmation")
    if after_confirm_header["case_name"] != "SAVE UX OVERWRITE CONFIRMED":
        raise AssertionError("confirmed overwrite should update DB")
    if after_confirm_meta["work_order_id"] != baseline_work_order_id:
        raise AssertionError("overwrite save should update existing work_order row")

    return {
        "work_order_id": baseline_work_order_id,
        "baseline_line_count": baseline_line_count,
        "cancel_confirmations": len(cancel_events["question"]),
        "confirm_confirmations": len(confirm_events["question"]),
        "case_name_after_confirm": after_confirm_header["case_name"],
        "build_label": BUILD_LABEL_TEXT,
    }


def chunked(items: list[dict[str, str]], size: int) -> list[list[dict[str, str]]]:
    if size <= 0:
        raise ValueError("chunk size must be positive")
    return [items[index : index + size] for index in range(0, len(items), size)]


def sanitize_filename_component(value: str) -> str:
    sanitized = "".join("_" if ch in '<>:"/\\|?*' else ch for ch in value.strip())
    sanitized = sanitized.replace("\n", " ").replace("\r", " ")
    sanitized = " ".join(sanitized.split())
    return sanitized.strip(" .") or "未命名"


def page_label(page_index: int) -> str:
    if 1 <= page_index <= len(PAGE_LABELS):
        return PAGE_LABELS[page_index - 1]
    return f"第{page_index}頁"


def build_billing_output_filename(work_number: str, customer_name: str, case_name: str, page_index: int, total_pages: int) -> str:
    parts = [
        sanitize_filename_component(work_number),
        sanitize_filename_component(customer_name),
        sanitize_filename_component(case_name),
    ]
    if total_pages > 1:
        parts.append(page_label(page_index))
    return "-".join(parts) + ".xlsx"


def set_merged_cell_value(sheet, cell_ref: str, value: str) -> None:
    sheet[cell_ref] = value


def set_cell_alignment(sheet, cell_ref: str, *, horizontal: str = "center", vertical: str = "center") -> None:
    if Alignment is None:
        return
    sheet[cell_ref].alignment = Alignment(horizontal=horizontal, vertical=vertical)


def set_cell_font(sheet, cell_ref: str, *, name: str = INPUT_FONT_FAMILY, size: int = 22) -> None:
    if Font is None:
        return
    sheet[cell_ref].font = Font(name=name, size=size)


def apply_billing_header_to_sheet(sheet, header: dict[str, str]) -> None:
    set_merged_cell_value(sheet, "C3", header["customer_full_name"])
    sheet["G3"] = header["case_name"]
    sheet["I3"] = header["contact_name"]
    sheet["K3"] = header["phone"]
    set_merged_cell_value(sheet, "C4", header["work_address"])
    sheet["I4"] = header["work_time"]
    sheet["K4"] = header["cleanup_time"]
    sheet["C5"] = header["company_tax_id"]
    set_merged_cell_value(sheet, "F5", header["company_address"])

    for cell_ref in ("C3", "G3", "I3", "K3", "C4", "I4", "K4", "C5", "F5"):
        set_cell_alignment(sheet, cell_ref, horizontal="left", vertical="center")
        set_cell_font(sheet, cell_ref)


def apply_billing_rows_to_sheet(sheet, page_rows: list[dict[str, str]]) -> None:
    for row_offset in range(BILLING_MAX_ROWS_PER_PAGE):
        excel_row = BILLING_DETAIL_START_ROW + row_offset
        row_payload = page_rows[row_offset] if row_offset < len(page_rows) else None
        values = row_payload or {
            "production_item": "", "size": "", "quantity": "", "material": "", "board": "",
            "extra_material": "", "extra_material_quantity": "", "cbm": "", "line_total": "", "extra_material_total": "",
        }
        sheet[f"B{excel_row}"] = values["production_item"]
        set_merged_cell_value(sheet, f"C{excel_row}", values["size"])
        sheet[f"E{excel_row}"] = values["quantity"]
        sheet[f"F{excel_row}"] = values["material"]
        sheet[f"G{excel_row}"] = values["board"]
        sheet[f"H{excel_row}"] = values["extra_material"]
        sheet[f"I{excel_row}"] = values["extra_material_quantity"]
        sheet[f"J{excel_row}"] = values["cbm"]
        sheet[f"K{excel_row}"] = values["line_total"]
        sheet[f"L{excel_row}"] = values["extra_material_total"]
        for cell_ref in (f"B{excel_row}", f"C{excel_row}", f"E{excel_row}", f"F{excel_row}", f"G{excel_row}", f"H{excel_row}", f"I{excel_row}", f"J{excel_row}", f"K{excel_row}", f"L{excel_row}"):
            set_cell_font(sheet, cell_ref)


def apply_billing_totals_to_sheet(sheet, totals: dict[str, str]) -> None:
    sheet["H22"] = totals["production_amount"]
    sheet["J22"] = totals["tax_amount"]
    sheet["L22"] = totals["total_amount"]
    for cell_ref in ("H22", "J22", "L22"):
        set_cell_font(sheet, cell_ref)
        set_cell_alignment(sheet, cell_ref, horizontal="center", vertical="center")


def clear_billing_totals_on_sheet(sheet) -> None:
    for cell_ref in ("H22", "J22", "L22"):
        sheet[cell_ref] = ""


def _seed_billing_rows(window: GeneratedUiPreviewWindow, row_count: int, *, fallback_row: int | None = None) -> None:
    rows: list[list[str]] = []
    for idx in range(row_count):
        use_fallback = fallback_row is not None and idx == fallback_row
        rows.append([
            f"製作項目{idx + 1}" if not use_fallback else "客製項目/特殊*測試",
            str(100 + idx),
            "x",
            str(200 + idx),
            str((idx % 3) + 1),
            f"材質{idx + 1}" if not use_fallback else "客製材質?",
            "",
            f"板材{idx + 1}" if not use_fallback else "特殊板",
            f"{(idx % 5) + 1}mm" if not use_fallback else "7mm",
            f"備料{idx + 1}" if not use_fallback else "其他備料:",
            str((idx % 2) + 1),
            str(10 + idx),
            str(100 + idx),
            str(500 + idx),
            str(30 + idx),
        ])
    window._populate_line_items_table_with_rows(rows)


def run_billing_export_verification(window: GeneratedUiPreviewWindow, output_dir: Path) -> dict[str, object]:
    customer_name = next(iter(window.client_rows_by_short_name.keys()), "")
    if not customer_name:
        raise AssertionError("clients table is empty; cannot verify billing export")
    client_row = window.client_rows_by_short_name[customer_name]
    customer_combo = getattr(window.ui, "cb_customerName", None)
    if isinstance(customer_combo, QComboBox):
        customer_combo.setCurrentText(customer_name)
    for attr, value in {
        "le_worknum": "BILL-EXPORT-01",
        "le_caseName": "請款測試/案件:特殊?",
        "le_contactName": "王小明",
        "le_phone": "02-1234-5678",
        "le_startTime": "2026/04/27 09:00",
        "le_endTime": "2026/04/27 18:00",
        "lle_address": "台北市信義區測試路 1 號",
    }.items():
        widget = getattr(window.ui, attr, None)
        if isinstance(widget, QLineEdit):
            widget.setText(value)

    output_dir.mkdir(parents=True, exist_ok=True)

    _seed_billing_rows(window, 15, fallback_row=2)
    window._set_summary_field_value("le_productionAmount", Decimal("9999"))
    window._set_summary_field_value("le_taxAmount", Decimal("500"))
    window._set_summary_field_value("le_totalAmount", Decimal("10499"))
    single_paths = window.export_billing_excels(output_dir / "single")
    if len(single_paths) != 1:
        raise AssertionError(f"expected 1 billing file, got {len(single_paths)}")
    wb = load_workbook(single_paths[0], data_only=True)
    ws = wb.active
    expected_full_name = str(client_row.get("full_name") or customer_name)
    if ws["C3"].value != expected_full_name:
        raise AssertionError(f"client full_name mismatch: {ws['C3'].value} vs {expected_full_name}")
    if ws["C4"].value != "台北市信義區測試路 1 號":
        raise AssertionError("work address mismatch")
    if ws["C5"].value != str(client_row.get("tax_id") or ""):
        raise AssertionError("client tax_id mismatch")
    if ws["F5"].value != str(client_row.get("address") or ""):
        raise AssertionError("client address mismatch")
    if ws["B9"].value != "客製項目/特殊*測試":
        raise AssertionError("fallback production_item export mismatch")
    if ws["F9"].value != "客製材質?":
        raise AssertionError("fallback material export mismatch")
    if ws["G9"].value != "7mm特殊板":
        raise AssertionError("fallback board export mismatch")
    if ws["H22"].value in (None, "") or ws["J22"].value in (None, "") or ws["L22"].value in (None, ""):
        raise AssertionError("single-page totals should be filled")

    if single_paths[0].name != "BILL-EXPORT-01-{}-{}-第一頁.xlsx".format(
        sanitize_filename_component(customer_name),
        sanitize_filename_component("請款測試/案件:特殊?"),
    ):
        raise AssertionError(f"unexpected single file name: {single_paths[0].name}")

    _seed_billing_rows(window, 17, fallback_row=16)
    multi_paths = window.export_billing_excels(output_dir / "multi")
    if len(multi_paths) != 2:
        raise AssertionError(f"expected 2 billing files, got {len(multi_paths)}")
    wb1 = load_workbook(multi_paths[0], data_only=True)
    ws1 = wb1.active
    wb2 = load_workbook(multi_paths[1], data_only=True)
    ws2 = wb2.active
    if ws1["B7"].value != "製作項目1" or ws1["B21"].value != "製作項目15":
        raise AssertionError("first page detail rows mismatch")
    if ws2["B7"].value != "製作項目16" or ws2["B8"].value != "客製項目/特殊*測試":
        raise AssertionError("second page remaining rows mismatch")
    if ws2["B9"].value not in (None, ""):
        raise AssertionError("second page should only contain remaining rows")
    if ws1["H22"].value not in (None, "") or ws1["J22"].value not in (None, "") or ws1["L22"].value not in (None, ""):
        raise AssertionError("non-final page totals must be blank")
    if ws2["H22"].value in (None, "") or ws2["J22"].value in (None, "") or ws2["L22"].value in (None, ""):
        raise AssertionError("final page totals must be filled")

    return {
        "single_paths": [str(path) for path in single_paths],
        "multi_paths": [str(path) for path in multi_paths],
        "single_header": {
            "customer_full_name": ws["C3"].value,
            "work_address": ws["C4"].value,
            "tax_id": ws["C5"].value,
            "company_address": ws["F5"].value,
        },
        "single_fallback_row": {
            "production_item": ws["B9"].value,
            "material": ws["F9"].value,
            "board": ws["G9"].value,
        },
        "multi_page_rows": {
            "page1_last": ws1["B21"].value,
            "page2_first": ws2["B7"].value,
            "page2_second": ws2["B8"].value,
        },
        "build_label": BUILD_LABEL_TEXT,
    }


def parse_args(argv: list[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--screenshot", type=Path, help="Save a screenshot to this path and exit.")
    parser.add_argument("--save-demo", action="store_true", help="Save the seeded header + line fields into work_orders/work_order_lines and print the inserted id.")
    parser.add_argument("--load-work-number", help="Load this work number after window init, print the loaded id/count, and exit.")
    parser.add_argument("--roundtrip-verify", help="Load an existing work number, edit header/lines, save, reload, verify DB/UI consistency, and exit.")
    parser.add_argument("--row-ux-verify", help="Seed demo data into this work number, verify clear/delete row UX, save/reload, and exit.")
    parser.add_argument("--calc-verify", action="store_true", help="Verify subtotal/calculate rules, blank handling, auto-append, and tab order.")
    parser.add_argument("--reset-verify", help="Seed demo data into this work number, verify reset UX/save-open flow, and exit.")
    parser.add_argument("--save-overwrite-verify", help="Seed demo data into this work number, verify save overwrite confirmation UX, and exit.")
    parser.add_argument("--open-dirty-verify", help="Seed demo data into this work number, verify open dirty-check UX, and exit.")
    parser.add_argument("--billing-export-verify", type=Path, help="Export billing excels into this directory, verify 1-page/2-page mappings, and exit.")
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv or sys.argv[1:])

    app = QApplication(sys.argv)
    apply_light_preview_theme(app)
    window = GeneratedUiPreviewWindow()
    window.show()

    if args.save_demo:
        def save_demo_and_quit() -> None:
            window._seed_demo_values()
            try:
                work_order_id, line_count = window.save_work_order_with_lines()
            except Exception as exc:
                print(f"SAVE_DEMO_ERROR: {exc}", file=sys.stderr)
                app.exit(1)
                return

            print(f"SAVE_DEMO_OK:{work_order_id}:{line_count}")
            app.exit(0)

        QTimer.singleShot(0, save_demo_and_quit)
    elif args.load_work_number:
        def load_and_quit() -> None:
            try:
                work_order_id, line_count = window.load_work_order_by_number(args.load_work_number)
            except Exception as exc:
                print(f"LOAD_WORK_ORDER_ERROR: {exc}", file=sys.stderr)
                app.exit(1)
                return

            header_values = _window_header_snapshot(window)
            first_rows = _window_line_snapshot(window)[:3]
            print(f"LOAD_WORK_ORDER_OK:{work_order_id}:{line_count}:{header_values}:{first_rows}")
            app.exit(0)

        QTimer.singleShot(0, load_and_quit)
    elif args.roundtrip_verify:
        def verify_and_quit() -> None:
            try:
                result = run_roundtrip_verification(window, args.roundtrip_verify)
            except Exception as exc:
                print(f"ROUNDTRIP_VERIFY_ERROR: {exc}", file=sys.stderr)
                app.exit(1)
                return

            print(f"ROUNDTRIP_VERIFY_OK:{result}")
            app.exit(0)

        QTimer.singleShot(0, verify_and_quit)
    elif args.row_ux_verify:
        def verify_row_ux_and_quit() -> None:
            try:
                result = run_row_action_verification(window, args.row_ux_verify)
            except Exception as exc:
                print(f"ROW_UX_VERIFY_ERROR: {exc}", file=sys.stderr)
                app.exit(1)
                return

            print(f"ROW_UX_VERIFY_OK:{result}")
            app.exit(0)

        QTimer.singleShot(0, verify_row_ux_and_quit)
    elif args.calc_verify:
        def verify_calc_and_quit() -> None:
            try:
                result = run_calculation_verification(window)
            except Exception as exc:
                print(f"CALC_VERIFY_ERROR: {exc}", file=sys.stderr)
                app.exit(1)
                return

            print(f"CALC_VERIFY_OK:{result}")
            app.exit(0)

        QTimer.singleShot(0, verify_calc_and_quit)
    elif args.reset_verify:
        def verify_reset_and_quit() -> None:
            try:
                result = run_reset_verification(window, args.reset_verify)
            except Exception as exc:
                print(f"RESET_VERIFY_ERROR: {exc}", file=sys.stderr)
                app.exit(1)
                return

            print(f"RESET_VERIFY_OK:{result}")
            app.exit(0)

        QTimer.singleShot(0, verify_reset_and_quit)
    elif args.save_overwrite_verify:
        def verify_save_overwrite_and_quit() -> None:
            try:
                result = run_save_overwrite_verification(window, args.save_overwrite_verify)
            except Exception as exc:
                print(f"SAVE_OVERWRITE_VERIFY_ERROR: {exc}", file=sys.stderr)
                app.exit(1)
                return

            print(f"SAVE_OVERWRITE_VERIFY_OK:{result}")
            app.exit(0)

        QTimer.singleShot(0, verify_save_overwrite_and_quit)
    elif args.open_dirty_verify:
        def verify_open_dirty_and_quit() -> None:
            try:
                result = run_open_dirty_check_verification(window, args.open_dirty_verify)
            except Exception as exc:
                print(f"OPEN_DIRTY_VERIFY_ERROR: {exc}", file=sys.stderr)
                app.exit(1)
                return

            print(f"OPEN_DIRTY_VERIFY_OK:{result}")
            app.exit(0)

        QTimer.singleShot(0, verify_open_dirty_and_quit)
    elif args.billing_export_verify:
        def verify_billing_export_and_quit() -> None:
            try:
                result = run_billing_export_verification(window, args.billing_export_verify)
            except Exception as exc:
                print(f"BILLING_EXPORT_VERIFY_ERROR: {exc}", file=sys.stderr)
                app.exit(1)
                return

            print(f"BILLING_EXPORT_VERIFY_OK:{result}")
            app.exit(0)

        QTimer.singleShot(0, verify_billing_export_and_quit)
    elif args.screenshot:
        target = args.screenshot.expanduser().resolve()
        target.parent.mkdir(parents=True, exist_ok=True)

        def save_and_quit() -> None:
            window.repaint()
            app.processEvents()
            window.grab().save(str(target))
            app.quit()

        QTimer.singleShot(250, save_and_quit)

    return app.exec()


if __name__ == "__main__":
    raise SystemExit(main())
